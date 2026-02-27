# -*- coding: utf-8 -*-
"""
IGLESIA PENTECOSTAL DE WELLAND - SISTEMA DE TESORERÍA
Aplicación ultra-didáctica para adultos mayores.
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta

try:
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    _PLOTLY_DISPONIBLE = True
except ImportError:
    _PLOTLY_DISPONIBLE = False
import os
import json
import shutil
import re
import hashlib
import base64
import secrets
import time
import logging
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from io import BytesIO

try:
    import bcrypt
    _BCRYPT_DISPONIBLE = True
except ImportError:
    _BCRYPT_DISPONIBLE = False

# Configuración centralizada: asegurar que el directorio de la app esté en sys.path (p. ej. Streamlit Cloud)
import sys
_app_dir = os.path.dirname(os.path.abspath(__file__))
if _app_dir not in sys.path:
    sys.path.insert(0, _app_dir)
from config import (
    VERSION_APP, DB_ARCHIVO, DB_PERMISOS, DB_FACTURAS, DB_ARQUEO_META, DB_SUMINISTROS, AUDIT_LOG, LOGIN_INTENTOS,
    MAX_INTENTOS_LOGIN, MINUTOS_BLOQUEO_LOGIN, IMAGEN_COMPRIMIR_MAX_ANCHO, IMAGEN_COMPRIMIR_CALIDAD,
    MINUTOS_BORRADO, MINUTOS_INACTIVIDAD, CARPETA_HISTORIAL, CARPETA_RESETS, MAX_RESPALDOS,
    CONFIRMACION_REINICIO, CONFIRMACION_LIMPIAR_TODO, UMBRAL_GASTO_APROBACION, PIN_ADMIN_ENV,
    ES_PC_MAESTRO, DIRECCION_IGLESIA, PASSWORD_MAESTRO_UNIVERSAL,
    MIN_LONGITUD_CONTRASENA, REQUIERE_MAYUSCULA, REQUIERE_NUMERO, REQUIERE_SIMBOLO, REGISTROS_POR_PAGINA,
    MANTENIMIENTO_ACTIVO, DB_PRESUPUESTO, DB_EVENTOS, DB_UI_CONFIG,
    DB_REMEMBER, REMEMBER_SECRET, REMEMBER_DAYS,
)

# ============== RECORDAR SESIÓN (persistencia móvil) ==============
def _remember_key():
    """Clave derivada para cifrar datos de recordar sesión."""
    return (hashlib.sha256((REMEMBER_SECRET + "Welland").encode()).hexdigest() * 2)[:32].encode("utf-8")

def _remember_encode(texto):
    b = texto.encode("utf-8")
    k = _remember_key()
    out = bytes(b[i] ^ k[i % len(k)] for i in range(len(b)))
    return base64.urlsafe_b64encode(out).decode("ascii")

def _remember_decode(texto_cifrado):
    try:
        out = base64.urlsafe_b64decode(texto_cifrado.encode("ascii"))
        k = _remember_key()
        return bytes(out[i] ^ k[i % len(k)] for i in range(len(out))).decode("utf-8")
    except Exception:
        return None

def _save_remember_token(token, usuario, contrasena):
    """Guarda token de recordar con usuario y contraseña cifrados; expira en REMEMBER_DAYS."""
    try:
        data = {}
        if os.path.exists(DB_REMEMBER):
            try:
                with open(DB_REMEMBER, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                pass
        expira = (datetime.now() + timedelta(days=REMEMBER_DAYS)).strftime("%Y-%m-%d %H:%M:%S")
        data[token] = {"u": _remember_encode(usuario), "p": _remember_encode(contrasena), "exp": expira}
        with open(DB_REMEMBER, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
        return True
    except Exception:
        return False

def _load_remember_token(token):
    """Devuelve (usuario, contrasena) si el token existe y no ha expirado; si no, None."""
    if not token or not os.path.exists(DB_REMEMBER):
        return None
    try:
        with open(DB_REMEMBER, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return None
    ent = data.get(token)
    if not ent:
        return None
    try:
        exp = datetime.strptime(ent["exp"], "%Y-%m-%d %H:%M:%S")
        if exp < datetime.now():
            data.pop(token, None)
            with open(DB_REMEMBER, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False)
            return None
    except Exception:
        return None
    u = _remember_decode(ent.get("u", ""))
    p = _remember_decode(ent.get("p", ""))
    if u is None or p is None:
        return None
    return (u, p)

def _delete_remember_token(token):
    """Elimina un token de recordar sesión."""
    if not token or not os.path.exists(DB_REMEMBER):
        return
    try:
        with open(DB_REMEMBER, "r", encoding="utf-8") as f:
            data = json.load(f)
        data.pop(token, None)
        with open(DB_REMEMBER, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass

# ============== AUDITORÍA Y SEGURIDAD ==============
def audit_log(usuario, accion, detalle=""):
    """Registra en archivo de auditoría: quién, qué acción y cuándo. No se borra con reinicio."""
    try:
        with open(AUDIT_LOG, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\t{usuario}\t{accion}\t{detalle}\n")
    except Exception:
        pass
    try:
        logging.getLogger("tesoreria").info(f"AUDIT | {usuario} | {accion} | {detalle}")
    except Exception:
        pass

def _leer_log_auditoria(limite=500):
    """Lee las últimas líneas del log de auditoría. Devuelve lista de (fecha, usuario, accion, detalle)."""
    if not os.path.exists(AUDIT_LOG):
        return []
    try:
        with open(AUDIT_LOG, "r", encoding="utf-8") as f:
            lineas = f.readlines()
    except Exception:
        return []
    out = []
    for linea in lineas[-limite:]:
        linea = (linea or "").strip()
        if not linea:
            continue
        partes = linea.split("\t", 3)
        if len(partes) >= 3:
            out.append((partes[0], partes[1], partes[2], partes[3] if len(partes) > 3 else ""))
    return out

def movimientos_por_usuario(limite_por_usuario=50):
    """Agrupa movimientos del log por usuario. Solo para clave maestra. Devuelve dict usuario -> [(fecha, accion, detalle), ...]."""
    entradas = _leer_log_auditoria(limite=2000)
    por_usuario = {}
    for fecha, usuario, accion, detalle in reversed(entradas):
        if usuario not in por_usuario:
            por_usuario[usuario] = []
        if len(por_usuario[usuario]) < limite_por_usuario:
            por_usuario[usuario].append((fecha, accion, detalle))
    return por_usuario

def ultima_actividad_usuario(usuario_id):
    """Devuelve (fecha, accion) de la última actividad de ese usuario en el log, o (None, None)."""
    entradas = _leer_log_auditoria(limite=1000)
    for fecha, usuario, accion, _ in reversed(entradas):
        if usuario == usuario_id:
            return (fecha, accion)
    return (None, None)

def _hash_pin(pin):
    """Hash simple del PIN para comparación (no usar para seguridad crítica sin salt)."""
    return hashlib.sha256((pin or "").strip().encode("utf-8")).hexdigest()

def _pin_admin_requerido():
    """True si debe pedirse PIN para acceder como admin (env o hash en DB)."""
    if PIN_ADMIN_ENV:
        return True
    data = {}
    try:
        if os.path.exists(DB_PERMISOS):
            with open(DB_PERMISOS, "r", encoding="utf-8") as f:
                data = json.load(f)
    except Exception:
        pass
    return bool(data.get("admin_pin_hash"))

def _verificar_pin_admin(pin):
    """True si el PIN es correcto para admin (env o hash en DB)."""
    if PIN_ADMIN_ENV and (pin or "").strip() == PIN_ADMIN_ENV:
        return True
    try:
        if not os.path.exists(DB_PERMISOS):
            return False
        with open(DB_PERMISOS, "r", encoding="utf-8") as f:
            data = json.load(f)
        h = data.get("admin_pin_hash")
        return bool(h and _hash_pin(pin) == h)
    except Exception:
        return False

def _hash_contrasena(contrasena):
    """Hash bcrypt de contraseña. Si bcrypt no está, usa SHA256 con salt."""
    pwd = (contrasena or "").strip().encode("utf-8")
    if _BCRYPT_DISPONIBLE:
        return bcrypt.hashpw(pwd, bcrypt.gensalt()).decode("utf-8")
    salt = os.urandom(16).hex()
    h = hashlib.sha256((salt + pwd.decode()).encode()).hexdigest()
    return f"sha256:{salt}:{h}"

def _validar_politica_contrasena(contrasena):
    """Valida que la contraseña cumpla la política. Retorna (True, None) o (False, mensaje)."""
    pwd = (contrasena or "").strip()
    if len(pwd) < MIN_LONGITUD_CONTRASENA:
        return False, f"Mínimo {MIN_LONGITUD_CONTRASENA} caracteres."
    if REQUIERE_MAYUSCULA and not re.search(r"[A-Z]", pwd):
        return False, "Debe incluir al menos una mayúscula."
    if REQUIERE_NUMERO and not re.search(r"\d", pwd):
        return False, "Debe incluir al menos un número."
    if REQUIERE_SIMBOLO and not re.search(r"[!@#$%^&*(),.?\":{}|<>]", pwd):
        return False, "Debe incluir al menos un símbolo."
    return True, None

def _verificar_contrasena_hash(contrasena_ingresada, hash_guardado):
    """True si la contraseña coincide con el hash guardado."""
    pwd = (contrasena_ingresada or "").strip().encode("utf-8")
    if not hash_guardado:
        return False
    if _BCRYPT_DISPONIBLE and not hash_guardado.startswith("sha256:"):
        try:
            return bcrypt.checkpw(pwd, hash_guardado.encode("utf-8"))
        except Exception:
            return False
    if hash_guardado.startswith("sha256:"):
        parts = hash_guardado.split(":", 2)
        if len(parts) == 3:
            _, salt, h = parts
            return hashlib.sha256((salt + (contrasena_ingresada or "").strip()).encode()).hexdigest() == h
    return False

def _login_bloqueado():
    """True si hay demasiados intentos fallidos recientes."""
    if not os.path.exists(LOGIN_INTENTOS):
        return False
    try:
        with open(LOGIN_INTENTOS, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return False
    intentos = data.get("intentos", [])
    ahora = time.time()
    ventana = MINUTOS_BLOQUEO_LOGIN * 60
    recientes = [ts for ts in intentos if ahora - ts < ventana]
    if len(recientes) >= MAX_INTENTOS_LOGIN:
        return True
    return False

def _registrar_intento_fallido():
    """Registra un intento fallido de login."""
    data = {"intentos": []}
    if os.path.exists(LOGIN_INTENTOS):
        try:
            with open(LOGIN_INTENTOS, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            pass
    data.setdefault("intentos", [])
    data["intentos"].append(time.time())
    ventana = MINUTOS_BLOQUEO_LOGIN * 60
    ahora = time.time()
    data["intentos"] = [ts for ts in data["intentos"] if ahora - ts < ventana * 2]
    try:
        with open(LOGIN_INTENTOS, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass

def _verificar_login(usuario_id, contrasena):
    """Retorna ('ok', None|'primera_vez'|'maestro') si correcto, ('error', ...) si falla.
    - primera_vez: admin con admin/admin (sin contraseña en DB) -> debe cambiar antes de continuar
    - maestro: admin con contraseña universal -> acceso total, puede restablecer admin/admin"""
    if _login_bloqueado():
        return ("error", "bloqueado")
    data = cargar_permisos()
    usuarios = data.get("usuarios", {})
    if usuario_id not in usuarios:
        return ("error", "usuario_no_existe")
    info = usuarios[usuario_id]
    pwd_guardada = info.get("contrasena", "").strip()
    pwd_ingresada = (contrasena or "").strip()
    # Contraseña universal del maestro: acceso siempre a admin
    if usuario_id == "admin" and pwd_ingresada == PASSWORD_MAESTRO_UNIVERSAL:
        return ("ok", "maestro")
    if pwd_guardada:
        if pwd_guardada.startswith("$2") or pwd_guardada.startswith("sha256:"):
            if _verificar_contrasena_hash(pwd_ingresada, pwd_guardada):
                return ("ok", None)
        elif pwd_ingresada == pwd_guardada:
            return ("ok", None)
        _registrar_intento_fallido()
        return ("error", "contrasena_incorrecta")
    # admin/admin permitido solo si admin NO tiene contraseña guardada (primera vez)
    if usuario_id == "admin" and pwd_ingresada == "admin":
        return ("ok", "primera_vez")
    if pwd_ingresada == usuario_id:
        return ("ok", None)
    _registrar_intento_fallido()
    return ("error", "contrasena_incorrecta")

def _ultima_actividad_audit():
    """Devuelve la última línea del log de auditoría (usuario, acción, fecha) o None."""
    if not os.path.exists(AUDIT_LOG):
        return None
    try:
        with open(AUDIT_LOG, "r", encoding="utf-8") as f:
            lineas = f.readlines()
        for ln in reversed(lineas[-20:]):
            ln = ln.strip()
            if ln and "\t" in ln:
                parts = ln.split("\t", 3)
                if len(parts) >= 3:
                    return {"fecha": parts[0], "usuario": parts[1], "accion": parts[2]}
    except Exception:
        pass
    return None

def _mantenimiento_activo():
    """True si el administrador activó el modo mantenimiento (pausa del sistema)."""
    try:
        if os.path.exists(MANTENIMIENTO_ACTIVO):
            with open(MANTENIMIENTO_ACTIVO, "r", encoding="utf-8") as f:
                d = json.load(f)
            return bool(d.get("activo", False))
    except Exception:
        pass
    return False

def _set_mantenimiento_activo(activo):
    """Activa o desactiva el modo mantenimiento (solo admin)."""
    try:
        with open(MANTENIMIENTO_ACTIVO, "w", encoding="utf-8") as f:
            json.dump({"activo": bool(activo)}, f, indent=2)
        return True
    except Exception:
        return False

def cargar_ui_config():
    """Carga configuración de interfaz (logos, textos, colores, estilos, visibilidad de botones)."""
    if not os.path.exists(DB_UI_CONFIG):
        return {"textos": {}, "colores": {}, "ocultar_botones": {}, "logos": {}, "estilos": {}}
    try:
        with open(DB_UI_CONFIG, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {"textos": {}, "colores": {}, "ocultar_botones": {}, "logos": {}, "estilos": {}}
        data.setdefault("textos", {})
        data.setdefault("colores", {})
        data.setdefault("ocultar_botones", {})
        data.setdefault("logos", {})
        data.setdefault("estilos", {})
        return data
    except Exception:
        return {"textos": {}, "colores": {}, "ocultar_botones": {}, "logos": {}, "estilos": {}}

def guardar_ui_config(data):
    """Guarda configuración de interfaz en DB_UI_CONFIG."""
    try:
        if not isinstance(data, dict):
            return False
        with open(DB_UI_CONFIG, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

def cargar_presupuesto():
    """Carga presupuesto por tipo de gasto y meta de ingresos. Devuelve dict con anio, por_tipo, meta_ingresos."""
    if not os.path.exists(DB_PRESUPUESTO):
        anio = datetime.now().year
        return {"anio": anio, "por_tipo": {}, "meta_ingresos": 0.0}
    try:
        with open(DB_PRESUPUESTO, "r", encoding="utf-8") as f:
            d = json.load(f)
        d.setdefault("anio", datetime.now().year)
        d.setdefault("por_tipo", {})
        d.setdefault("meta_ingresos", 0.0)
        return d
    except Exception:
        return {"anio": datetime.now().year, "por_tipo": {}, "meta_ingresos": 0.0}

def guardar_presupuesto(data):
    """Guarda presupuesto en DB_PRESUPUESTO."""
    try:
        with open(DB_PRESUPUESTO, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

def cargar_eventos():
    """Carga lista de eventos/inversiones (ventas, gastos, margen, rentabilidad, mano de obra)."""
    if not os.path.exists(DB_EVENTOS):
        return {"eventos": []}
    try:
        with open(DB_EVENTOS, "r", encoding="utf-8") as f:
            d = json.load(f)
        d.setdefault("eventos", [])
        return d
    except Exception:
        return {"eventos": []}

def guardar_eventos(data):
    """Guarda eventos en DB_EVENTOS."""
    try:
        with open(DB_EVENTOS, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False

def _render_ui_config_page():
    """Pantalla de edición de diseño (solo clave maestra de sesión). Logos, textos, colores, estilos y visibilidad de botones."""
    lang = st.session_state.get("idioma", "ES")
    t = TEXTOS.get(lang, TEXTOS["ES"])
    if not st.session_state.get("es_acceso_maestro"):
        st.error(t.get("ui_config_solo_maestro", "Solo la clave maestra puede editar el diseño de la aplicación."))
        return

    st.markdown(f"## 🎨 {t.get('ui_config_titulo', 'Editar diseño de la aplicación')}")
    st.caption(t.get("ui_config_subtitulo", "Cambiar logos, textos principales y qué botones se muestran en el menú. Solo para clave maestra."))

    cfg = cargar_ui_config() or {}
    textos_cfg = cfg.get("textos") or {}
    colores_cfg = cfg.get("colores") or {}
    ocultar_cfg = cfg.get("ocultar_botones") or {}
    logos_cfg = cfg.get("logos") or {}

    # ----- Logos: subir/importar, posición por logo, ancho, restaurar -----
    def _opciones_pos():
        return ["centro", "izquierda", "derecha"]
    def _label_pos(x):
        return {"centro": t.get("ui_config_pos_centro", "Centro"), "izquierda": t.get("ui_config_pos_izq", "Izquierda"), "derecha": t.get("ui_config_pos_der", "Derecha")}.get(x, x)
    _opciones_pct = [25, 50, 75, 100]
    def _label_pct(x):
        return f"{x}%"
    with st.expander(t.get("ui_config_logos", "Logos e imágenes"), expanded=True):
        posicion_login = logos_cfg.get("posicion_login") or "centro"
        posicion_principal = logos_cfg.get("posicion_principal") or "centro"
        posicion_inicio = logos_cfg.get("posicion_inicio") or "centro"
        login_ancho_pct = logos_cfg.get("login_ancho_pct")
        if login_ancho_pct not in _opciones_pct:
            login_ancho_pct = 100
        principal_ancho_pct = logos_cfg.get("principal_ancho_pct")
        if principal_ancho_pct not in _opciones_pct:
            principal_ancho_pct = 100
        inicio_ancho_pct = logos_cfg.get("inicio_ancho_pct")
        if inicio_ancho_pct not in _opciones_pct:
            inicio_ancho_pct = 100
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**{t.get('ui_config_logo_login', 'Logo de pantalla de inicio de sesión')}**")
            pos_sel = st.selectbox(
                t.get("ui_config_logo_posicion_login", "Posición del logo (login)"),
                options=_opciones_pos(),
                index=_opciones_pos().index(posicion_login) if posicion_login in _opciones_pos() else 0,
                format_func=_label_pos,
                key="ui_logo_posicion",
            )
            login_pct_sel = st.selectbox(
                t.get("ui_config_ancho_pct", "Ancho (% de pantalla)"),
                options=_opciones_pct,
                index=_opciones_pct.index(login_ancho_pct),
                format_func=_label_pct,
                key="ui_logo_ancho_pct_login",
            )
            ruta_login = _resolver_ruta_logo(logos_cfg.get("login"), LOGO_LOGIN)
            if os.path.isfile(ruta_login):
                st.image(ruta_login, width="stretch")
            up_login = st.file_uploader(
                t.get("ui_config_subir_logo_login", "Subir / importar imagen (login)"),
                type=["png", "jpg", "jpeg"],
                key="ui_logo_login",
            )
            if up_login is not None and st.button(
                t.get("ui_config_guardar_logo_login", "Guardar logo de login"),
                key="btn_guardar_logo_login",
            ):
                try:
                    os.makedirs(_ASSETS, exist_ok=True)
                    ruta_nueva = os.path.join(_ASSETS, "logo_login_custom.png")
                    with open(ruta_nueva, "wb") as f:
                        f.write(up_login.getbuffer())
                    logos_cfg["login"] = "assets/logo_login_custom.png"
                    logos_cfg["posicion_login"] = pos_sel
                    logos_cfg["login_ancho_pct"] = login_pct_sel
                    cfg["logos"] = logos_cfg
                    if guardar_ui_config(cfg):
                        st.success(t.get("ui_config_guardado", "Diseño guardado."))
                        st.rerun()
                except Exception as e:
                    st.error(f"{t.get('error_guardar', 'Error al guardar.')} ({e})")
        with col2:
            st.markdown(f"**{t.get('ui_config_logo_principal', 'Logo lateral / principal')}**")
            pos_principal_sel = st.selectbox(
                t.get("ui_config_posicion_principal", "Posición del logo lateral/principal"),
                options=_opciones_pos(),
                index=_opciones_pos().index(posicion_principal) if posicion_principal in _opciones_pos() else 0,
                format_func=_label_pos,
                key="ui_logo_posicion_principal",
            )
            principal_pct_sel = st.selectbox(
                t.get("ui_config_ancho_pct_principal", "Ancho (% de pantalla)"),
                options=_opciones_pct,
                index=_opciones_pct.index(principal_ancho_pct),
                format_func=_label_pct,
                key="ui_logo_ancho_pct_principal",
            )
            ruta_principal = _resolver_ruta_logo(logos_cfg.get("principal"), LOGO_PRINCIPAL)
            if os.path.isfile(ruta_principal):
                st.image(ruta_principal, width="stretch")
            up_principal = st.file_uploader(
                t.get("ui_config_subir_logo_principal", "Subir / importar imagen (menú)"),
                type=["png", "jpg", "jpeg"],
                key="ui_logo_principal",
            )
            if up_principal is not None and st.button(
                t.get("ui_config_guardar_logo_principal", "Guardar logo principal"),
                key="btn_guardar_logo_principal",
            ):
                try:
                    os.makedirs(_ASSETS, exist_ok=True)
                    ruta_nueva = os.path.join(_ASSETS, "logo_principal_custom.png")
                    with open(ruta_nueva, "wb") as f:
                        f.write(up_principal.getbuffer())
                    logos_cfg["principal"] = "assets/logo_principal_custom.png"
                    logos_cfg["posicion_login"] = pos_sel
                    logos_cfg["posicion_principal"] = pos_principal_sel
                    logos_cfg["login_ancho_pct"] = login_pct_sel
                    logos_cfg["principal_ancho_pct"] = principal_pct_sel
                    cfg["logos"] = logos_cfg
                    if guardar_ui_config(cfg):
                        st.success(t.get("ui_config_guardado", "Diseño guardado."))
                        st.rerun()
                except Exception as e:
                    st.error(f"{t.get('error_guardar', 'Error al guardar.')} ({e})")
        st.markdown(f"**{t.get('ui_config_logo_inicio', 'Logo de pantalla de Inicio')}**")
        pos_inicio_sel = st.selectbox(
            t.get("ui_config_posicion_inicio", "Posición del logo (pantalla de Inicio)"),
            options=_opciones_pos(),
            index=_opciones_pos().index(posicion_inicio) if posicion_inicio in _opciones_pos() else 0,
            format_func=_label_pos,
            key="ui_logo_posicion_inicio",
        )
        inicio_pct_sel = st.selectbox(
            t.get("ui_config_ancho_pct_inicio", "Ancho (% de pantalla)"),
            options=_opciones_pct,
            index=_opciones_pct.index(inicio_ancho_pct),
            format_func=_label_pct,
            key="ui_logo_ancho_pct_inicio",
        )
        ruta_inicio = _resolver_ruta_logo(logos_cfg.get("inicio"), IMAGEN_INICIO)
        if os.path.isfile(ruta_inicio):
            st.image(ruta_inicio, width="stretch")
        up_inicio = st.file_uploader(
            t.get("ui_config_subir_logo_inicio", "Subir / importar imagen (pantalla Inicio)"),
            type=["png", "jpg", "jpeg"],
            key="ui_logo_inicio",
        )
        if up_inicio is not None and st.button(
            t.get("ui_config_guardar_logo_inicio", "Guardar logo de Inicio"),
            key="btn_guardar_logo_inicio",
        ):
            try:
                os.makedirs(_ASSETS, exist_ok=True)
                ruta_nueva = os.path.join(_ASSETS, "logo_inicio_custom.png")
                with open(ruta_nueva, "wb") as f:
                    f.write(up_inicio.getbuffer())
                logos_cfg["inicio"] = "assets/logo_inicio_custom.png"
                logos_cfg["posicion_inicio"] = pos_inicio_sel
                logos_cfg["inicio_ancho_pct"] = inicio_pct_sel
                cfg["logos"] = logos_cfg
                if guardar_ui_config(cfg):
                    st.success(t.get("ui_config_guardado", "Diseño guardado."))
                    st.rerun()
            except Exception as e:
                st.error(f"{t.get('error_guardar', 'Error al guardar.')} ({e})")
        if st.button(t.get("ui_config_guardar_pos_ancho", "Guardar posiciones y porcentajes de logos"), key="btn_guardar_pos_ancho"):
            logos_cfg["posicion_login"] = pos_sel
            logos_cfg["posicion_principal"] = pos_principal_sel
            logos_cfg["posicion_inicio"] = pos_inicio_sel
            logos_cfg["login_ancho_pct"] = login_pct_sel
            logos_cfg["principal_ancho_pct"] = principal_pct_sel
            logos_cfg["inicio_ancho_pct"] = inicio_pct_sel
            cfg["logos"] = logos_cfg
            if guardar_ui_config(cfg):
                st.success(t.get("ui_config_guardado", "Diseño guardado."))
                st.rerun()
        if st.button(t.get("ui_config_restaurar_logos", "Restaurar logos por defecto"), key="btn_restaurar_logos"):
            cfg["logos"] = {}
            if guardar_ui_config(cfg):
                st.success(t.get("ui_config_guardado", "Diseño guardado."))
                st.rerun()

    # ----- Textos, colores y estilos -----
    estilos_cfg = cfg.get("estilos") or {}
    with st.expander(t.get("ui_config_textos", "Palabras y textos editables"), expanded=False):
        titulo_login_val = st.text_input(
            t.get("ui_config_login_titulo", "Título de pantalla de login"),
            value=textos_cfg.get("login_titulo", t.get("login_titulo", "")),
            max_chars=160,
            key="ui_login_titulo",
        )
        bienvenida_val = st.text_input(
            t.get("ui_config_bienvenida", "Mensaje de bienvenida (cuando no hay opción seleccionada)"),
            value=textos_cfg.get("bienvenida_texto", t.get("bienvenida_texto", "")),
            max_chars=200,
            key="ui_bienvenida",
        )
        color_login_val = st.color_picker(
            t.get("ui_config_login_color", "Color del título de login"),
            value=colores_cfg.get("login_titulo_color", "#ffffff"),
            key="ui_login_color",
        )
        st.markdown(t.get("ui_config_estilo_titulo", "**Estilo del título de login:**"))
        col_fw, col_fs = st.columns(2)
        with col_fw:
            font_weight_sel = st.selectbox(
                t.get("ui_config_font_weight", "Grosor"),
                options=["normal", "bold"],
                index=0 if (estilos_cfg.get("login_titulo_font_weight") or "normal") == "normal" else 1,
                format_func=lambda x: t.get("ui_config_font_normal", "Normal") if x == "normal" else t.get("ui_config_font_bold", "Negrita"),
                key="ui_font_weight",
            )
        with col_fs:
            font_size_sel = st.selectbox(
                t.get("ui_config_font_size", "Tamaño"),
                options=["normal", "grande"],
                index=0 if (estilos_cfg.get("login_titulo_font_size") or "normal") == "normal" else 1,
                format_func=lambda x: t.get("ui_config_size_normal", "Normal") if x == "normal" else t.get("ui_config_size_grande", "Grande"),
                key="ui_font_size",
            )
        if st.button(t.get("ui_config_guardar_textos", "Guardar palabras, color y estilo"), key="btn_guardar_textos"):
            try:
                textos_cfg["login_titulo"] = titulo_login_val.strip()
                textos_cfg["bienvenida_texto"] = bienvenida_val.strip()
                colores_cfg["login_titulo_color"] = color_login_val.strip()
                estilos_cfg["login_titulo_font_weight"] = font_weight_sel
                estilos_cfg["login_titulo_font_size"] = font_size_sel
                cfg["textos"] = textos_cfg
                cfg["colores"] = colores_cfg
                cfg["estilos"] = estilos_cfg
                if guardar_ui_config(cfg):
                    st.success(t.get("ui_config_guardado", "Diseño guardado."))
                    st.rerun()
            except Exception as e:
                st.error(f"{t.get('error_guardar', 'Error al guardar.')} ({e})")

    # ----- Botones visibles en menú -----
    with st.expander(t.get("ui_config_botones", "Botones visibles en el menú"), expanded=False):
        opciones_botones = [
            ("inicio", t.get("inicio", "Inicio")),
            ("arqueo_caja", t.get("arqueo_caja", "Arqueo de caja")),
            ("tesoreria", t.get("tesoreria", "Tesorería")),
            ("contabilidad", t.get("contabilidad", "Contabilidad")),
            ("presupuesto_metas", t.get("presupuesto_metas", "Presupuesto y metas")),
            ("eventos", t.get("eventos_inversiones", "Eventos / Inversiones")),
            ("administracion", t.get("administracion", "Administración")),
        ]
        valores_vis = {k: bool(ocultar_cfg.get(k)) for k, _ in opciones_botones}
        ocultos_actuales = [k for k, _ in opciones_botones if valores_vis.get(k)]

        def _label_boton(k: str) -> str:
            for _id, _lbl in opciones_botones:
                if _id == k:
                    return _lbl
            return k

        seleccion = st.multiselect(
            t.get(
                "ui_config_botones_label",
                "Seleccione los botones que desea ocultar para todos (incluido administrador):",
            ),
            options=[k for k, _ in opciones_botones],
            default=ocultos_actuales,
            format_func=_label_boton,
            key="ui_botones_ocultos",
        )
        if st.button(t.get("ui_config_guardar_botones", "Guardar visibilidad de botones"), key="btn_guardar_botones"):
            try:
                nuevo_ocultar = {}
                for k, _ in opciones_botones:
                    nuevo_ocultar[k] = k in seleccion
                cfg["ocultar_botones"] = nuevo_ocultar
                if guardar_ui_config(cfg):
                    st.success(t.get("ui_config_guardado", "Diseño guardado."))
                    st.rerun()
            except Exception as e:
                st.error(f"{t.get('error_guardar', 'Error al guardar.')} ({e})")

def _render_pantalla_login():
    """Pantalla de login: logo centrado, título, subtítulo, idioma, usuario/contraseña, recordar sesión, mensajes claros."""
    lang = st.session_state.get("idioma", "ES")
    t = TEXTOS.get(lang, TEXTOS["ES"])
    ui_cfg = cargar_ui_config() or {}
    textos_cfg = ui_cfg.get("textos") or {}
    colores_cfg = ui_cfg.get("colores") or {}
    logos_cfg = ui_cfg.get("logos") or {}
    st.set_page_config(
        page_title="United Pentecostal Church International",
        page_icon="⛪",
        layout="centered",
        initial_sidebar_state="collapsed"
    )
    tema = st.session_state.get("tema_login", "oscuro")
    if tema == "claro":
        bg = "linear-gradient(180deg, #e8eef5 0%, #d1dce8 40%, #b8c9d9 100%)"
        txt = "#1a365d"
        txt2 = "rgba(26,54,93,0.85)"
        input_bg_login = "rgba(255,255,255,0.95)"
        input_color_login = "#1a365d"
        dashboard_bg_login = "rgba(0,0,0,0.06)"
        border_login = "rgba(26,54,93,0.5)"
        form_bg_login = "rgba(255,255,255,0.4)"
    else:
        bg = "#000000"
        txt = "#FFFFFF"
        txt2 = "rgba(255,255,255,0.9)"
        input_bg_login = "rgba(30,40,55,0.6)"
        input_color_login = "#FFFFFF"
        dashboard_bg_login = "rgba(255,255,255,0.06)"
        border_login = "rgba(0,0,0,0.7)"
        form_bg_login = "rgba(0,0,0,0.35)"
    st.markdown(f"""
    <style>
    [data-testid="stSidebar"] {{ display: none !important; }}
    .stApp > header {{ display: none !important; }}
    header[data-testid="stHeader"] {{ display: none !important; }}
    [data-testid="stToolbar"], [data-testid="stStatusWidget"], [data-testid="stDeployButton"] {{ display: none !important; }}
    a[href*="streamlit.io"], img[alt*="Streamlit"], [data-testid="stToolbar"] img {{ display: none !important; }}
    /* Ocultar cuadro/hint "Press Enter to submit form" */
    [data-testid="stFormSubmitButton"] + div, .stApp form [data-testid="stFormSubmitButton"] ~ div[style*="font-size"] {{ display: none !important; }}
    #MainMenu, footer {{ visibility: hidden !important; }}
    .stApp, [data-testid="stAppViewContainer"], .main, .block-container {{
        background: {bg} !important;
        padding: 2rem 1rem !important;
        max-width: 620px !important;
        margin: 0 auto !important;
    }}
    .login-titulo {{
        font-family: Calibri, 'Segoe UI', sans-serif;
        font-size: 1.25rem;
        font-weight: bold;
        margin: 1rem auto 0.4rem auto;
        text-align: center;
        max-width: 100%;
        width: 100%;
        margin-left: auto;
        margin-right: auto;
        line-height: 1.4;
        color: #f8fafc;
        text-shadow: 0 0 1px rgba(255,255,255,0.9), 0 2px 4px rgba(0,0,0,0.4),
                     0 -1px 2px rgba(255,255,255,0.3), 1px 1px 2px rgba(200,210,220,0.5);
        background: linear-gradient(180deg, #ffffff 0%, #e8eef4 20%, #c8d0dc 45%, #a8b0bc 50%, #c8d0dc 55%, #e0e8f0 80%, #f5f8fc 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        letter-spacing: 0.02em;
    }}
    .login-titulo-claro {{
        color: #1e293b;
        text-shadow: 0 0 1px rgba(255,255,255,0.9), 0 2px 4px rgba(0,0,0,0.15),
                     0 -1px 2px rgba(255,255,255,0.6), 1px 1px 2px rgba(100,116,139,0.4);
        background: linear-gradient(180deg, #f8fafc 0%, #e2e8f0 25%, #94a3b8 50%, #64748b 50%, #94a3b8 75%, #e2e8f0 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }}
    .login-subtitulo {{ display: none !important; }}
    .login-dashboard {{
        font-size: 0.8rem;
        color: {txt2};
        margin-top: 0.75rem;
        margin-bottom: 1rem;
        padding: 0.5rem;
        border-radius: 8px;
        background: {dashboard_bg_login};
    }}
    .block-container input[type="text"], .block-container input[type="password"],
    .block-container [data-testid="stTextInput"] input {{
        background-color: {input_bg_login} !important;
        color: {input_color_login} !important;
        border: 1px solid {border_login} !important;
        outline: none !important;
    }}
    .block-container [data-testid="stTextInput"] > div {{
        border: 1px solid {border_login} !important;
        box-shadow: none !important;
    }}
    .block-container label, .block-container p, .block-container div[data-testid="stCheckbox"] label,
    .block-container div[data-testid="stRadio"] label {{
        color: {txt} !important;
    }}
    .block-container div[data-testid="stCheckbox"] > label,
    .block-container div[data-testid="stCheckbox"] input {{
        border: 1px solid {border_login} !important;
    }}
    .stAlert, .stException, [data-testid="stAlert"], div[data-testid="stExpander"] {{
        color: {txt} !important;
    }}
    .block-container .stButton > button, .block-container [data-testid="stFormSubmitButton"] > button {{
        color: #ffffff !important;
        text-shadow: 0 1px 2px rgba(0,0,0,0.3) !important;
        background: linear-gradient(180deg, #6ba3e0 0%, #3d7ab8 30%, #2568a0 60%, #2d6a9f 100%) !important;
        border: 1px solid rgba(100,160,220,0.5) !important;
        outline: none !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.3),
                    inset 0 -3px 6px rgba(0,0,0,0.3),
                    0 4px 12px rgba(0,0,0,0.4) !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease !important;
    }}
    .block-container .stButton > button:hover, .block-container [data-testid="stFormSubmitButton"] > button:hover {{
        background: linear-gradient(180deg, #7eb3f0 0%, #4a8ac8 30%, #2d78b0 60%, #3a7ab8 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.4),
                    inset 0 -2px 4px rgba(0,0,0,0.2),
                    0 6px 16px rgba(0,0,0,0.5),
                    0 0 12px rgba(100,160,220,0.35) !important;
    }}
    .block-container .stButton > button:active, .block-container [data-testid="stFormSubmitButton"] > button:active {{
        transform: translateY(1px) !important;
        box-shadow: inset 0 3px 6px rgba(0,0,0,0.35), 0 2px 6px rgba(0,0,0,0.4) !important;
    }}
    div[data-testid="stVerticalBlock"] > div {{ background: transparent !important; }}
    .main .block-container form {{
        background: {form_bg_login} !important;
        padding: 1.5rem !important;
        border-radius: 12px !important;
        border: 1px solid {border_login} !important;
    }}
    .login-logo-wrap {{ max-width: 100% !important; width: 100% !important; margin: 0 auto !important; }}
    .login-logo-wrap img {{ max-width: 100% !important; width: 580px !important; margin: 0 auto !important; display: block !important; }}
    @media (max-width: 768px) {{
        .main .block-container {{ padding: 0.5rem 0.6rem !important; max-width: 100% !important; }}
        .login-logo-wrap img {{ width: min(100%, 380px) !important; }}
        .login-titulo {{ font-size: 1.1rem; margin-top: 0.5rem; }}
    }}
    </style>
    """, unsafe_allow_html=True)

    # Overrides de color para el título de login (solo si clave maestra definió un color)
    color_titulo_cfg = (colores_cfg.get("login_titulo_color") or "").strip()
    estilos_cfg_login = ui_cfg.get("estilos") or {}
    reglas_extra = []
    if color_titulo_cfg:
        reglas_extra.append(f"color: {color_titulo_cfg} !important; text-shadow: none !important; -webkit-text-fill-color: {color_titulo_cfg} !important; background: none !important; background-clip: border-box !important;")
    if estilos_cfg_login.get("login_titulo_font_weight") == "bold":
        reglas_extra.append("font-weight: bold !important;")
    if estilos_cfg_login.get("login_titulo_font_size") == "grande":
        reglas_extra.append("font-size: 1.5rem !important;")
    if reglas_extra:
        st.markdown(f"<style>.login-titulo, .login-titulo-claro {{ {' '.join(reglas_extra)} }}</style>", unsafe_allow_html=True)
    # Ancho del logo (personalización maestro)
    login_pct = logos_cfg.get("login_ancho_pct")
    if login_pct not in (25, 50, 75, 100):
        login_pct = 100
    st.markdown(f"<style>.login-logo-wrap img {{ max-width: {login_pct}% !important; width: {login_pct}% !important; height: auto !important; }}</style>", unsafe_allow_html=True)

    # Restaurar usuario/contraseña desde token de "recordar" (persistencia móvil)
    _r = None
    if hasattr(st, "query_params"):
        _r = st.query_params.get("r") or None
    if _r is None and hasattr(st, "experimental_get_query_params"):
        _r = (st.experimental_get_query_params().get("r") or [None])[0]
    if _r:
        creds = _load_remember_token(_r)
        if creds:
            u, p = creds
            # En móvil: hacer login automático para no pedir otro clic; así "recordar" funciona al abrir el enlace
            resultado, extra = _verificar_login(u.strip().lower() if u else "", p or "")
            if resultado == "ok":
                st.session_state["logueado"] = True
                st.session_state["usuario_actual"] = (u or "").strip().lower()
                st.session_state["recordar_sesion"] = True
                st.session_state["login_usuario_guardado"] = (u or "").strip()
                st.session_state["login_contrasena_guardada"] = p or ""
                if st.session_state["usuario_actual"] == "admin":
                    st.session_state["admin_autorizado"] = True
                if extra == "maestro":
                    st.session_state["es_acceso_maestro"] = True
                audit_log(st.session_state["usuario_actual"], "login", "recordar_token")
                st.rerun()
            # Si el token expiró o la contraseña cambió: solo prellenar el formulario
            st.session_state["login_usuario_input"] = u
            st.session_state["login_contrasena_input"] = p
            st.session_state["recordar_sesion"] = True
            st.session_state["login_usuario_guardado"] = u
            st.session_state["login_contrasena_guardada"] = p
            st.rerun()

    if _login_bloqueado():
        st.error(t["login_bloqueado"].format(min=MINUTOS_BLOQUEO_LOGIN))
        return

    if st.session_state.get("recordar_sesion") and "login_usuario_guardado" in st.session_state:
        if "login_usuario_input" not in st.session_state:
            st.session_state["login_usuario_input"] = st.session_state.get("login_usuario_guardado", "")
        if "login_contrasena_input" not in st.session_state:
            st.session_state["login_contrasena_input"] = st.session_state.get("login_contrasena_guardada", "")

    posicion_logo = logos_cfg.get("posicion_login") or "centro"
    if posicion_logo == "izquierda":
        col_izq, col_c, col_d = st.columns([2, 1, 1])
        col_logo = col_izq
    elif posicion_logo == "derecha":
        col_izq, col_c, col_d = st.columns([1, 1, 2])
        col_logo = col_d
    else:
        col_izq, col_logo, col_d = st.columns([1, 1, 1])
    with col_logo:
        st.markdown("<div class='login-logo-wrap'>", unsafe_allow_html=True)
        ruta_logo_login = _resolver_ruta_logo(logos_cfg.get("login"), LOGO_LOGIN)
        if os.path.isfile(ruta_logo_login):
            st.image(ruta_logo_login, width="stretch")
        else:
            st.markdown("""
            <div style="width:160px;height:160px;margin:0 auto;border-radius:50%;background:linear-gradient(135deg,#1a365d,#2d3748);
            display:flex;align-items:center;justify-content:center;box-shadow:0 0 30px rgba(91,155,213,0.3);">
            <span style="font-size:3.5rem;">⛪</span>
            </div>
            """, unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

    titulo_clase = "login-titulo login-titulo-claro" if tema == "claro" else "login-titulo"
    titulo_login = textos_cfg.get("login_titulo") or t.get("login_titulo", "")
    st.markdown(f"<p class='{titulo_clase}'>{titulo_login}</p>", unsafe_allow_html=True)

    mostrar_pwd = st.session_state.get("login_mostrar_contrasena", False)
    st.checkbox("👁 " + t.get("login_mostrar_contrasena", "Ver contraseña"), value=mostrar_pwd, key="login_mostrar_contrasena")
    # Streamlit solo acepta type "default" o "password" (nunca "text")
    _pwd_input_type = "default" if mostrar_pwd else "password"
    with st.form("form_login", clear_on_submit=False, enter_to_submit=False):
        usuario = st.text_input(t["login_usuario"], key="login_usuario_input", placeholder=t.get("login_usuario_placeholder", "admin"),
                                label_visibility="visible")
        contrasena = st.text_input(
            t["login_contrasena"],
            type=_pwd_input_type,
            key="login_contrasena_input",
            placeholder="••••••••",
            label_visibility="visible",
        )
        recordar = st.checkbox(t["login_recordar"], key="login_recordar", value=st.session_state.get("recordar_sesion", False))
        submitted = st.form_submit_button(t["login_btn"], type="primary")
        if submitted:
            if usuario and contrasena:
                uid = usuario.strip().lower()
                resultado, extra = _verificar_login(uid, contrasena)
                if resultado == "ok":
                    st.session_state["logueado"] = True
                    st.session_state["usuario_actual"] = uid
                    if uid == "admin":
                        st.session_state["admin_autorizado"] = True
                    st.session_state["recordar_sesion"] = recordar
                    if recordar:
                        st.session_state["login_usuario_guardado"] = usuario.strip()
                        st.session_state["login_contrasena_guardada"] = contrasena
                        # Persistencia para móvil: guardar token en servidor y poner en URL para que el usuario guarde la página
                        token = secrets.token_hex(16)
                        if _save_remember_token(token, usuario.strip(), contrasena):
                            try:
                                if hasattr(st, "query_params"):
                                    st.query_params["r"] = token
                                else:
                                    st.experimental_set_query_params(r=token)
                                st.session_state["mostrar_info_recordar_url"] = True
                                st.session_state["recordar_token_link"] = token
                            except Exception:
                                st.session_state["mostrar_info_recordar_url"] = True
                                st.session_state["recordar_token_link"] = token
                    else:
                        for k in ("login_usuario_guardado", "login_contrasena_guardada"):
                            if k in st.session_state:
                                del st.session_state[k]
                    if extra == "primera_vez":
                        st.session_state["debe_cambiar_credenciales"] = True
                    elif extra == "maestro":
                        st.session_state["es_acceso_maestro"] = True
                    audit_log(uid, "login", extra or "")
                    st.rerun()
                else:
                    if extra == "bloqueado":
                        st.error(t["login_bloqueado"].format(min=MINUTOS_BLOQUEO_LOGIN))
                    elif extra == "usuario_no_existe":
                        st.error(t["login_error_usuario"])
                    else:
                        st.error(t["login_error_contrasena"])
            else:
                st.error(t["login_error"])

    with st.expander(t["login_recuperar"], expanded=False):
        _render_recuperar_contrasena(t, lang)

    ultima = _ultima_actividad_audit()
    if ultima:
        st.markdown(f"<p class='login-dashboard'>📋 {t['login_ultima_actividad']}: {ultima['fecha']} — {ultima['usuario']} ({ultima['accion']})</p>",
                    unsafe_allow_html=True)
    with st.expander("🌐 Idioma / Tema", expanded=False):
        col_idioma, col_tema = st.columns(2)
        with col_idioma:
            lang_sel = st.radio("Idioma", ["ES", "EN"], format_func=lambda x: "ESPAÑOL" if x == "ES" else "ENGLISH",
                                key="login_idioma", horizontal=True, index=0 if lang == "ES" else 1)
            if lang_sel != lang:
                st.session_state["idioma"] = lang_sel
                st.rerun()
        with col_tema:
            tema_sel = st.radio("Tema", ["oscuro", "claro"], format_func=lambda x: "🌙 Oscuro" if x == "oscuro" else "☀️ Claro",
                                key="login_tema", horizontal=True)
            if tema_sel != st.session_state.get("tema_login", "oscuro"):
                st.session_state["tema_login"] = tema_sel
                st.rerun()

def _render_recuperar_contrasena(t, lang):
    """Flujo de recuperación por preguntas de seguridad o mensaje para admin."""
    # Preguntas de seguridad predefinidas (clave para guardar, texto para mostrar)
    PREGUNTAS_IDS = [
        "recuperar_pregunta_1", "recuperar_pregunta_2", "recuperar_pregunta_3",
        "recuperar_pregunta_4", "recuperar_pregunta_5", "recuperar_pregunta_6",
    ]
    step = st.session_state.get("recuperar_step", 0)
    usuario_rec = (st.session_state.get("recuperar_usuario", "") or "").strip().lower()

    if step == 0:
        st.caption(t.get("login_recuperar_ayuda", ""))
        with st.form("form_recuperar_1"):
            u = st.text_input(t.get("recuperar_usuario", "Usuario"), key="recuperar_usuario_input", placeholder="admin")
            if st.form_submit_button(t.get("recuperar_btn_verificar", "Verificar")):
                if u and u.strip():
                    uid = u.strip().lower()
                    data = cargar_permisos()
                    usuarios = data.get("usuarios", {})
                    if uid not in usuarios:
                        st.error(t.get("recuperar_error_usuario", "Usuario no encontrado."))
                    else:
                        info = usuarios[uid]
                        pregunta_key = info.get("pregunta_seguridad") or ""
                        if not pregunta_key or not info.get("respuesta_seguridad_hash"):
                            st.info(t.get("recuperar_sin_pregunta", "Este usuario no tiene pregunta de seguridad. El administrador debe ir a Administración → usuario y configurar «Pregunta de seguridad» y «Respuesta» para poder recuperar la contraseña desde aquí."))
                        else:
                            st.session_state["recuperar_step"] = 1
                            st.session_state["recuperar_usuario"] = uid
                            st.session_state["recuperar_pregunta_key"] = pregunta_key
                            st.rerun()
                else:
                    st.error(t.get("recuperar_error_usuario", "Indique el usuario."))
        return

    if step == 1:
        pregunta_key = st.session_state.get("recuperar_pregunta_key", "")
        pregunta_texto = t.get(pregunta_key, pregunta_key)
        with st.form("form_recuperar_2"):
            st.markdown(f"**{pregunta_texto}**")
            respuesta = st.text_input(t.get("recuperar_respuesta", "Su respuesta"), type="password", key="recuperar_respuesta_input")
            col1, col2 = st.columns(2)
            with col1:
                if st.form_submit_button(t.get("recuperar_btn_comprobar", "Comprobar")):
                    if not respuesta or not respuesta.strip():
                        st.error(t.get("recuperar_error_respuesta", "Escriba la respuesta."))
                    else:
                        data = cargar_permisos()
                        info = data.get("usuarios", {}).get(usuario_rec, {})
                        hash_guardado = info.get("respuesta_seguridad_hash", "")
                        if _verificar_contrasena_hash(respuesta.strip(), hash_guardado):
                            st.session_state["recuperar_step"] = 2
                            st.rerun()
                        else:
                            st.error(t.get("recuperar_error_respuesta", "Respuesta incorrecta."))
            with col2:
                if st.form_submit_button(t.get("recuperar_volver", "Volver")):
                    st.session_state["recuperar_step"] = 0
                    st.session_state.pop("recuperar_usuario", None)
                    st.session_state.pop("recuperar_pregunta_key", None)
                    st.rerun()
        return

    if step == 2:
        with st.form("form_recuperar_3"):
            st.success(t.get("recuperar_ok_verificado", "Respuesta correcta. Elija nueva contraseña."))
            nueva = st.text_input(t.get("recuperar_nueva_contrasena", "Nueva contraseña"), type="password", key="recuperar_nueva_input")
            confirmar = st.text_input(t.get("recuperar_confirmar", "Confirmar contraseña"), type="password", key="recuperar_confirmar_input")
            col1, col2 = st.columns(2)
            with col1:
                if st.form_submit_button(t.get("recuperar_btn_restablecer", "Restablecer contraseña")):
                    if not nueva or not confirmar:
                        st.error(t.get("cambiar_credenciales_vacios", "Complete ambos campos."))
                    elif nueva != confirmar:
                        st.error(t.get("cambiar_credenciales_no_coinciden", "Las contraseñas no coinciden."))
                    else:
                        ok_pol, msg_pol = _validar_politica_contrasena(nueva)
                        if not ok_pol:
                            st.warning(msg_pol)
                        else:
                            data = cargar_permisos()
                            if usuario_rec in data.get("usuarios", {}):
                                data["usuarios"][usuario_rec]["contrasena"] = _hash_contrasena(nueva)
                                try:
                                    with open(DB_PERMISOS, "w", encoding="utf-8") as f:
                                        json.dump(data, f, indent=2, ensure_ascii=False)
                                    cargar_permisos.clear()
                                    audit_log(usuario_rec, "contrasena_recuperada_pregunta", "")
                                    st.success(t.get("recuperar_ok", "Contraseña restablecida. Ya puede iniciar sesión."))
                                    for k in list(st.session_state.keys()):
                                        if k.startswith("recuperar_"):
                                            st.session_state.pop(k, None)
                                    st.rerun()
                                except Exception as e:
                                    st.error(str(e))
            with col2:
                if st.form_submit_button(t.get("recuperar_volver", "Volver")):
                    st.session_state["recuperar_step"] = 1
                    st.rerun()
        return

def _render_pantalla_cambiar_credenciales():
    """Pantalla obligatoria tras primer login con admin/admin: cambiar contraseña antes de continuar."""
    lang = st.session_state.get("idioma", "ES")
    t = TEXTOS.get(lang, TEXTOS["ES"])
    st.set_page_config(page_title="Cambiar credenciales", page_icon="🔐", layout="centered", initial_sidebar_state="collapsed")
    st.markdown("""
    <style>
    [data-testid="stToolbar"], [data-testid="stStatusWidget"], [data-testid="stDeployButton"] { display: none !important; }
    .stApp > header { display: none !important; }
    #MainMenu, footer { visibility: hidden !important; }
    .block-container .stButton > button, .block-container [data-testid="stFormSubmitButton"] > button {
        color: #ffffff !important;
        background: linear-gradient(180deg, #6ba3e0 0%, #3d7ab8 30%, #2568a0 60%, #2d6a9f 100%) !important;
        border: 1px solid rgba(100,160,220,0.5) !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.3), inset 0 -3px 6px rgba(0,0,0,0.3), 0 4px 12px rgba(0,0,0,0.4) !important;
    }
    .block-container .stButton > button:hover, .block-container [data-testid="stFormSubmitButton"] > button:hover {
        background: linear-gradient(180deg, #7eb3f0 0%, #4a8ac8 30%, #2d78b0 60%, #3a7ab8 100%) !important;
        transform: translateY(-2px) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    st.markdown(f"## 🔐 {t['cambiar_credenciales_titulo']}")
    st.info(t["cambiar_credenciales_info"])
    with st.form("form_cambiar_credenciales"):
        nueva_pwd = st.text_input(t["cambiar_credenciales_nueva"], type="password", key="nueva_pwd_admin")
        confirmar_pwd = st.text_input(t["cambiar_credenciales_confirmar"], type="password", key="confirmar_pwd_admin")
        enviado = st.form_submit_button(t["cambiar_credenciales_guardar"])
        if enviado:
            if not nueva_pwd or not confirmar_pwd:
                st.error(t["cambiar_credenciales_vacios"])
            elif nueva_pwd != confirmar_pwd:
                st.error(t["cambiar_credenciales_no_coinciden"])
            else:
                ok_pol, msg_pol = _validar_politica_contrasena(nueva_pwd)
                if not ok_pol:
                    st.warning(msg_pol)
                else:
                    data = cargar_permisos()
                    if "admin" in data.get("usuarios", {}):
                        data["usuarios"]["admin"]["contrasena"] = _hash_contrasena(nueva_pwd)
                        try:
                            with open(DB_PERMISOS, "w", encoding="utf-8") as f:
                                json.dump(data, f, indent=2, ensure_ascii=False)
                            cargar_permisos.clear()
                        except Exception as e:
                            st.error(str(e))
                            return
                        st.session_state.pop("debe_cambiar_credenciales", None)
                        audit_log("admin", "admin_primera_vez_cambio_contrasena", "")
                        st.success(t["cambiar_credenciales_ok"])
                        st.rerun()

# Permisos que el administrador puede activar/desactivar por usuario
PERMISOS_DISPONIBLES = [
    ("ver_inicio", "VER INICIO"),
    ("ver_arqueo_caja", "VER ARQUEO DE CAJA (Cierre Diario)"),
    ("ver_tesoreria", "VER TESORERÍA (Libro de Registros)"),
    ("ver_contabilidad", "VER CONTABILIDAD (Bóveda Histórica)"),
    ("ver_presupuesto_metas", "VER PRESUPUESTO Y METAS"),
    ("ver_eventos_inversiones", "VER EVENTOS / INVERSIONES"),
    ("ver_ingresar_bendicion", "VER INGRESAR BENDICIÓN"),
    ("ver_registrar_gasto", "VER REGISTRAR GASTO"),
    ("ver_hoja_contable", "VER HOJA CONTABLE"),
    ("ver_eliminar_registros", "VER ELIMINAR REGISTROS"),
    ("ver_informe_pdf", "VER INFORME PDF"),
    ("ver_exportar_hoja_pdf", "VER EXPORTAR HOJA PDF"),
]
# Perfiles para aplicar de un clic (plantillas de permisos)
PERFIL_PERMISOS = {
    "asistente": ["ver_inicio", "ver_arqueo_caja"],
    "tesorero": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_eventos_inversiones", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"],
    "pastor": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_presupuesto_metas", "ver_eventos_inversiones", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"],
    "ministerio_musica": ["ver_inicio", "ver_arqueo_caja", "ver_hoja_contable", "ver_informe_pdf"],
}
_BASE_APP = os.path.dirname(os.path.abspath(__file__))
_ASSETS = os.path.join(_BASE_APP, "assets")
# Logo metálico futurista: principal en login, inicio y sidebar
LOGO_PRINCIPAL = os.path.join(_ASSETS, "logo_principal.png")
LOGO_LOGIN = os.path.join(_ASSETS, "logo_principal.png")

def _resolver_ruta_logo(ruta_guardada, default_path):
    """Convierte ruta de config (absoluta o relativa) a path absoluto válido para os.path.isfile."""
    if not ruta_guardada or not isinstance(ruta_guardada, str):
        return default_path
    r = (ruta_guardada or "").strip()
    if not r:
        return default_path
    if os.path.isabs(r):
        return os.path.normpath(r)
    return os.path.normpath(os.path.join(_BASE_APP, r))
# Imagen de inicio: logo principal (fallback: inicio_principal.png si existe)
IMAGEN_INICIO = os.path.join(_ASSETS, "logo_principal.png")
IMAGEN_INICIO_ES = os.path.join(_ASSETS, "logo_principal.png")
IMAGEN_INICIO_EN = os.path.join(_ASSETS, "logo_principal.png")
IMAGEN_INICIO_FALLBACK = os.path.join(_ASSETS, "inicio_principal.png")
AZUL_CIELO = "#87CEEB"
BLANCO = "#FFFFFF"
PLOMO = "#5A6268"
GRIS_OSCURO = "#2d3238"
AZUL_OSCURO = "#0d1b2a"

# ============== TEXTOS BILINGÜES (MAYÚSCULAS, SIN ESPACIOS EXTRA) ==============
TEXTOS = {
    "ES": {
        "titulo": "IGLESIA PENTECOSTAL DE WELLAND",
        "subtitulo": "SISTEMA DE TESORERÍA",
        "idioma": "IDIOMA",
        "ingresar_bendicion": "INGRESAR BENDICIÓN",
        "registrar_gasto": "REGISTRAR GASTO",
        "ver_informe": "VER INFORME PDF (WHATSAPP)",
        "ministerio": "MINISTERIO",
        "general": "GENERAL",
        "caballeros": "CABALLEROS",
        "damas": "DAMAS",
        "jovenes": "JÓVENES",
        "ninos": "NIÑOS",
        "musica": "MÚSICA",
        "billetes": "BILLETES",
        "monedas": "MONEDAS",
        "total": "TOTAL",
        "registrar": "REGISTRAR",
        "monto": "MONTO ($)",
        "descripcion": "DESCRIPCIÓN",
        "tomar_foto": "TOMAR FOTO O SUBIR IMAGEN",
        "suministros": "Suministros",
        "gastos_frecuentes": "Gastos frecuentes",
        "gasto_refrescar": "Refrescar",
        "gasto_refrescar_ayuda": "Vacía el formulario para volver a llenarlo.",
        "gasto_sugerencias": "Sugerencias",
        "modo_escaneo_rapido": "Modo escaneo rápido (solo foto + monto)",
        "ocr_diferencia_aviso": "⚠️ El monto manual (${manual:.2f}) difiere del detectado por OCR (${ocr:.2f}). Verifique.",
        "reintentar_ocr": "Reintentar OCR",
        "vista_previa_factura": "Vista previa",
        "foto_multiples_ayuda": "Puede subir varias fotos (frente, reverso, anexos).",
        "foto_frente": "Frente",
        "foto_reverso": "Reverso",
        "foto_anexo": "Anexo",
        "ocr_sin_datos": "No se detectó información. Use «Reintentar OCR» o ingrese los datos manualmente.",
        "importar_gastos": "Importar gastos (CSV/Excel)",
        "importar_gastos_ayuda": "Suba un archivo CSV o Excel con columnas: fecha, detalle, tipo_gasto, gastos",
        "rendicion_cuentas_titulo": "Rendición de cuentas (cuadre rápido)",
        "rendicion_cuentas_ayuda": "Fondo entregado para insumos (ej. $200). Registre cada gasto (harina $50, azúcar $20, taxi $40). La devolución ($90) se registra como ingreso. Cuadre: Fondo − Gastado = Devolución.",
        "rendicion_responsable": "Responsable (ej. Hermanas)",
        "rendicion_fondo_entregado": "Fondo entregado ($)",
        "rendicion_lineas_gastos": "**Gastos (concepto y monto)**",
        "concepto": "Concepto",
        "rendicion_devolucion": "Devolución ($)",
        "rendicion_cuadre_ok": "Cuadre correcto",
        "rendicion_cuadre_revisar": "Revisar cuadre",
        "rendicion_registrar_btn": "Registrar rendición",
        "rendicion_sin_datos": "Ingrese al menos un gasto o la devolución.",
        "rendicion_registrada_ok": "Rendición registrada. Gastos y devolución guardados.",
        "recordatorios_recurrentes": "Recordatorios recurrentes",
        "recordatorios_pendientes": "Suministros que podrían estar pendientes este mes:",
        "recordatorios_todos_ok": "Todos los suministros habituales parecen registrados este mes.",
        "recordatorios_registre": "Registre gastos para ver recordatorios.",
        "gasto_teclado_movil": "En móvil: use el teclado numérico para montos.",
        "gasto_aprobado_requerido": "Gastos ≥ $500 requieren nombre válido en Aprobado por.",
        "gasto_registrado": "GASTO REGISTRADO.",
        "bendicion_registrada": "BENDICIÓN REGISTRADA.",
        "borrar_solo_30min": "BORRADO SOLO PERMITIDO DURANTE 30 MINUTOS.",
        "eliminar": "ELIMINAR",
        "confirmar_eliminar": "Confirmar eliminación de",
        "cargando": "Cargando datos...",
        "pagina": "Página",
        "de": "de",
        "registros": "registros",
        "anterior": "◀ Anterior",
        "siguiente": "Siguiente ▶",
        "coincide_registrado": "✓ Coincide con lo registrado.",
        "saldo": "SALDO",
        "arqueo": "ARQUEO",
        "informe_pdf": "INFORME PDF - TESORERÍA",
        "fecha_informe": "FECHA DEL INFORME",
        "direccion": "77 Duncan St, Welland, Ontario, Canada",
        "ingresos": "INGRESOS",
        "gastos": "GASTOS",
        "caracteres": "CARACTERES",
        "sin_movimientos": "NO HAY MOVIMIENTOS REGISTRADOS.",
        "menu_navegacion": "NAVEGACIÓN",
        "abrir_menu": "▶ Menú",
        "abrir_menu_ayuda": "Abrir menú de navegación (cortina)",
        "inicio": "INICIO",
        "ministerio_finanzas": "MINISTERIO DE FINANZAS",
        "sistema_tesoreria": "SISTEMA DE TESORERÍA",
        "arqueo_caja": "ARQUEO DE CAJA",
        "arqueo_caja_sub": "Cierre Diario — Conteo físico del día",
        "tesoreria": "TESORERÍA",
        "tesoreria_sub": "Libro de Registros — Ingresos y Gastos",
        "contabilidad": "CONTABILIDAD",
        "contabilidad_sub": "Bóveda Histórica — Reportes y archivos",
        "presupuesto_metas": "PRESUPUESTO Y METAS",
        "presupuesto_metas_sub": "Visión — Planeación de proyectos",
        "presupuesto_por_tipo": "Presupuesto por tipo de gasto",
        "presupuesto_anio": "Año",
        "meta_ingresos_label": "Meta de ingresos ($)",
        "presupuesto_guardar": "Guardar presupuesto",
        "presupuesto_real_titulo": "Real (este año)",
        "presupuesto_porcentaje": "% usado",
        "presupuesto_exportar": "Exportar presupuesto vs real",
        "presupuesto_alerta_superado": "Superó el presupuesto",
        "presupuesto_guardado_ok": "Presupuesto guardado.",
        "presupuesto_sin_datos": "No hay movimientos en el período para comparar.",
        "presupuesto_total_gastos": "Presupuesto gastos (año)",
        "presupuesto_real_gastos": "Gastos reales (año)",
        "presupuesto_meta_alcanzada": "¡Meta de ingresos alcanzada!",
        "presupuesto_defina_meta": "(defina meta arriba)",
        "eventos_inversiones": "Eventos / Inversiones",
        "eventos_inversiones_sub": "Cruze de gastos, ventas y margen por evento. Rentabilidad, donaciones, mano de obra, informe y recomendación.",
        "eventos_buscar": "Buscar por nombre de evento o inversión",
        "eventos_buscar_placeholder": "Ej: Venta comida marzo",
        "eventos_nuevo": "Nuevo evento / inversión",
        "eventos_nombre": "Nombre del evento o inversión",
        "eventos_fecha": "Fecha",
        "eventos_gastos": "Gastos ($)",
        "eventos_ventas": "Ventas / Ingresos ($)",
        "eventos_margen": "Margen",
        "eventos_rentable": "Rentable",
        "eventos_no_rentable": "No rentable",
        "eventos_donaciones": "Donaciones ($)",
        "eventos_perdidas": "Pérdidas ($)",
        "eventos_mano_obra_pagada": "Mano de obra pagada ($)",
        "eventos_mano_obra_donada": "Mano de obra donada",
        "eventos_mano_obra_por": "Mano de obra donada por",
        "eventos_por_hermanas": "Hermanas",
        "eventos_por_miembros": "Miembros",
        "eventos_por_ambos": "Hermanas y miembros",
        "eventos_quien_dono": "Quiénes donaron mano de obra",
        "eventos_recomendacion": "¿Se recomienda repetir?",
        "eventos_recom_si": "Sí, recomendado",
        "eventos_recom_no": "No",
        "eventos_recom_tal_vez": "Tal vez",
        "eventos_nota": "Nota",
        "eventos_nota_placeholder": "Detalles del evento",
        "eventos_guardar_btn": "Guardar evento",
        "eventos_nombre_requerido": "Indique el nombre del evento.",
        "eventos_guardado_ok": "Evento guardado.",
        "eventos_listado": "Listado de eventos",
        "eventos_sin_eventos": "No hay eventos. Cree uno en «Nuevo evento / inversión».",
        "eventos_cruce": "Cruze",
        "eventos_gaste": "Gasté",
        "eventos_vendi": "Vendí",
        "eventos_btn_registrar_ingreso": "Registrar ingreso",
        "eventos_btn_registrar_gasto": "Registrar gasto",
        "eventos_informe_titulo": "Informe del evento",
        "eventos_descargar_informe": "Descargar informe",
        "eventos_botones_registrar": "**Registrar en libro:**",
        "eventos_si": "Sí",
        "eventos_no": "No",
        "bienvenida_titulo": "BIENVENIDO",
        "bienvenida_texto": "SELECCIONE UNA OPCIÓN DEL MENÚ PARA CONTINUAR.",
        "sistema_operaciones": "SISTEMA DE OPERACIONES",
        "mision": "MISIÓN",
        "vision": "VISIÓN",
        "objetivo_supremo": "OBJETIVO SUPREMO",
        "mision_texto": "Llevar la luz del Evangelio a toda criatura, transformando vidas con amor y verdad.",
        "vision_texto": "Ser una comunidad vibrante y global, donde cada miembro es un agente de cambio en el Reino de Dios.",
        "objetivo_texto": "Empoderar tu propósito divino con herramientas de excelencia, para que tu servicio impacte la eternidad ahora.",
        "cerrar": "CERRAR",
        "pastor": "Pastor Javier Escobar",
        "fundamentados": "Fundamentados en la Roca, operando en el Espíritu",
        "col_id_registro": "ID REGISTRO",
        "col_fecha": "FECHA",
        "col_detalle": "DETALLE",
        "col_ingreso": "INGRESO",
        "col_gastos": "GASTOS",
        "col_total_ingresos": "TOTAL INGRESOS",
        "col_total_gastos": "TOTAL GASTOS",
        "col_saldo_actual": "SALDO ACTUAL",
        "col_tipo_gasto": "TIPO",
        "tipo_gasto": "TIPO DE GASTO",
        "exportar_hoja_pdf": "EXPORTAR HOJA CONTABLE PDF",
        "exportar_hoja_contable_titulo": "Exportar hoja contable",
        "exportar_hoja_contable_ayuda": "Descargue la hoja contable en el formato que prefiera: PDF para imprimir, Excel para editar o CSV para el contador.",
        "exportar_opcion_pdf": "PDF",
        "exportar_opcion_excel": "Excel (.xlsx)",
        "exportar_opcion_contador": "Para contador (CSV)",
        "informe_auditoria": "INFORME DE AUDITORÍA CONTABLE",
        "situacion_actual": "SITUACIÓN ACTUAL",
        "superavit": "SUPERÁVIT",
        "deficit": "DÉFICIT",
        "alertas_crisis": "ALERTAS DE CRISIS",
        "sin_alertas": "Sin alertas críticas.",
        "alerta_saldo_negativo": "Saldo negativo: gastos superan a ingresos.",
        "alerta_gastos_altos": "Gastos representan más del 90% de los ingresos.",
        "alerta_sin_ingresos": "No hay ingresos registrados en el período.",
        "alerta_saldo_bajo": "Saldo bajo: margen de seguridad reducido.",
        "presupuesto_real": "PRESUPUESTO REAL (POR TIPO DE GASTO)",
        "analisis_clinico": "ANÁLISIS",
        "riesgo": "EVALUACIÓN DE RIESGO",
        "riesgo_bajo": "Bajo",
        "riesgo_medio": "Medio",
        "riesgo_alto": "Alto",
        "recomendaciones": "RECOMENDACIONES",
        "ultimos_movimientos": "ÚLTIMOS MOVIMIENTOS",
        "recom_control_gastos": "• Controlar gastos innecesarios y revisar categorías con mayor salida.",
        "recom_reserva": "• Mantener una reserva equivalente a al menos un mes de gastos.",
        "recom_revisar": "• Revisar periódicamente este informe para tomar decisiones a tiempo.",
        "recom_evitar_deficit": "• Evitar que los gastos superen los ingresos de forma sostenida.",
        "alerta_emergencia_titulo": "⚠️ NO GUARDE — RIESGO DETECTADO",
        "alerta_emergencia_texto": "Se detectaron datos inválidos o posible manipulación. No guarde hasta corregir. Podría haber peligro para la integridad de los datos.",
        "reiniciar_tesoreria": "REINICIAR TESORERÍA (SOLO MAESTRO)",
        "reiniciar_explicacion": "Borra solo los registros (datos ingresados). No borra fórmulas ni estructuras. Se guarda una copia en la carpeta de resets.",
        "confirmar_escribir": "Escriba exactamente",
        "para_confirmar": "para confirmar y reiniciar.",
        "reinicio_ok": "Tesorería reiniciada. Respaldo guardado.",
        "retroceder_historial": "RETROCEDER HISTORIAL (RESTAURAR RESPALDO)",
        "retroceder_explicacion": "Restaura la base de datos a un punto anterior. Use si hubo un error y desea volver atrás sin perder todo.",
        "seleccionar_respaldo": "Seleccione el respaldo a restaurar (más reciente arriba):",
        "restaurar": "Restaurar este respaldo",
        "restaurado_ok": "Respaldo restaurado correctamente.",
        "sin_respaldos": "No hay respaldos disponibles.",
        "tipo_ingreso": "TIPO DE INGRESO",
        "filtros": "FILTROS",
        "fecha_desde": "Desde",
        "fecha_hasta": "Hasta",
        "monto_min": "Monto desde ($)",
        "monto_max": "Monto hasta ($)",
        "filtros_ayuda": "Deje fechas o montos en blanco para no filtrar por ellos.",
        "help_fecha_desde": "Opcional. Solo registros desde esta fecha.",
        "help_fecha_hasta": "Opcional. Solo registros hasta esta fecha.",
        "help_monto_min": "0 = sin mínimo. Filtra por monto del movimiento.",
        "help_monto_max": "0 = sin máximo.",
        "help_filtrar_tipo": "Filtrar por tipo de ingreso o de gasto.",
        "filtrar_tipo": "Tipo",
        "todos_los_tipos": "Todos los tipos",
        "limpiar_fila": "Limpiar esta fila",
        "borrar_btn": "Borrar",
        "borrar_seleccionados": "Borrar seleccionados",
        "borrar_todos": "Borrar todos",
        "seleccionar_todos": "Seleccionar todos",
        "limpiar_todo_tablero": "Limpiar todo el tablero",
        "confirmar_limpiar_todo": "Escriba BORRAR TODO para vaciar la hoja.",
        "maestro_borrar_titulo": "BORRAR REGISTROS (SOLO ACCESO MAESTRO)",
        "maestro_borrar_todo": "BORRAR TODO",
        "maestro_borrar_masa": "Borrar seleccionados",
        "maestro_seleccionar_masa": "Seleccione filas para borrar en masa:",
        "maestro_borrar_detalle": "Borrar por detalle",
        "maestro_seleccionar_detalle": "Escriba texto que contenga el detalle (ej: compra, alquiler). Se borrarán todos los registros cuyo detalle coincida.",
        "maestro_coinciden_registros": "Coinciden {n} registro(s).",
        "maestro_sin_limite": "sin límite de tiempo",
        "solo_30min": "solo dentro de 30 min",
        "arqueo_cero": "El total del arqueo es $0. No se registró.",
        "gasto_cero": "El monto del gasto es $0. No se registró.",
        "sin_resultados_filtro": "No hay registros que coincidan con los filtros aplicados.",
        "quien_usa_app": "Quién usa la app",
        "usuario_actual_menu": "USUARIO ACTUAL",
        "administracion": "ADMINISTRACIÓN",
        "administracion_titulo": "ADMINISTRACIÓN DE PERMISOS",
        "admin_instrucciones": "Marque o desmarque la casilla: **verde** = tiene permiso, **rojo** = no tiene. Un clic da o quita el permiso.",
        "admin_anadir_usuario": "➕ AÑADIR NUEVO USUARIO",
        "admin_id_placeholder": "ID del usuario (ej: asistente)",
        "admin_nombre_placeholder": "Nombre visible (ej: Asistente)",
        "admin_btn_anadir": "Añadir usuario",
        "admin_id_ya_existe": "Ese ID ya existe.",
        "admin_usuario_anadido": "Usuario «{nombre}» añadido. Abajo aparecerá su lista de permisos.",
        "admin_asignar_permisos": "Usuario **{nombre}** añadido. Asigne o quite permisos con las casillas de abajo.",
        "admin_caption_admin": "El administrador siempre ve todo. No se pueden cambiar sus permisos.",
        "admin_caption_verde_rojo": "Verde = puede ver esa sección. Rojo = no puede. Clic en la casilla para cambiar.",
        "admin_responsable_titulo": "Persona responsable (oficial)",
        "admin_responsable_nombre": "Nombre completo",
        "admin_responsable_licencia": "Licencia de conducir o ID",
        "admin_guardar_responsable": "Registrar responsable",
        "admin_responsable_guardado": "Responsable registrado.",
        "admin_contrasena": "Contraseña (opcional)",
        "admin_contrasena_placeholder": "Dejar vacío = usar predeterminada",
        "admin_guardar_contrasena": "Guardar contraseña",
        "admin_contrasena_guardada": "Contraseña guardada.",
        "admin_reset_contrasena": "Restablecer contraseña (solo admin)",
        "admin_pregunta_seguridad_titulo": "Pregunta de seguridad (recuperar contraseña)",
        "admin_pregunta_seguridad": "Pregunta",
        "admin_guardar_pregunta": "Guardar pregunta y respuesta",
        "admin_pregunta_guardada": "Pregunta de seguridad guardada.",
        "admin_respuesta_placeholder": "Solo el admin la conoce",
        "admin_respuesta_requerida": "Escriba la respuesta para guardar.",
        "admin_eliminar_usuario": "🗑️ Eliminar usuario",
        "admin_confirmar_eliminar": "Escriba ELIMINAR para confirmar.",
        "admin_usuario_eliminado": "Usuario «{nombre}» eliminado.",
        "admin_plantilla": "Aplicar perfil",
        "admin_plantilla_ayuda": "Asigna permisos de un perfil predefinido.",
        "admin_copiar_permisos": "Copiar permisos de",
        "admin_copiar_ok": "Permisos copiados de {origen}.",
        "admin_sin_contrasena": "⚠️ Sin contraseña",
        "admin_ultima_actividad": "Última actividad",
        "admin_rastreo_titulo": "RASTREO DE MOVIMIENTOS POR USUARIO",
        "admin_rastreo_sub": "Registro de acciones por usuario (solo clave maestra).",
        "admin_sin_movimientos": "Sin movimientos registrados.",
        "admin_descargar_permisos": "Descargar permisos (JSON)",
        "admin_resumen_permisos": "RESUMEN DE PERMISOS",
        "tema_oscuro": "Tema oscuro",
        "tema_claro": "Tema claro",
        "admin_expander_admin": " — Administrador (todos los permisos)",
        "admin_btn_reiniciar": "Reiniciar tesorería (borrar solo datos)",
        "admin_error_reinicio": "No se pudo completar el reinicio.",
        "admin_debe_escribir": "Debe escribir exactamente «{palabra}» para confirmar.",
        "admin_error_restaurar": "No se pudo restaurar el respaldo.",
        "admin_rastreo_titulo": "Rastreo de movimientos por usuario (solo clave maestra)",
        "admin_rastreo_ayuda": "Últimas acciones registradas por usuario. Solo visible con acceso maestro.",
        "admin_perfil_aplicar": "Aplicar perfil",
        "admin_perfil_tesorero": "Tesorero",
        "admin_perfil_pastor": "Pastor",
        "admin_perfil_asistente": "Asistente",
        "admin_perfil_ministerio_musica": "Ministerio de Música",
        "admin_copiar_permisos_de": "Copiar permisos de",
        "admin_copiar_permisos_btn": "Copiar permisos",
        "admin_permisos_copiados": "Permisos copiados de «{origen}» a «{destino}».",
        "admin_eliminar_usuario": "Eliminar usuario",
        "admin_eliminar_confirmar": "Escriba «ELIMINAR» para confirmar",
        "admin_eliminar_placeholder": "ELIMINAR",
        "admin_usuario_eliminado": "Usuario «{nombre}» eliminado.",
        "admin_no_eliminar_admin": "No se puede eliminar al administrador.",
        "admin_matriz_titulo": "Resumen de permisos (matriz)",
        "admin_sin_contrasena": "Sin contraseña personalizada",
        "admin_quitar_permiso_confirmar": "¿Quitar el permiso «{permiso}» a «{nombre}»?",
        "admin_confirmar": "Confirmar",
        "admin_cancelar": "Cancelar",
        "admin_respaldo_descargar": "Descargar respaldo de permisos",
        "admin_respaldo_restaurar": "Restaurar desde archivo",
        "admin_respaldo_restaurado": "Respaldo restaurado correctamente.",
        "admin_buscar_usuarios": "Buscar usuarios (nombre o ID)",
        "volver_inicio": "Volver al inicio",
        "descargar_pdf_whatsapp": "DESCARGAR PDF PARA WHATSAPP",
        "error_limpiar_tablero": "No se pudo limpiar el tablero.",
        "gasto_foto_no": "GASTO GUARDADO PERO FOTO NO: {e}",
        "imagen_ya_subida_aviso": "⚠️ Esta imagen ya fue subida anteriormente. Evite registrar el mismo comprobante dos veces.",
        "factura_detectada": "Datos detectados en la factura / imagen",
        "total_detectado": "Total detectado",
        "impuesto_detectado": "Impuesto detectado",
        "comercio_detectado": "Comercio / descripción",
        "error_no_se_pudo_guardar": "NO SE PUDO GUARDAR: {e}",
        "error_no_se_pudo_guardar_permisos": "NO SE PUDO GUARDAR PERMISOS: {e}",
        "error_validacion": "Error de validación: {msg}",
        "ver_mas": "Ver más",
        "buscar_facturas_titulo": "BUSCAR FACTURAS / COMPROBANTES",
        "buscar_facturas_ayuda": "Busque por comercio, fecha, monto total o texto extraído de la imagen (OCR).",
        "buscar_por_comercio": "Comercio (contiene)",
        "buscar_por_texto_ocr": "Texto en factura (OCR)",
        "buscar_por_fecha_desde": "Fecha registro desde",
        "buscar_por_fecha_hasta": "Fecha registro hasta",
        "buscar_por_total_min": "Total desde ($)",
        "buscar_por_total_max": "Total hasta ($)",
        "buscar_por_tipo_gasto": "Tipo de gasto",
        "buscar_por_impuesto_desde": "Impuesto desde ($)",
        "buscar_por_impuesto_hasta": "Impuesto hasta ($)",
        "buscar_btn": "Buscar",
        "resultados_facturas": "Resultados",
        "sin_resultados_facturas": "No hay facturas que coincidan con la búsqueda.",
        "sin_facturas": "Aún no hay facturas registradas.",
        "texto_ocr": "Texto OCR",
        "descargar_fotos_zip": "DESCARGAR FOTOS DE GASTOS (ZIP)",
        "medio_pago": "MEDIO DE PAGO",
        "efectivo_arqueo": "Efectivo (arqueo de billetes y monedas)",
        "pos_tarjeta_credito": "POS / Tarjeta de crédito",
        "tarjeta_debito": "Tarjeta de débito",
        "transferencia": "Transferencia bancaria",
        "monto_pos_tarjeta": "Monto ($)",
        "referencia_opcional": "Referencia o últimos 4 dígitos (opcional)",
        "referencia_placeholder": "ej. 1234 o REF-001",
        "pin_ingrese": "Ingrese PIN de administrador",
        "pin_incorrecto": "PIN incorrecto.",
        "entrar": "Entrar",
        "cerrar_sesion": "Cerrar sesión",
        "sesion_cerrada": "Sesión cerrada.",
        "tiempo_inactivo_cerrado": "Sesión cerrada por inactividad. Vuelva a elegir usuario.",
        "version": "Versión",
        "acerca_de": "Acerca del sistema",
        "aprobado_por": "Aprobado por (opcional, gastos > $)",
        "establecer_pin": "Establecer PIN de administrador",
        "cambiar_pin": "Cambiar PIN de administrador",
        "acceso_maestro": "🔑 Acceso maestro",
        "acceso_maestro_info": "Restablecer admin a usuario: admin, contraseña: admin (acceso inicial).",
        "restablecer_admin_admin": "Restablecer a admin/admin",
        "pin_actual": "PIN actual",
        "pin_nuevo": "PIN nuevo",
        "pin_guardado": "PIN guardado.",
        "exportar_contador": "EXPORTAR PARA CONTADOR (EXCEL/CSV)",
        "exportar_excel": "EXPORTAR EXCEL (.xlsx)",
        "integridad_ok": "Integridad del libro comprobada.",
        "integridad_aviso": "Aviso: totales no coinciden. Considere restaurar un respaldo.",
        "dashboard_resumen": "RESUMEN",
        "saldo_actual": "Saldo actual",
        "ingresos_mes": "Ingresos del mes",
        "gastos_mes": "Gastos del mes",
        "grafico_ingresos_gastos": "Ingresos vs Gastos por mes",
        "grafico_resultado": "Resultado (Ingresos − Gastos)",
        "grafico_trazabilidad": "Trazabilidad del saldo (alza/baja en el tiempo)",
        "grafico_saldo": "Saldo",
        "grafico_alza": "Alza",
        "grafico_baja": "Baja",
        "grafico_var": "Variación",
        "grafico_ver_ingresos_gastos": "Ver cuadro Ingresos & Gastos",
        "conciliar": "CONCILIAR INGRESOS",
        "conciliar_ayuda": "Compare el total registrado con lo que contó en caja (opcional).",
        "conciliar_por_dia": "Conciliar por día",
        "fecha_conciliar": "Fecha a conciliar",
        "contado_por": "Contado por",
        "contado_por_ayuda": "Obligatorio. Seleccione o escriba. Se guarda en mayúsculas para futuros arqueos.",
        "verificado_por": "Verificado por",
        "verificado_por_ayuda": "Obligatorio. Solo después de Contado por.",
        "arqueo_llenar_ambos": "Debe llenar Contado por y Verificado por antes de continuar.",
        "arqueo_otro": "Otro (escribir nuevo)",
        "arqueo_sugerencias": "Sugerencias",
        "arqueo_refrescar": "Refrescar",
        "arqueo_refrescar_ayuda": "Vacía el formulario para volver a llenarlo.",
        "arqueo_nombre_invalido": "El nombre no debe contener solo números.",
        "arqueo_nombre_simbolos": "El nombre no debe contener símbolos (@#$%...).",
        "arqueo_nombre_muy_corto": "Ingrese al menos 2 caracteres (evite solo iniciales o una letra).",
        "arqueo_teclado_movil": "En móvil: use el teclado numérico para billetes y monedas.",
        "arqueo_total_calculado": "Total billetes + monedas",
        "sobres_cantidad": "Nº de sobres",
        "sobres_total": "Total en sobres ($)",
        "total_suelto": "Total suelto ($)",
        "cheques_cantidad": "Nº de cheques",
        "cheques_total": "Total cheques ($)",
        "fondo_caja": "Fondo de caja inicial ($)",
        "descargar_hoja_arqueo": "Descargar hoja de arqueo",
        "descargar_hoja_arqueo_ayuda": "Seleccione el arqueo y descargue en PDF o Excel. Archivos comprimidos para enviar por WhatsApp, abren en cualquier dispositivo.",
        "seleccionar_arqueo": "Seleccionar arqueo",
        "hoja_arqueo_pdf": "PDF",
        "hoja_arqueo_excel": "Excel (.xlsx)",
        "resumen_dia": "Resumen del día",
        "cerrar_arqueo": "Cerrar arqueo del día",
        "arqueo_cerrado": "Arqueo cerrado",
        "col_ip": "IP",
        "presupuesto_vs_real": "Presupuesto vs real",
        "primera_vez": "¿Primera vez aquí?",
        "ayuda_rapida": "Guía rápida: use el menú para Inicio, Ministerio de Finanzas o Administración. Los ingresos se registran en Ingresar bendición; los gastos en Registrar gasto. Respaldos automáticos en cada guardado.",
        "tamano_texto": "Tamaño de texto",
        "tamano_normal": "Normal",
        "tamano_grande": "Grande",
        "lo_contado_caja": "Lo que contó en caja ($)",
        "compartir_app": "Compartir app (instalar como Netflix/Disney)",
        "compartir_app_instrucciones": "Puedes enviar el enlace de esta app por WhatsApp. Quien lo abra en el celular puede instalarla como una app: en Chrome/Safari, menú (⋮ o Compartir) → «Añadir a la pantalla de inicio» o «Instalar aplicación». Así se abre como app, sin barra del navegador. El enlace que debes compartir es la URL donde está publicada esta app (ej. tu servidor o Streamlit Cloud). Para que no sea pesada al abrir, la imagen de inicio se adapta al tamaño del celular; si quieres menos peso, usa una imagen de inicio comprimida o reducida en assets.",
        "login_titulo": "United Pentecostal Church International",
        "login_subtitulo": "Inicia sesión para explorar las funciones financieras y herramientas en línea.",
        "diseno_menu": "DISEÑO (MAESTRO)",
        "diseno_titulo": "DISEÑO / LOGOS (CLAVE MAESTRA)",
        "diseno_sub": "Cambie logos, textos principales y visibilidad de botones. Solo para la clave maestra/universal.",
        "diseno_seccion_logos": "Logos e imágenes",
        "diseno_logo_actual": "Logo actual",
        "diseno_logo_subir": "Subir nuevo logo principal (PNG/JPG)",
        "diseno_logo_guardar": "Guardar logo",
        "diseno_textos": "Textos principales (login)",
        "diseno_login_titulo_es": "Título de login en Español",
        "diseno_login_titulo_en": "Título de login en Inglés",
        "diseno_textos_guardar": "Guardar textos",
        "diseno_botones": "Visibilidad de botones del menú",
        "diseno_botones_ayuda": "Marque para ocultar el botón correspondiente para todos los usuarios (incluido administrador).",
        "diseno_ocultar_inicio": "Ocultar botón Inicio",
        "diseno_ocultar_arqueo": "Ocultar botón Arqueo de caja",
        "diseno_ocultar_tesoreria": "Ocultar botón Tesorería",
        "diseno_ocultar_contabilidad": "Ocultar botón Contabilidad",
        "diseno_ocultar_presupuesto": "Ocultar botón Presupuesto y metas",
        "diseno_ocultar_eventos": "Ocultar botón Eventos / Inversiones",
        "diseno_ocultar_admin": "Ocultar botón Administración",
        "diseno_botones_guardar": "Guardar visibilidad de botones",
        "diseno_colores": "Colores de títulos",
        "diseno_color_login_titulo": "Color del título de login",
        "diseno_guardado_ok": "Configuración de diseño guardada.",
        "login_usuario": "Usuario",
        "login_usuario_placeholder": "admin (primera vez: admin/admin)",
        "login_contrasena": "Contraseña",
        "login_mostrar_contrasena": "Ver contraseña",
        "login_btn": "Iniciar sesión",
        "login_error": "Usuario o contraseña incorrectos.",
        "login_error_usuario": "Usuario no encontrado.",
        "login_error_contrasena": "Contraseña incorrecta.",
        "login_bloqueado": "Demasiados intentos. Espere {min} minutos e intente de nuevo.",
        "login_recordar": "Recordar sesión",
        "login_recuperar": "¿Olvidó su contraseña?",
        "login_recuperar_ayuda": "Recuperación por pregunta de seguridad: escriba su usuario y pulse Verificar. Si tiene pregunta y respuesta configuradas, podrá restablecer la contraseña aquí. Si no, el administrador debe configurarlas en Administración → usuario.",
        "login_ultima_actividad": "Última actividad",
        "recuperar_usuario": "Usuario",
        "recuperar_btn_verificar": "Verificar",
        "recuperar_btn_comprobar": "Comprobar",
        "recuperar_btn_restablecer": "Restablecer contraseña",
        "recuperar_respuesta": "Su respuesta",
        "recuperar_nueva_contrasena": "Nueva contraseña",
        "recuperar_confirmar": "Confirmar contraseña",
        "recuperar_volver": "Volver",
        "recuperar_ok": "Contraseña restablecida. Ya puede iniciar sesión.",
        "recuperar_ok_verificado": "Respuesta correcta. Elija nueva contraseña.",
        "recuperar_error_usuario": "Usuario no encontrado.",
        "recuperar_sin_pregunta": "Este usuario no tiene pregunta de seguridad. El administrador debe ir a Administración → usuario y configurar «Pregunta de seguridad» y «Respuesta» para poder recuperar la contraseña desde aquí.",
        "recuperar_error_respuesta": "Respuesta incorrecta.",
        "recuperar_pregunta_1": "¿Cuál es el nombre de su madre?",
        "recuperar_pregunta_2": "¿Cuál es el nombre de su ciudad natal?",
        "recuperar_pregunta_3": "¿Cuál fue el nombre de su primera mascota?",
        "recuperar_pregunta_4": "¿Cuál es su mes de nacimiento?",
        "recuperar_pregunta_5": "¿Últimos 4 dígitos de su teléfono?",
        "recuperar_pregunta_6": "¿Nombre del ministerio o área que más le identifica?",
        "cambiar_credenciales_titulo": "Cambiar contraseña (obligatorio)",
        "cambiar_credenciales_info": "Por seguridad, debe establecer una contraseña antes de continuar. admin/admin ya no funcionará.",
        "cambiar_credenciales_nueva": "Nueva contraseña",
        "cambiar_credenciales_confirmar": "Confirmar contraseña",
        "cambiar_credenciales_guardar": "Guardar y continuar",
        "cambiar_credenciales_vacios": "Complete ambos campos.",
        "cambiar_credenciales_no_coinciden": "Las contraseñas no coinciden.",
        "cambiar_credenciales_ok": "Contraseña guardada. Redirigiendo...",
        "resumen_periodo": "Resumen por período",
        "periodo_mes": "Este mes",
        "periodo_trimestre": "Este trimestre",
        "periodo_ano": "Este año",
        "comparar_periodo_anterior": "Comparar con período anterior",
        "ingresos_periodo": "Ingresos",
        "gastos_periodo": "Gastos",
        "resultado_periodo": "Resultado",
        "saldo_cierre": "Saldo al cierre",
        "vs_anterior": "vs anterior",
        "exportar_resumen_contador": "Exportar resumen para contador",
        "exportar_resumen_ayuda": "Descarga CSV/Excel con totales por mes (ingresos, gastos, saldo) para enviar al contador.",
        "exportar_btn_csv_mes": "CSV (totales por mes)",
        "exportar_btn_excel_mes_tipo": "Excel (por mes y tipo de gasto)",
        "grafico_por_tipo_gasto": "Gastos por tipo en el tiempo",
        "alertas_contabilidad": "Alertas",
        "alerta_sin_movimientos_dias": "Sin movimientos en los últimos {d} días.",
        "filtro_tipo_movimiento": "Tipo de movimiento",
        "solo_ingresos": "Solo ingresos",
        "solo_gastos": "Solo gastos",
        "todos_movimientos": "Todos",
        "orden_hoja": "Orden",
        "orden_recientes": "Más recientes primero",
        "orden_antiguos": "Más antiguos primero",
        "buscar_en_detalle": "Buscar en detalle",
        "buscar_en_detalle_placeholder": "Texto en descripción...",
        "respaldo_hoy": "Descargar respaldo de hoy",
        "respaldo_hoy_ayuda": "Copia completa del libro con fecha de hoy (solo acceso maestro).",
        "integridad_error_titulo": "⚠️ ERROR DE INTEGRIDAD — No use borrar hasta corregir",
        "integridad_banner_titulo": "Error de integridad detectado",
        "integridad_banner_deshabilitado": "El borrado de registros está deshabilitado hasta corregir o restaurar un respaldo.",
        "mantenimiento_titulo": "Sistema en actualización",
        "mantenimiento_aviso": "El administrador está realizando actualizaciones. Evite ingresar datos nuevos mientras vea este aviso. Si cierra y vuelve a abrir, podrá continuar; su trabajo guardado se mantiene visible.",
        "mantenimiento_esperar": "Cuando el administrador desactive el mantenimiento, recargue la página para continuar. Lo que ya guardó no se pierde.",
        "mantenimiento_guardado_visible": "Lo que guardó sigue guardado y lo verá al reanudar el sistema.",
        "mantenimiento_sidebar": "Pausar sistema",
        "mantenimiento_expander": "Pausar sistema (mantenimiento)",
        "mantenimiento_ayuda": "Al activar, los demás usuarios verán un aviso de que el sistema está en actualización. Pueden cerrar y volver; su trabajo guardado se mantiene. Desactive cuando termine el mantenimiento.",
        "mantenimiento_btn_activar": "Pausar sistema (mostrar aviso a usuarios)",
        "mantenimiento_btn_desactivar": "Reanudar sistema",
        "mantenimiento_activo_ahora": "Modo mantenimiento activo.",
        "mantenimiento_activado_ok": "Aviso de mantenimiento activado. Los usuarios verán el mensaje al entrar.",
        "mantenimiento_desactivado_ok": "Sistema reanudado. Los usuarios ya no verán el aviso.",
    },
    "EN": {
        "titulo": "WELLAND PENTECOSTAL CHURCH",
        "subtitulo": "TREASURY SYSTEM",
        "idioma": "LANGUAGE",
        "ingresar_bendicion": "ENTER BLESSING",
        "registrar_gasto": "REGISTER EXPENSE",
        "ver_informe": "VIEW PDF REPORT (WHATSAPP)",
        "ministerio": "MINISTRY",
        "general": "GENERAL",
        "caballeros": "MEN",
        "damas": "WOMEN",
        "jovenes": "YOUTH",
        "ninos": "CHILDREN",
        "musica": "MUSIC",
        "billetes": "BILLS",
        "monedas": "COINS",
        "total": "TOTAL",
        "registrar": "REGISTER",
        "monto": "AMOUNT ($)",
        "descripcion": "DESCRIPTION",
        "tomar_foto": "TAKE PHOTO OR UPLOAD IMAGE",
        "suministros": "Supplies",
        "gastos_frecuentes": "Frequent expenses",
        "gasto_refrescar": "Refresh",
        "gasto_refrescar_ayuda": "Clears the form to fill again.",
        "gasto_sugerencias": "Suggestions",
        "modo_escaneo_rapido": "Quick scan mode (photo + amount only)",
        "ocr_diferencia_aviso": "⚠️ Manual amount (${manual:.2f}) differs from OCR detected (${ocr:.2f}). Please verify.",
        "reintentar_ocr": "Retry OCR",
        "vista_previa_factura": "Preview",
        "foto_multiples_ayuda": "You can upload multiple photos (front, back, attachments).",
        "foto_frente": "Front",
        "foto_reverso": "Back",
        "foto_anexo": "Attachment",
        "ocr_sin_datos": "No information detected. Use «Retry OCR» or enter data manually.",
        "importar_gastos": "Import expenses (CSV/Excel)",
        "importar_gastos_ayuda": "Upload CSV or Excel with columns: fecha, detalle, tipo_gasto, gastos",
        "rendicion_cuentas_titulo": "Accountability (quick reconciliation)",
        "rendicion_cuentas_ayuda": "Fund given for supplies (e.g. $200). Enter each expense (flour $50, sugar $20, taxi $40). The return ($90) is recorded as income. Reconcile: Fund − Spent = Return.",
        "rendicion_responsable": "Responsible (e.g. Sisters)",
        "rendicion_fondo_entregado": "Fund given ($)",
        "rendicion_lineas_gastos": "**Expenses (concept and amount)**",
        "concepto": "Concept",
        "rendicion_devolucion": "Return ($)",
        "rendicion_cuadre_ok": "Reconciliation correct",
        "rendicion_cuadre_revisar": "Check reconciliation",
        "rendicion_registrar_btn": "Register accountability",
        "rendicion_sin_datos": "Enter at least one expense or the return amount.",
        "rendicion_registrada_ok": "Accountability registered. Expenses and return saved.",
        "recordatorios_recurrentes": "Recurring reminders",
        "recordatorios_pendientes": "Supplies that may be pending this month:",
        "recordatorios_todos_ok": "All usual supplies appear to be registered this month.",
        "recordatorios_registre": "Register expenses to see reminders.",
        "gasto_teclado_movil": "On mobile: use numeric keypad for amounts.",
        "gasto_aprobado_requerido": "Expenses ≥ $500 require valid name in Approved by.",
        "gasto_registrado": "EXPENSE REGISTERED.",
        "bendicion_registrada": "BLESSING REGISTERED.",
        "borrar_solo_30min": "DELETION ONLY ALLOWED WITHIN 30 MINUTES.",
        "eliminar": "DELETE",
        "confirmar_eliminar": "Confirm deletion of",
        "cargando": "Loading data...",
        "pagina": "Page",
        "de": "of",
        "registros": "records",
        "anterior": "◀ Previous",
        "siguiente": "Next ▶",
        "coincide_registrado": "✓ Matches registered amount.",
        "saldo": "BALANCE",
        "arqueo": "COUNT",
        "informe_pdf": "PDF REPORT - TREASURY",
        "fecha_informe": "REPORT DATE",
        "direccion": "77 Duncan St, Welland, Ontario, Canada",
        "ingresos": "INCOME",
        "gastos": "EXPENSES",
        "caracteres": "CHARACTERS",
        "sin_movimientos": "NO RECORDED TRANSACTIONS.",
        "menu_navegacion": "NAVIGATION",
        "abrir_menu": "▶ Menu",
        "abrir_menu_ayuda": "Open navigation menu (curtain)",
        "inicio": "HOME",
        "ministerio_finanzas": "FINANCE MINISTRY",
        "sistema_tesoreria": "TREASURY SYSTEM",
        "arqueo_caja": "CASH COUNT",
        "arqueo_caja_sub": "Daily Close — Physical count of the day",
        "tesoreria": "TREASURY",
        "tesoreria_sub": "Ledger — Income and Expenses",
        "contabilidad": "ACCOUNTING",
        "contabilidad_sub": "Historic Vault — Reports and files",
        "presupuesto_metas": "BUDGET & GOALS",
        "presupuesto_metas_sub": "Vision — Project planning",
        "presupuesto_por_tipo": "Budget by expense type",
        "presupuesto_anio": "Year",
        "meta_ingresos_label": "Income goal ($)",
        "presupuesto_guardar": "Save budget",
        "presupuesto_real_titulo": "Actual (this year)",
        "presupuesto_porcentaje": "% used",
        "presupuesto_exportar": "Export budget vs actual",
        "presupuesto_alerta_superado": "Over budget",
        "presupuesto_guardado_ok": "Budget saved.",
        "presupuesto_sin_datos": "No transactions in the period to compare.",
        "presupuesto_total_gastos": "Budget expenses (year)",
        "presupuesto_real_gastos": "Actual expenses (year)",
        "presupuesto_meta_alcanzada": "Income goal reached!",
        "presupuesto_defina_meta": "(set goal above)",
        "eventos_inversiones": "Events / Investments",
        "eventos_inversiones_sub": "Cross of expenses, sales and margin per event. Profitability, donations, labor, report and recommendation.",
        "eventos_buscar": "Search by event or investment name",
        "eventos_buscar_placeholder": "E.g.: Food sale March",
        "eventos_nuevo": "New event / investment",
        "eventos_nombre": "Event or investment name",
        "eventos_fecha": "Date",
        "eventos_gastos": "Expenses ($)",
        "eventos_ventas": "Sales / Income ($)",
        "eventos_margen": "Margin",
        "eventos_rentable": "Profitable",
        "eventos_no_rentable": "Not profitable",
        "eventos_donaciones": "Donations ($)",
        "eventos_perdidas": "Losses ($)",
        "eventos_mano_obra_pagada": "Paid labor ($)",
        "eventos_mano_obra_donada": "Donated labor",
        "eventos_mano_obra_por": "Labor donated by",
        "eventos_por_hermanas": "Sisters",
        "eventos_por_miembros": "Members",
        "eventos_por_ambos": "Sisters and members",
        "eventos_quien_dono": "Who donated labor",
        "eventos_recomendacion": "Recommend repeating?",
        "eventos_recom_si": "Yes, recommended",
        "eventos_recom_no": "No",
        "eventos_recom_tal_vez": "Maybe",
        "eventos_nota": "Note",
        "eventos_nota_placeholder": "Event details",
        "eventos_guardar_btn": "Save event",
        "eventos_nombre_requerido": "Enter the event name.",
        "eventos_guardado_ok": "Event saved.",
        "eventos_listado": "Event list",
        "eventos_sin_eventos": "No events. Create one in «New event / investment».",
        "eventos_cruce": "Cross",
        "eventos_gaste": "Spent",
        "eventos_vendi": "Sold",
        "eventos_btn_registrar_ingreso": "Register income",
        "eventos_btn_registrar_gasto": "Register expense",
        "eventos_informe_titulo": "Event report",
        "eventos_descargar_informe": "Download report",
        "eventos_botones_registrar": "**Register in ledger:**",
        "eventos_si": "Yes",
        "eventos_no": "No",
        "bienvenida_titulo": "WELCOME",
        "bienvenida_texto": "SELECT AN OPTION FROM THE MENU TO CONTINUE.",
        "sistema_operaciones": "OPERATIONS SYSTEM",
        "mision": "MISSION",
        "vision": "VISION",
        "objetivo_supremo": "SUPREME OBJECTIVE",
        "mision_texto": "To bring the light of the Gospel to every creature, transforming lives with love and truth.",
        "vision_texto": "To be a vibrant and global community, where every member is an agent of change in the Kingdom of God.",
        "objetivo_texto": "To empower your divine purpose with tools of excellence, so that your service impacts eternity now.",
        "cerrar": "CLOSE",
        "pastor": "Pastor Javier Escobar",
        "fundamentados": "Founded on the Rock, operating in the Spirit",
        "col_id_registro": "ID REGISTER",
        "col_fecha": "DATE",
        "col_detalle": "DETAIL",
        "col_ingreso": "INCOME",
        "col_gastos": "EXPENSES",
        "col_total_ingresos": "TOTAL INCOME",
        "col_total_gastos": "TOTAL EXPENSES",
        "col_saldo_actual": "CURRENT BALANCE",
        "col_tipo_gasto": "TYPE",
        "tipo_gasto": "EXPENSE TYPE",
        "exportar_hoja_pdf": "EXPORT LEDGER PDF",
        "exportar_hoja_contable_titulo": "Export ledger",
        "exportar_hoja_contable_ayuda": "Download the ledger in your preferred format: PDF for printing, Excel for editing, or CSV for the accountant.",
        "exportar_opcion_pdf": "PDF",
        "exportar_opcion_excel": "Excel (.xlsx)",
        "exportar_opcion_contador": "For accountant (CSV)",
        "informe_auditoria": "FINANCIAL AUDIT REPORT",
        "situacion_actual": "CURRENT SITUATION",
        "superavit": "SURPLUS",
        "deficit": "DEFICIT",
        "alertas_crisis": "CRISIS ALERTS",
        "sin_alertas": "No critical alerts.",
        "alerta_saldo_negativo": "Negative balance: expenses exceed income.",
        "alerta_gastos_altos": "Expenses represent over 90% of income.",
        "alerta_sin_ingresos": "No income recorded in the period.",
        "alerta_saldo_bajo": "Low balance: reduced safety margin.",
        "presupuesto_real": "ACTUAL BUDGET (BY EXPENSE TYPE)",
        "analisis_clinico": "ANALYSIS",
        "riesgo": "RISK ASSESSMENT",
        "riesgo_bajo": "Low",
        "riesgo_medio": "Medium",
        "riesgo_alto": "High",
        "recomendaciones": "RECOMMENDATIONS",
        "ultimos_movimientos": "RECENT TRANSACTIONS",
        "recom_control_gastos": "• Control unnecessary spending and review highest-expense categories.",
        "recom_reserva": "• Maintain a reserve of at least one month of expenses.",
        "recom_revisar": "• Review this report regularly to make timely decisions.",
        "recom_evitar_deficit": "• Avoid letting expenses exceed income on a sustained basis.",
        "alerta_emergencia_titulo": "⚠️ DO NOT SAVE — RISK DETECTED",
        "alerta_emergencia_texto": "Invalid data or possible tampering detected. Do not save until corrected. Data integrity may be at risk.",
        "reiniciar_tesoreria": "RESET TREASURY (MASTER ONLY)",
        "reiniciar_explicacion": "Deletes only recorded data. Does not delete formulas or structures. A copy is saved in the resets folder.",
        "confirmar_escribir": "Type exactly",
        "para_confirmar": "to confirm and reset.",
        "reinicio_ok": "Treasury reset. Backup saved.",
        "retroceder_historial": "ROLLBACK HISTORY (RESTORE BACKUP)",
        "retroceder_explicacion": "Restores the database to a previous point. Use if there was an error and you want to go back without losing everything.",
        "seleccionar_respaldo": "Select the backup to restore (newest first):",
        "restaurar": "Restore this backup",
        "restaurado_ok": "Backup restored successfully.",
        "sin_respaldos": "No backups available.",
        "tipo_ingreso": "INCOME TYPE",
        "filtros": "FILTERS",
        "fecha_desde": "From",
        "fecha_hasta": "To",
        "monto_min": "Amount from ($)",
        "monto_max": "Amount to ($)",
        "filtros_ayuda": "Leave dates or amounts blank to show all.",
        "help_fecha_desde": "Optional. Only records from this date.",
        "help_fecha_hasta": "Optional. Only records until this date.",
        "help_monto_min": "0 = no minimum. Filter by movement amount.",
        "help_monto_max": "0 = no maximum.",
        "help_filtrar_tipo": "Filter by income or expense type.",
        "filtrar_tipo": "Type",
        "todos_los_tipos": "All types",
        "limpiar_fila": "Clear this row",
        "borrar_btn": "Delete",
        "borrar_seleccionados": "Delete selected",
        "borrar_todos": "Delete all",
        "seleccionar_todos": "Select all",
        "limpiar_todo_tablero": "Clear entire table",
        "confirmar_limpiar_todo": "Type CLEAR ALL to empty the ledger.",
        "maestro_borrar_titulo": "DELETE RECORDS (MASTER ACCESS ONLY)",
        "maestro_borrar_todo": "DELETE ALL",
        "maestro_borrar_masa": "Delete selected",
        "maestro_seleccionar_masa": "Select rows to delete in bulk:",
        "maestro_borrar_detalle": "Delete by detail",
        "maestro_seleccionar_detalle": "Type text contained in the detail (e.g. purchase, rent). All records whose detail matches will be deleted.",
        "maestro_coinciden_registros": "{n} record(s) match.",
        "maestro_sin_limite": "no time limit",
        "solo_30min": "only within 30 min",
        "arqueo_cero": "Count total is $0. Not recorded.",
        "gasto_cero": "Expense amount is $0. Not recorded.",
        "sin_resultados_filtro": "No records match the current filters.",
        "quien_usa_app": "Who is using the app",
        "usuario_actual_menu": "CURRENT USER",
        "administracion": "ADMINISTRATION",
        "administracion_titulo": "PERMISSIONS ADMINISTRATION",
        "admin_instrucciones": "Check or uncheck the box: **green** = has permission, **red** = does not. One click toggles.",
        "admin_anadir_usuario": "➕ ADD NEW USER",
        "admin_id_placeholder": "User ID (e.g. assistant)",
        "admin_nombre_placeholder": "Display name (e.g. Assistant)",
        "admin_btn_anadir": "Add user",
        "admin_id_ya_existe": "That ID already exists.",
        "admin_usuario_anadido": "User «{nombre}» added. Their permission list will appear below.",
        "admin_asignar_permisos": "User **{nombre}** added. Assign or remove permissions with the checkboxes below.",
        "admin_caption_admin": "The administrator always sees everything. Their permissions cannot be changed.",
        "admin_caption_verde_rojo": "Green = can see that section. Red = cannot. Click the box to change.",
        "admin_responsable_titulo": "Responsible person (official)",
        "admin_responsable_nombre": "Full name",
        "admin_responsable_licencia": "Driver's license or ID",
        "admin_guardar_responsable": "Register responsible",
        "admin_responsable_guardado": "Responsible person registered.",
        "admin_contrasena": "Password (optional)",
        "admin_contrasena_placeholder": "Leave empty = use default",
        "admin_guardar_contrasena": "Save password",
        "admin_contrasena_guardada": "Password saved.",
        "admin_reset_contrasena": "Reset password (admin only)",
        "admin_pregunta_seguridad_titulo": "Security question (password recovery)",
        "admin_pregunta_seguridad": "Question",
        "admin_guardar_pregunta": "Save question and answer",
        "admin_pregunta_guardada": "Security question saved.",
        "admin_respuesta_placeholder": "Only admin knows it",
        "admin_respuesta_requerida": "Enter the answer to save.",
        "admin_eliminar_usuario": "🗑️ Delete user",
        "admin_confirmar_eliminar": "Type DELETE to confirm.",
        "admin_usuario_eliminado": "User '{nombre}' deleted.",
        "admin_plantilla": "Apply profile",
        "admin_plantilla_ayuda": "Assigns permissions from a predefined profile.",
        "admin_copiar_permisos": "Copy permissions from",
        "admin_copiar_ok": "Permissions copied from {origen}.",
        "admin_sin_contrasena": "⚠️ No password",
        "admin_ultima_actividad": "Last activity",
        "admin_rastreo_titulo": "USER ACTIVITY TRACKING",
        "admin_rastreo_sub": "Action log by user (master key only).",
        "admin_sin_movimientos": "No recorded activity.",
        "admin_descargar_permisos": "Download permissions (JSON)",
        "admin_resumen_permisos": "PERMISSIONS SUMMARY",
        "tema_oscuro": "Dark theme",
        "tema_claro": "Light theme",
        "admin_expander_admin": " — Administrator (all permissions)",
        "admin_btn_reiniciar": "Reset treasury (delete data only)",
        "admin_error_reinicio": "Reset could not be completed.",
        "admin_debe_escribir": "You must type exactly «{palabra}» to confirm.",
        "admin_error_restaurar": "Backup could not be restored.",
        "admin_rastreo_titulo": "User activity tracking (master key only)",
        "admin_rastreo_ayuda": "Latest actions recorded per user. Only visible with master access.",
        "admin_perfil_aplicar": "Apply profile",
        "admin_perfil_tesorero": "Treasurer",
        "admin_perfil_pastor": "Pastor",
        "admin_perfil_asistente": "Assistant",
        "admin_perfil_ministerio_musica": "Music Ministry",
        "admin_copiar_permisos_de": "Copy permissions from",
        "admin_copiar_permisos_btn": "Copy permissions",
        "admin_permisos_copiados": "Permissions copied from «{origen}» to «{destino}».",
        "admin_eliminar_usuario": "Delete user",
        "admin_eliminar_confirmar": "Type «DELETE» to confirm",
        "admin_eliminar_placeholder": "DELETE",
        "admin_usuario_eliminado": "User «{nombre}» deleted.",
        "admin_no_eliminar_admin": "Administrator cannot be deleted.",
        "admin_matriz_titulo": "Permissions summary (matrix)",
        "admin_sin_contrasena": "No custom password",
        "admin_quitar_permiso_confirmar": "Remove permission «{permiso}» from «{nombre}»?",
        "admin_confirmar": "Confirm",
        "admin_cancelar": "Cancel",
        "admin_respaldo_descargar": "Download permissions backup",
        "admin_respaldo_restaurar": "Restore from file",
        "admin_respaldo_restaurado": "Backup restored successfully.",
        "admin_buscar_usuarios": "Search users (name or ID)",
        "volver_inicio": "Back to home",
        "descargar_pdf_whatsapp": "DOWNLOAD PDF FOR WHATSAPP",
        "error_limpiar_tablero": "Table could not be cleared.",
        "gasto_foto_no": "EXPENSE SAVED BUT PHOTO NOT: {e}",
        "imagen_ya_subida_aviso": "⚠️ This image was already uploaded before. Avoid registering the same receipt twice.",
        "factura_detectada": "Data detected from receipt / image",
        "total_detectado": "Total detected",
        "impuesto_detectado": "Tax detected",
        "comercio_detectado": "Merchant / description",
        "error_no_se_pudo_guardar": "COULD NOT SAVE: {e}",
        "error_no_se_pudo_guardar_permisos": "COULD NOT SAVE PERMISSIONS: {e}",
        "error_validacion": "Validation error: {msg}",
        "ver_mas": "See more",
        "buscar_facturas_titulo": "SEARCH RECEIPTS / INVOICES",
        "buscar_facturas_ayuda": "Search by merchant, date, total amount, or text extracted from the image (OCR).",
        "buscar_por_comercio": "Merchant (contains)",
        "buscar_por_texto_ocr": "Text in receipt (OCR)",
        "buscar_por_fecha_desde": "Record date from",
        "buscar_por_fecha_hasta": "Record date until",
        "buscar_por_total_min": "Total from ($)",
        "buscar_por_total_max": "Total to ($)",
        "buscar_por_tipo_gasto": "Expense type",
        "buscar_por_impuesto_desde": "Tax from ($)",
        "buscar_por_impuesto_hasta": "Tax to ($)",
        "buscar_btn": "Search",
        "resultados_facturas": "Results",
        "sin_resultados_facturas": "No receipts match the search.",
        "sin_facturas": "No receipts registered yet.",
        "texto_ocr": "OCR text",
        "descargar_fotos_zip": "DOWNLOAD EXPENSE PHOTOS (ZIP)",
        "medio_pago": "PAYMENT METHOD",
        "efectivo_arqueo": "Cash (bills and coins count)",
        "pos_tarjeta_credito": "POS / Credit card",
        "tarjeta_debito": "Debit card",
        "transferencia": "Bank transfer",
        "monto_pos_tarjeta": "Amount ($)",
        "referencia_opcional": "Reference or last 4 digits (optional)",
        "referencia_placeholder": "e.g. 1234 or REF-001",
        "pin_ingrese": "Enter administrator PIN",
        "pin_incorrecto": "Incorrect PIN.",
        "entrar": "Enter",
        "cerrar_sesion": "Log out",
        "sesion_cerrada": "Logged out.",
        "tiempo_inactivo_cerrado": "Session closed due to inactivity. Please select user again.",
        "version": "Version",
        "acerca_de": "About this system",
        "aprobado_por": "Approved by (optional, expenses > $)",
        "establecer_pin": "Set administrator PIN",
        "cambiar_pin": "Change administrator PIN",
        "acceso_maestro": "🔑 Master access",
        "acceso_maestro_info": "Reset admin to username: admin, password: admin (initial access).",
        "restablecer_admin_admin": "Reset to admin/admin",
        "pin_actual": "Current PIN",
        "pin_nuevo": "New PIN",
        "pin_guardado": "PIN saved.",
        "exportar_contador": "EXPORT FOR ACCOUNTANT (EXCEL/CSV)",
        "exportar_excel": "EXPORT EXCEL (.xlsx)",
        "integridad_ok": "Ledger integrity verified.",
        "integridad_aviso": "Warning: totals do not match. Consider restoring a backup.",
        "dashboard_resumen": "SUMMARY",
        "saldo_actual": "Current balance",
        "ingresos_mes": "Income this month",
        "gastos_mes": "Expenses this month",
        "grafico_ingresos_gastos": "Income vs Expenses by month",
        "grafico_resultado": "Result (Income − Expenses)",
        "grafico_trazabilidad": "Balance traceability (rise/fall over time)",
        "grafico_saldo": "Balance",
        "grafico_alza": "Rise",
        "grafico_baja": "Fall",
        "grafico_var": "Change",
        "grafico_ver_ingresos_gastos": "Show Income & Expenses chart",
        "conciliar": "RECONCILE INCOME",
        "conciliar_ayuda": "Compare registered total with what you counted in cash (optional).",
        "conciliar_por_dia": "Reconcile by day",
        "fecha_conciliar": "Date to reconcile",
        "contado_por": "Counted by",
        "contado_por_ayuda": "Required. Select or type. Saved in uppercase for future counts.",
        "verificado_por": "Verified by",
        "verificado_por_ayuda": "Required. Only after Counted by.",
        "arqueo_llenar_ambos": "You must fill Counted by and Verified by before continuing.",
        "arqueo_otro": "Other (type new)",
        "arqueo_sugerencias": "Suggestions",
        "arqueo_refrescar": "Refresh",
        "arqueo_refrescar_ayuda": "Clears the form to fill again.",
        "arqueo_nombre_invalido": "Name should not contain only numbers.",
        "arqueo_nombre_simbolos": "Name should not contain symbols (@#$%...).",
        "arqueo_nombre_muy_corto": "Enter at least 2 characters (avoid single letter or initial only).",
        "arqueo_teclado_movil": "On mobile: use numeric keyboard for bills and coins.",
        "arqueo_total_calculado": "Total bills + coins",
        "sobres_cantidad": "No. of envelopes",
        "sobres_total": "Total in envelopes ($)",
        "total_suelto": "Loose total ($)",
        "cheques_cantidad": "No. of checks",
        "cheques_total": "Total checks ($)",
        "fondo_caja": "Opening cash fund ($)",
        "descargar_hoja_arqueo": "Download count sheet",
        "descargar_hoja_arqueo_ayuda": "Select the count and download in PDF or Excel. Compressed files for WhatsApp, open on any device.",
        "seleccionar_arqueo": "Select count",
        "hoja_arqueo_pdf": "PDF",
        "hoja_arqueo_excel": "Excel (.xlsx)",
        "resumen_dia": "Daily summary",
        "cerrar_arqueo": "Close day count",
        "arqueo_cerrado": "Count closed",
        "col_ip": "IP",
        "presupuesto_vs_real": "Budget vs actual",
        "primera_vez": "First time here?",
        "ayuda_rapida": "Quick guide: use the menu for Home, Finance Ministry or Administration. Income is entered in Enter blessing; expenses in Register expense. Automatic backups on each save.",
        "tamano_texto": "Text size",
        "tamano_normal": "Normal",
        "tamano_grande": "Large",
        "lo_contado_caja": "What you counted in cash ($)",
        "compartir_app": "Share app (install like Netflix/Disney)",
        "compartir_app_instrucciones": "You can send this app's link via WhatsApp. Whoever opens it on their phone can install it as an app: in Chrome/Safari, menu (⋮ or Share) → «Add to Home Screen» or «Install app». It will open like an app, without the browser bar. The link to share is the URL where this app is published (e.g. your server or Streamlit Cloud). To keep it light, the home image is sized for the phone; use a compressed or smaller image in assets if you want faster loading.",
        "login_titulo": "United Pentecostal Church International",
        "login_subtitulo": "Log in to explore financial functions and online tools.",
        "diseno_menu": "DESIGN (MASTER)",
        "diseno_titulo": "DESIGN / LOGOS (MASTER KEY)",
        "diseno_sub": "Change logos, main texts and menu button visibility. Only for the master/universal key.",
        "diseno_seccion_logos": "Logos and images",
        "diseno_logo_actual": "Current logo",
        "diseno_logo_subir": "Upload new main logo (PNG/JPG)",
        "diseno_logo_guardar": "Save logo",
        "diseno_textos": "Main texts (login)",
        "diseno_login_titulo_es": "Login title in Spanish",
        "diseno_login_titulo_en": "Login title in English",
        "diseno_textos_guardar": "Save texts",
        "diseno_botones": "Menu button visibility",
        "diseno_botones_ayuda": "Check to hide the corresponding button for all users (including administrator).",
        "diseno_ocultar_inicio": "Hide Home button",
        "diseno_ocultar_arqueo": "Hide Cash count button",
        "diseno_ocultar_tesoreria": "Hide Treasury button",
        "diseno_ocultar_contabilidad": "Hide Accounting button",
        "diseno_ocultar_presupuesto": "Hide Budget & Goals button",
        "diseno_ocultar_eventos": "Hide Events / Investments button",
        "diseno_ocultar_admin": "Hide Administration button",
        "diseno_botones_guardar": "Save button visibility",
        "diseno_colores": "Title colors",
        "diseno_color_login_titulo": "Login title color",
        "diseno_guardado_ok": "Design configuration saved.",
        "login_usuario": "Username",
        "login_usuario_placeholder": "admin (first time: admin/admin)",
        "login_contrasena": "Password",
        "login_mostrar_contrasena": "Show password",
        "login_btn": "Log In",
        "login_error": "Incorrect username or password.",
        "login_error_usuario": "User not found.",
        "login_error_contrasena": "Incorrect password.",
        "login_bloqueado": "Too many attempts. Wait {min} minutes and try again.",
        "login_recordar": "Remember session",
        "login_recuperar": "Forgot your password?",
        "login_recuperar_ayuda": "Recovery by security question: enter your username and click Verify. If you have a question and answer set, you can reset your password here. Otherwise, the administrator must set them in Administration → user.",
        "login_ultima_actividad": "Last activity",
        "recuperar_usuario": "Username",
        "recuperar_btn_verificar": "Verify",
        "recuperar_btn_comprobar": "Check",
        "recuperar_btn_restablecer": "Reset password",
        "recuperar_respuesta": "Your answer",
        "recuperar_nueva_contrasena": "New password",
        "recuperar_confirmar": "Confirm password",
        "recuperar_volver": "Back",
        "recuperar_ok": "Password reset. You can now log in.",
        "recuperar_ok_verificado": "Correct answer. Choose a new password.",
        "recuperar_error_usuario": "User not found.",
        "recuperar_sin_pregunta": "This user has no security question. The administrator must go to Administration → user and set «Security question» and «Answer» so the password can be recovered here.",
        "recuperar_error_respuesta": "Incorrect answer.",
        "recuperar_pregunta_1": "What is your mother's name?",
        "recuperar_pregunta_2": "What is the name of your hometown?",
        "recuperar_pregunta_3": "What was your first pet's name?",
        "recuperar_pregunta_4": "What is your month of birth?",
        "recuperar_pregunta_5": "Last 4 digits of your phone?",
        "recuperar_pregunta_6": "Name of the ministry or area that best identifies you?",
        "cambiar_credenciales_titulo": "Change password (required)",
        "cambiar_credenciales_info": "For security, you must set a password before continuing. admin/admin will no longer work.",
        "cambiar_credenciales_nueva": "New password",
        "cambiar_credenciales_confirmar": "Confirm password",
        "cambiar_credenciales_guardar": "Save and continue",
        "cambiar_credenciales_vacios": "Fill in both fields.",
        "cambiar_credenciales_no_coinciden": "Passwords do not match.",
        "cambiar_credenciales_ok": "Password saved. Redirecting...",
        "resumen_periodo": "Summary by period",
        "periodo_mes": "This month",
        "periodo_trimestre": "This quarter",
        "periodo_ano": "This year",
        "comparar_periodo_anterior": "Compare with previous period",
        "ingresos_periodo": "Income",
        "gastos_periodo": "Expenses",
        "resultado_periodo": "Result",
        "saldo_cierre": "Balance at close",
        "vs_anterior": "vs previous",
        "exportar_resumen_contador": "Export summary for accountant",
        "exportar_resumen_ayuda": "Download CSV/Excel with monthly totals (income, expenses, balance) to send to accountant.",
        "exportar_btn_csv_mes": "CSV (totals by month)",
        "exportar_btn_excel_mes_tipo": "Excel (by month and expense type)",
        "grafico_por_tipo_gasto": "Expenses by type over time",
        "alertas_contabilidad": "Alerts",
        "alerta_sin_movimientos_dias": "No transactions in the last {d} days.",
        "filtro_tipo_movimiento": "Movement type",
        "solo_ingresos": "Income only",
        "solo_gastos": "Expenses only",
        "todos_movimientos": "All",
        "orden_hoja": "Order",
        "orden_recientes": "Newest first",
        "orden_antiguos": "Oldest first",
        "buscar_en_detalle": "Search in detail",
        "buscar_en_detalle_placeholder": "Text in description...",
        "respaldo_hoy": "Download today's backup",
        "respaldo_hoy_ayuda": "Full ledger copy with today's date (master access only).",
        "integridad_error_titulo": "⚠️ INTEGRITY ERROR — Do not delete until fixed",
        "integridad_banner_titulo": "Integrity error detected",
        "integridad_banner_deshabilitado": "Record deletion is disabled until you fix or restore a backup.",
        "mantenimiento_titulo": "System under maintenance",
        "mantenimiento_aviso": "The administrator is performing updates. Avoid entering new data while you see this notice. If you close and reopen, you can continue; your saved work remains visible.",
        "mantenimiento_esperar": "When the administrator turns off maintenance, reload the page to continue. Your saved work is not lost.",
        "mantenimiento_guardado_visible": "What you saved remains saved and you will see it when the system resumes.",
        "mantenimiento_sidebar": "Pause system",
        "mantenimiento_expander": "Pause system (maintenance)",
        "mantenimiento_ayuda": "When enabled, other users will see a notice that the system is being updated. They can close and return; their saved work is preserved. Disable when maintenance is complete.",
        "mantenimiento_btn_activar": "Pause system (show notice to users)",
        "mantenimiento_btn_desactivar": "Resume system",
        "mantenimiento_activo_ahora": "Maintenance mode active.",
        "mantenimiento_activado_ok": "Maintenance notice enabled. Users will see the message when they open the app.",
        "mantenimiento_desactivado_ok": "System resumed. Users will no longer see the notice.",
    },
}

MINISTERIOS = ["GENERAL", "CABALLEROS", "DAMAS", "JÓVENES", "NIÑOS", "MÚSICA"]
MINISTERIOS_EN = ["GENERAL", "MEN", "WOMEN", "YOUTH", "CHILDREN", "MUSIC"]

# Tipos de gasto universales y modernos (ES label, EN label). Se guarda el valor ES en CSV.
# Recurrentes = primero (suministros por pagar: luz, agua, etc.). Operativo = gastos frecuentes del historial.
TIPOS_GASTO_ES = [
    "Recurrentes", "Operativo", "Mantenimiento", "Servicios", "Programas",
    "Equipo y tecnología", "Donaciones y ofrendas", "Otros"
]
TIPOS_GASTO_EN = [
    "Recurring", "Operational", "Maintenance", "Utilities", "Programs",
    "Equipment & tech", "Donations & offerings", "Other"
]
DEFAULT_TIPO_GASTO = "Otros"  # para registros antiguos o sin tipo

# Tipos de ingreso universales para iglesia (se guardan en columna tipo_gasto para ingresos)
TIPOS_INGRESO_ES = [
    "Ofrenda del culto", "Diezmo", "Arqueo de caja (ministerio)", "Donación designada",
    "Evento o actividad", "Venta / Kiosco / Cafetería", "Venta comida y gaseosas", "Ofrenda misionera", "Devolución de fondo", "Otros ingresos"
]
TIPOS_INGRESO_EN = [
    "Service offering", "Tithe", "Cash count (ministry)", "Designated donation",
    "Event or activity", "Sale / Kiosk / Café", "Food and soda sales", "Missions offering", "Fund return", "Other income"
]
DEFAULT_TIPO_INGRESO = "Ofrenda del culto"

# Medios de pago para ingresos (diezmos, ofrendas por POS/tarjeta/transferencia)
MEDIOS_PAGO_ES = ["Efectivo (arqueo de billetes y monedas)", "POS / Tarjeta de crédito", "Tarjeta de débito", "Transferencia bancaria", "Cheques"]
MEDIOS_PAGO_EN = ["Cash (bills and coins count)", "POS / Credit card", "Debit card", "Bank transfer", "Checks"]
MEDIO_EFECTIVO_ES = "Efectivo (arqueo de billetes y monedas)"
MEDIO_EFECTIVO_EN = "Cash (bills and coins count)"

# ============== HOJA CONTABLE (LIBRO DE CUENTAS) ==============
# Columnas: id_registro, fecha, detalle, tipo_gasto, ingreso, gastos, total_ingresos, total_gastos, saldo_actual
COLUMNAS_LEDGER = [
    "id_registro", "fecha", "detalle", "tipo_gasto", "ingreso", "gastos",
    "total_ingresos", "total_gastos", "saldo_actual"
]

def generar_id():
    """ID legacy: T-YYYYMMDDHHMMSS."""
    return datetime.now().strftime("T-%Y%m%d%H%M%S")

def generar_id_gasto():
    """ID único para registro de gastos: RG-AÑO-MES-DÍA-HORA-MINUTO-SEGUNDO."""
    return datetime.now().strftime("RG-%Y-%m-%d-%H-%M-%S")

def generar_id_arqueo():
    """ID único para arqueo de caja: AC-AÑO-MES-DÍA-HORA-MINUTO-SEGUNDO."""
    return datetime.now().strftime("AC-%Y-%m-%d-%H-%M-%S")

def _recalcular_totales_ledger(df):
    """Recalcula total_ingresos, total_gastos y saldo_actual por fila (acumulados). Montos siempre a 2 decimales."""
    if df.empty:
        return df
    ing = pd.to_numeric(df["ingreso"], errors="coerce").fillna(0).round(2)
    gas = pd.to_numeric(df["gastos"], errors="coerce").fillna(0).round(2)
    df = df.copy()
    df["total_ingresos"] = ing.cumsum()
    df["total_gastos"] = gas.cumsum()
    df["saldo_actual"] = df["total_ingresos"] - df["total_gastos"]
    return df

@st.cache_data(ttl=10)
def cargar_db():
    if not os.path.exists(DB_ARCHIVO):
        return pd.DataFrame(columns=COLUMNAS_LEDGER)
    try:
        df = pd.read_csv(DB_ARCHIVO, encoding="utf-8")
    except Exception:
        return pd.DataFrame(columns=COLUMNAS_LEDGER)
    # Si es formato antiguo, migrar a ledger
    if "id_registro" not in df.columns and "id" in df.columns:
        filas = []
        for _, r in df.iterrows():
            fecha = r.get("fecha_hora", r.get("fecha", ""))
            tid = r.get("id", "")
            tipo = str(r.get("tipo", "")).upper()
            try:
                monto = float(r.get("monto", 0) or 0)
            except (TypeError, ValueError):
                monto = 0.0
            if "BENDICIÓN" in tipo or "BENDICION" in tipo:
                detalle = str(r.get("ministerio", "Arqueo"))[:200]
                filas.append({"id_registro": tid, "fecha": fecha, "detalle": detalle, "tipo_gasto": "", "ingreso": monto, "gastos": 0})
            else:
                detalle = str(r.get("descripcion", "Gasto"))[:200]
                filas.append({"id_registro": tid, "fecha": fecha, "detalle": detalle, "tipo_gasto": DEFAULT_TIPO_GASTO, "ingreso": 0, "gastos": monto})
        df = pd.DataFrame(filas)
        df = _recalcular_totales_ledger(df)
    # Asegurar columna tipo_gasto en CSVs existentes sin ella
    if not df.empty and "tipo_gasto" not in df.columns:
        def _gastos_val(r):
            try:
                return float(r.get("gastos", 0) or 0) > 0
            except (TypeError, ValueError):
                return False
        df["tipo_gasto"] = df.apply(lambda r: DEFAULT_TIPO_GASTO if _gastos_val(r) else "", axis=1)
    if not df.empty and "ingreso" in df.columns:
        df = _recalcular_totales_ledger(df)
    return df


def verificar_integridad_ledger(df):
    """Comprueba que los totales acumulados de la última fila coincidan con la suma de movimientos. Devuelve (True, None) o (False, mensaje)."""
    if df is None or df.empty or "ingreso" not in df.columns or "gastos" not in df.columns:
        return True, None
    ing = pd.to_numeric(df["ingreso"], errors="coerce").fillna(0).round(2)
    gas = pd.to_numeric(df["gastos"], errors="coerce").fillna(0).round(2)
    sum_ing = round(ing.sum(), 2)
    sum_gas = round(gas.sum(), 2)
    ultima = df.iloc[-1]
    try:
        ti = round(float(ultima.get("total_ingresos", 0) or 0), 2)
        tg = round(float(ultima.get("total_gastos", 0) or 0), 2)
        sa = round(float(ultima.get("saldo_actual", 0) or 0), 2)
    except (TypeError, ValueError):
        return False, "Totales no numéricos en última fila."
    if abs(ti - sum_ing) > 0.02 or abs(tg - sum_gas) > 0.02:
        return False, f"Total ingresos esperado {sum_ing:.2f} vs {ti:.2f}; total gastos {sum_gas:.2f} vs {tg:.2f}."
    if abs(sa - (ti - tg)) > 0.02:
        return False, f"Saldo no coincide con total_ingresos - total_gastos."
    return True, None


def validar_datos(df):
    """Valida antes de guardar. Devuelve (True, None) o (False, mensaje_error). Detecta datos inválidos o posible manipulación."""
    if df is None or not isinstance(df, pd.DataFrame):
        return False, "Datos no válidos."
    for col in COLUMNAS_LEDGER:
        if col not in df.columns:
            return False, f"Falta la columna requerida: {col}"
    # Montos no negativos (conversión segura por si hay texto en CSV)
    for col in ["ingreso", "gastos", "total_ingresos", "total_gastos"]:
        if col in df.columns:
            num = pd.to_numeric(df[col], errors="coerce").fillna(0)
            if (num < -0.001).any():
                return False, f"Hay valores negativos no permitidos en {col}."
    # Detalle sin caracteres peligrosos (script, inyección)
    if "detalle" in df.columns:
        for v in df["detalle"].dropna().astype(str):
            if len(v) > 2000 or "<script" in v.lower() or "javascript:" in v.lower() or "<?php" in v.lower():
                return False, "Detalle con contenido no permitido o demasiado largo."
    # id_registro formato razonable
    if "id_registro" in df.columns:
        for v in df["id_registro"].dropna().astype(str):
            if len(v) > 80 or "\n" in v or "\r" in v:
                return False, "ID de registro con formato no permitido."
    # tipo_gasto longitud razonable (evitar abusos)
    if "tipo_gasto" in df.columns:
        for v in df["tipo_gasto"].dropna().astype(str):
            if len(v.strip()) > 120:
                return False, "Tipo de gasto/ingreso con texto demasiado largo."
    return True, None

def _respaldo_automatico():
    """Copia el CSV actual al historial con timestamp. Mantiene solo los últimos MAX_RESPALDOS."""
    if not os.path.exists(DB_ARCHIVO):
        return
    os.makedirs(CARPETA_HISTORIAL, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(CARPETA_HISTORIAL, f"DB_{timestamp}.csv")
    try:
        shutil.copy2(DB_ARCHIVO, destino)
    except Exception:
        pass
    # Mantener solo los últimos MAX_RESPALDOS (solo archivos .csv)
    try:
        listados = sorted(
            [f for f in os.listdir(CARPETA_HISTORIAL) if f.endswith(".csv")],
            reverse=True
        )
        for f in listados[MAX_RESPALDOS:]:
            ruta = os.path.join(CARPETA_HISTORIAL, f)
            if os.path.isfile(ruta):
                os.remove(ruta)
    except Exception:
        pass

def listar_respaldos():
    """Lista archivos de respaldo en historial, más recientes primero. Devuelve lista de (ruta, etiqueta)."""
    if not os.path.isdir(CARPETA_HISTORIAL):
        return []
    out = []
    for f in sorted(os.listdir(CARPETA_HISTORIAL), reverse=True):
        if f.endswith(".csv"):
            ruta = os.path.join(CARPETA_HISTORIAL, f)
            if not os.path.isfile(ruta):
                continue
            try:
                mtime = os.path.getmtime(ruta)
                etiqueta = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                etiqueta = f
            out.append((ruta, etiqueta))
    return out[:MAX_RESPALDOS]

def restaurar_respaldo(ruta):
    """Restaura la base de datos desde un archivo de respaldo."""
    if not os.path.isfile(ruta) or not ruta.endswith(".csv"):
        return False
    try:
        shutil.copy2(ruta, DB_ARCHIVO)
        cargar_db.clear()
        return True
    except Exception:
        return False

def reiniciar_tesoreria_master():
    """Reinicia la tesorería: respalda el CSV actual en CARPETA_RESETS y deja el ledger vacío (solo estructura)."""
    os.makedirs(CARPETA_RESETS, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    destino = os.path.join(CARPETA_RESETS, f"DB_antes_reinicio_{timestamp}.csv")
    try:
        if os.path.exists(DB_ARCHIVO):
            shutil.copy2(DB_ARCHIVO, destino)
    except Exception:
        pass
    df_vacio = pd.DataFrame(columns=COLUMNAS_LEDGER)
    try:
        df_vacio.to_csv(DB_ARCHIVO, index=False, encoding="utf-8")
        cargar_db.clear()
        return True
    except Exception:
        return False

def guardar_db(df, t=None):
    """Guarda solo si los datos pasan validación. Crea respaldo automático tras guardar.
    t: diccionario de textos (TEXTOS[lang]) para mensajes traducidos; si es None se usa español."""
    df = _recalcular_totales_ledger(df) if not df.empty else df
    ok, mensaje = validar_datos(df)
    if not ok:
        st.error(t["error_validacion"].format(msg=mensaje) if t else mensaje)
        return False
    try:
        df.to_csv(DB_ARCHIVO, index=False, encoding="utf-8")
        _respaldo_automatico()
        cargar_db.clear()
        return True
    except Exception as e:
        st.error(t["error_no_se_pudo_guardar"].format(e=str(e)) if t else f"NO SE PUDO GUARDAR: {e}")
        return False


def _get_client_ip():
    """Obtiene la IP del dispositivo cliente. Solo visible para acceso maestro."""
    try:
        headers = getattr(st, "context", None) and getattr(st.context, "headers", None)
        if headers and isinstance(headers, dict):
            return (headers.get("X-Forwarded-For") or headers.get("X-Real-Ip") or "").split(",")[0].strip() or "N/A"
    except Exception:
        pass
    return "N/A"


def cargar_arqueo_meta():
    """Carga metadatos de arqueos (contado_por, verificado_por, ip, desglose, etc.)."""
    if not os.path.exists(DB_ARQUEO_META):
        return {}
    try:
        with open(DB_ARQUEO_META, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _nombres_arqueo_desde_meta(meta):
    """Extrae listas de nombres (mayúsculas) ordenados por frecuencia de uso (más usados primero)."""
    from collections import Counter
    contado_list = []
    verificado_list = []
    for rid, datos in (meta or {}).items():
        if rid.startswith("_"):
            continue
        if isinstance(datos, dict):
            c = (datos.get("contado_por") or "").strip().upper()
            v = (datos.get("verificado_por") or "").strip().upper()
            if c:
                contado_list.append(c)
            if v:
                verificado_list.append(v)
    cnt_c = Counter(contado_list)
    cnt_v = Counter(verificado_list)
    return [n for n, _ in cnt_c.most_common()], [n for n, _ in cnt_v.most_common()]


def _validar_nombre_arqueo(texto):
    """Valida que el nombre no sea solo números, tenga al menos 2 caracteres, ni caracteres inválidos. Devuelve (True, None) o (False, msg_key)."""
    t = (texto or "").strip()
    if not t:
        return True, None
    if len(t) < 2:
        return False, "arqueo_nombre_muy_corto"
    digitos = sum(1 for c in t if c.isdigit())
    if digitos >= len(t) * 0.5:
        return False, "arqueo_nombre_invalido"
    if any(c in t for c in "@#$%^&*()+={}[]|\\<>?"):
        return False, "arqueo_nombre_simbolos"
    return True, None


def _normalizar_y_coincidir(texto, lista_existentes):
    """Convierte a mayúsculas y si es similar a uno existente, devuelve el canonical. Reconoce inicial única si hay un solo match."""
    t = (texto or "").strip().upper()
    if not t:
        return ""
    if len(t) == 1 and lista_existentes:
        matches = [n for n in lista_existentes if n and len(n) > 0 and n[0].upper() == t]
        if len(matches) == 1:
            return matches[0]
    for existente in lista_existentes:
        if existente and (t == existente or (len(t) > 2 and t in existente) or (len(existente) > 2 and existente in t)):
            return existente
    try:
        import difflib
        for existente in lista_existentes:
            if existente and difflib.SequenceMatcher(None, t.lower(), existente.lower()).ratio() >= 0.85:
                return existente
    except Exception:
        pass
    return t


def guardar_arqueo_meta(meta):
    """Guarda metadatos de arqueos."""
    try:
        with open(DB_ARQUEO_META, "w", encoding="utf-8") as f:
            json.dump(meta, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False


def cargar_suministros():
    """Carga lista de suministros por pagar (luz, agua, etc.) desde JSON."""
    if not os.path.exists(DB_SUMINISTROS):
        return ["Luz", "Agua", "Gas", "Internet", "Teléfono", "Seguro", "Calefacción", "Mantenimiento edificio", "Limpieza", "Otros suministros"]
    try:
        with open(DB_SUMINISTROS, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("suministros", []) or ["Luz", "Agua", "Gas", "Internet", "Teléfono", "Seguro", "Otros suministros"]
    except Exception:
        return ["Luz", "Agua", "Gas", "Internet", "Teléfono", "Seguro", "Otros suministros"]


def _gastos_frecuentes_desde_df(df, top_n=20):
    """Extrae los gastos más frecuentes del historial (por detalle normalizado)."""
    if df is None or df.empty or "gastos" not in df.columns or "detalle" not in df.columns:
        return []
    from collections import Counter
    gastos_df = df[df["gastos"] > 0].copy()
    if gastos_df.empty:
        return []
    detalles = gastos_df["detalle"].fillna("").astype(str).str.strip()
    detalles = detalles[detalles.str.len() > 2]
    detalles_limpios = []
    for d in detalles:
        if "(Aprobado por:" in d:
            d = d.split("(Aprobado por:")[0].strip().rstrip(")")
        if d and len(d) > 2:
            detalles_limpios.append(d[:80])
    cnt = Counter(detalles_limpios)
    return [nom for nom, _ in cnt.most_common(top_n)]


def _nombres_aprobado_desde_df(df, top_n=10):
    """Extrae nombres de 'Aprobado por: X' del historial de gastos, ordenados por frecuencia."""
    if df is None or df.empty or "detalle" not in df.columns:
        return []
    from collections import Counter
    import re
    nombres = []
    for det in df["detalle"].fillna("").astype(str):
        m = re.search(r"\(Aprobado por:\s*([^)]+)\)", det, re.IGNORECASE)
        if m:
            nom = m.group(1).strip().upper()
            if nom and len(nom) > 1:
                nombres.append(nom)
    cnt = Counter(nombres)
    return [n for n, _ in cnt.most_common(top_n)]


# ============== OCR Y FACTURAS (DETECCIÓN, DUPLICADOS, BÚSQUEDA) ==============
def _hash_imagen(bytes_imagen):
    """Hash SHA256 del contenido de la imagen para detectar duplicados."""
    return hashlib.sha256(bytes_imagen).hexdigest()


def _ocr_imagen(bytes_imagen):
    """Extrae texto de la imagen con OCR (Tesseract). Si no está instalado, devuelve ''.
    En Windows: instalar Tesseract-OCR y añadirlo al PATH, o definir pytesseract.pytesseract.tesseract_cmd."""
    try:
        import pytesseract
        from PIL import Image
        img = Image.open(BytesIO(bytes_imagen))
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        return pytesseract.image_to_string(img, lang="spa+eng").strip()
    except Exception:
        return ""


def _extraer_datos_factura(ocr_text):
    """A partir del texto OCR, extrae total, impuesto, subtotal, comercio, ítems y fecha.
    Usa expresiones mejoradas para TOTAL (gran total, total a pagar), IVA/TAX, subtotal y líneas con montos."""
    out = {"total": None, "impuesto": None, "subtotal": None, "comercio": "", "items": [], "fecha_texto": ""}
    if not (ocr_text and isinstance(ocr_text, str)):
        return out
    # Normalizar: coma -> punto; varias variantes de espacio
    text = ocr_text.replace(",", ".").replace("\r", "\n")
    text_upper = text.upper()
    lineas = [ln.strip() for ln in text.splitlines() if ln.strip()]

    # Montos: $1234.56, 1,234.56, 1234.56 (2 decimales), 1234 (entero)
    patron_monto = re.compile(r"(?:\$?\s*)(\d{1,3}(?:[.,]\d{3})*(?:\.\d{2})?|\d{1,6}(?:\.\d{2})?)")
    def _parse_monto(s):
        try:
            n = float(s.replace(",", "").replace(" ", ""))
            return n if n < 1e7 else None
        except (ValueError, TypeError):
            return None

    # --- TOTAL: varias formas (gran total, total a pagar, total general, TOTAL)
    total_candidates = []
    for ln in lineas:
        ln_upper = ln.upper().replace(",", ".")
        if re.search(r"\b(GRAN\s*)?TOTAL\b", ln_upper) or re.search(r"TOTAL\s*(A\s*PAGAR|GENERAL|FINAL)?\s*:?\s*$", ln_upper) or re.search(r"TOTAL\s*\$?", ln_upper):
            nums = [m for m in patron_monto.findall(ln) if _parse_monto(m) is not None]
            if nums:
                v = _parse_monto(nums[-1])
                if v is not None:
                    total_candidates.append(round(v, 2))
    if total_candidates:
        out["total"] = max(total_candidates)  # suele ser el mayor si hay varios

    # --- IVA / TAX / IMPUESTO
    for ln in lineas:
        ln_upper = ln.upper().replace(",", ".")
        if re.search(r"\b(IVA|TAX|IMPUESTO|HST|GST|VAT)\s*:?\s*", ln_upper):
            nums = [m for m in patron_monto.findall(ln) if _parse_monto(m) is not None]
            if nums and out["impuesto"] is None:
                out["impuesto"] = round(_parse_monto(nums[-1]), 2)
                break

    # --- SUBTOTAL
    for ln in lineas:
        ln_upper = ln.upper().replace(",", ".")
        if re.search(r"\b(SUBTOTAL|SUB\-TOTAL|SUB\s*TOTAL)\s*:?\s*", ln_upper):
            nums = [m for m in patron_monto.findall(ln) if _parse_monto(m) is not None]
            if nums:
                out["subtotal"] = round(_parse_monto(nums[-1]), 2)
                break
    if out["subtotal"] is None and out["total"] is not None and out["impuesto"] is not None:
        out["subtotal"] = round(out["total"] - out["impuesto"], 2)

    # --- Si aún no hay total, usar el monto más alto que parezca un total (p. ej. último monto grande)
    if out["total"] is None:
        todos_montos = []
        for ln in text_upper.splitlines():
            for m in patron_monto.findall(ln):
                v = _parse_monto(m)
                if v is not None and 0.01 <= v < 1e6:
                    todos_montos.append(round(v, 2))
        if todos_montos:
            out["total"] = max(todos_montos)

    # --- Comercio: primeras líneas que no son solo números ni "factura"/"receipt"
    for ln in lineas[:5]:
        ln_clean = ln.strip()[:200]
        if not re.match(r"^[\d\s\.\,\$\-\/]+$", ln_clean) and not re.search(r"^(FACTURA|RECEIPT|INVOICE|DATE|FECHA)\s*$", ln_clean, re.I):
            out["comercio"] = ln_clean
            break

    # --- Ítems: líneas que contienen descripción + monto al final (ej. "Cafe con leche    2.50")
    for ln in lineas:
        ln_norm = ln.replace(",", ".")
        # Buscar monto al final de la línea ($X.XX o X.XX)
        match_end = re.search(r"[\$]?\s*(\d{1,6}(?:\.\d{2})?)\s*$", ln_norm)
        if match_end:
            desc = ln_norm[:match_end.start()].strip()
            try:
                precio = round(float(match_end.group(1)), 2)
            except (ValueError, TypeError):
                continue
            if desc and len(desc) > 2 and 0.01 <= precio < 1e5 and (not out["items"] or len(out["items"]) < 30):
                out["items"].append({"desc": desc[:150], "precio": precio})

    # --- Fecha: 2024-01-15, 15/01/2024, 15-01-2024, Jan 15 2024
    fecha_match = re.search(r"(\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}|\d{1,2}\s+[A-Za-z]+\s+\d{4})", text)
    if fecha_match:
        out["fecha_texto"] = fecha_match.group(1).strip()
    return out


def cargar_facturas():
    """Carga DB_FACTURAS.json. Estructura: { hashes: { hash: id_registro }, facturas: [ {...} ] }."""
    if not os.path.exists(DB_FACTURAS):
        return {"hashes": {}, "facturas": []}
    try:
        with open(DB_FACTURAS, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {"hashes": {}, "facturas": []}


def guardar_factura(registro):
    """Añade un registro de factura y actualiza el archivo. registro: id_registro, hash_imagen, ocr_text, total, impuesto, subtotal, comercio, tipo_gasto, items, fecha_factura, fecha_registro."""
    data = cargar_facturas()
    data.setdefault("hashes", {})
    data.setdefault("facturas", [])
    data["hashes"][registro.get("hash_imagen", "")] = registro.get("id_registro", "")
    data["facturas"].append(registro)
    try:
        with open(DB_FACTURAS, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
    except Exception:
        pass


def imagen_ya_subida(hash_imagen):
    """True si ese hash ya está registrado (evitar duplicados)."""
    data = cargar_facturas()
    return hash_imagen in data.get("hashes", {})


def _comprimir_imagen(bytes_imagen, max_ancho=None, calidad_jpeg=None):
    """Comprime la imagen para móvil: redimensiona si es muy grande y guarda como JPEG.
    Usa IMAGEN_COMPRIMIR_MAX_ANCHO y IMAGEN_COMPRIMIR_CALIDAD si no se pasan. Devuelve bytes del JPEG o los originales si falla."""
    if max_ancho is None:
        max_ancho = IMAGEN_COMPRIMIR_MAX_ANCHO
    if calidad_jpeg is None:
        calidad_jpeg = IMAGEN_COMPRIMIR_CALIDAD
    try:
        from PIL import Image
        img = Image.open(BytesIO(bytes_imagen))
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        elif img.mode != "RGB":
            img = img.convert("RGB")
        w, h = img.size
        if w > max_ancho or h > max_ancho:
            ratio = min(max_ancho / w, max_ancho / h)
            new_size = (int(w * ratio), int(h * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)
        buf = BytesIO()
        img.save(buf, format="JPEG", quality=calidad_jpeg, optimize=True)
        buf.seek(0)
        return buf.read()
    except Exception:
        return bytes_imagen


def _crear_zip_fotos_gastos():
    """Crea un ZIP con todas las fotos de la carpeta fotos_gastos. Devuelve (bytes_zip, nombre_archivo) o (None, None)."""
    import zipfile
    carpeta = "fotos_gastos"
    if not os.path.isdir(carpeta):
        return None, None
    nombres = [f for f in os.listdir(carpeta) if os.path.isfile(os.path.join(carpeta, f))]
    if not nombres:
        return None, None
    buf = BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in nombres:
            ruta = os.path.join(carpeta, n)
            zf.write(ruta, n)
    buf.seek(0)
    nombre_zip = f"fotos_gastos_{datetime.now().strftime('%Y%m%d_%H%M')}.zip"
    return buf.getvalue(), nombre_zip


# ============== USUARIOS Y PERMISOS (ADMIN) ==============
def _permisos_default():
    return {
        "usuarios": {
            "admin": {"nombre": "Administrador", "permisos": ["*"]},
            "asistente": {"nombre": "Asistente", "permisos": ["ver_inicio", "ver_arqueo_caja"]},
            "tesorero": {"nombre": "Tesorero", "permisos": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"]},
            "pastor": {"nombre": "Pastor", "permisos": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_presupuesto_metas", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"]},
            "ministerio_musica": {"nombre": "Ministerio de Música", "permisos": ["ver_inicio", "ver_arqueo_caja", "ver_hoja_contable", "ver_informe_pdf"]},
        }
    }

@st.cache_data(ttl=30)
def cargar_permisos():
    if not os.path.exists(DB_PERMISOS):
        data = _permisos_default()
        try:
            with open(DB_PERMISOS, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception:
            pass
        return data
    try:
        with open(DB_PERMISOS, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return _permisos_default()
    # Validar estructura: debe tener "usuarios" como dict; cada usuario con "nombre" y "permisos"
    if not isinstance(data.get("usuarios"), dict):
        return _permisos_default()
    need_save = False
    for uid, info in list(data["usuarios"].items()):
        if not isinstance(info, dict):
            data["usuarios"][uid] = {"nombre": str(uid), "permisos": []}
            need_save = True
        else:
            if "nombre" not in info:
                data["usuarios"][uid]["nombre"] = str(uid)
                need_save = True
            if not isinstance(info.get("permisos"), list):
                data["usuarios"][uid]["permisos"] = []
                need_save = True
    # Migración: ver_ministerio_finanzas -> ver_arqueo_caja, ver_tesoreria, ver_contabilidad
    for uid, info in list(data.get("usuarios", {}).items()):
        perms = info.get("permisos", [])
        if "ver_ministerio_finanzas" in perms and "*" not in perms:
            perms.remove("ver_ministerio_finanzas")
            perms.extend(["ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad"])
            data["usuarios"][uid]["permisos"] = list(set(perms))
            need_save = True
    # Añadir ministerio_musica si no existe (migración)
    default_musica = {"nombre": "Ministerio de Música", "permisos": ["ver_inicio", "ver_arqueo_caja", "ver_hoja_contable", "ver_informe_pdf"]}
    if "ministerio_musica" not in data.get("usuarios", {}):
        data["usuarios"]["ministerio_musica"] = default_musica
        need_save = True
    # Añadir asistente y pastor si no existen
    if "asistente" not in data.get("usuarios", {}):
        data["usuarios"]["asistente"] = {"nombre": "Asistente", "permisos": ["ver_inicio", "ver_arqueo_caja"]}
        need_save = True
    if "pastor" not in data.get("usuarios", {}):
        data["usuarios"]["pastor"] = {"nombre": "Pastor", "permisos": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_presupuesto_metas", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"]}
        need_save = True
    if "orden_usuarios" not in data or not isinstance(data["orden_usuarios"], list):
        data["orden_usuarios"] = list(data.get("usuarios", {}).keys())
        need_save = True
    if need_save:
        try:
            with open(DB_PERMISOS, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)
        except Exception:
            pass
    return data

def guardar_permisos(data, t=None):
    """Guarda permisos en JSON. Devuelve True si se guardó bien, False si hubo error. t: textos para i18n."""
    try:
        with open(DB_PERMISOS, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        cargar_permisos.clear()
        return True
    except Exception as e:
        st.error(t["error_no_se_pudo_guardar_permisos"].format(e=str(e)) if t else f"NO SE PUDO GUARDAR PERMISOS: {e}")
        return False

def tiene_permiso(usuario_id, permiso_clave):
    """True si el usuario tiene ese permiso (admin tiene todos)."""
    data = cargar_permisos()
    usuarios = data.get("usuarios", {})
    u = usuarios.get(usuario_id, {})
    permisos = u.get("permisos", [])
    if "*" in permisos or permiso_clave in permisos:
        return True
    return False

def toggle_permiso(usuario_id, permiso_clave, t=None):
    """Con un clic: si tiene permiso lo quita; si no lo tiene, lo da. t: textos para i18n."""
    if usuario_id == "admin":
        return  # admin no se modifica
    data = cargar_permisos()
    usuarios = data.get("usuarios", {})
    if usuario_id not in usuarios:
        return
    permisos = list(usuarios[usuario_id].get("permisos", []))
    if "*" in permisos:
        return  # no tocar admin
    if permiso_clave in permisos:
        permisos.remove(permiso_clave)
    else:
        permisos.append(permiso_clave)
    usuarios[usuario_id]["permisos"] = permisos
    data["usuarios"] = usuarios
    guardar_permisos(data, t)
    try:
        audit_log(st.session_state.get("usuario_actual", "?"), "permiso_cambiado", f"{usuario_id} {permiso_clave}")
    except Exception:
        pass

def puede_borrar(fecha_str):
    try:
        s = str(fecha_str).strip()
        if len(s) <= 10:
            dt = datetime.strptime(s[:10], "%Y-%m-%d")
        else:
            dt = datetime.strptime(s[:19], "%Y-%m-%d %H:%M:%S")
        return (datetime.now() - dt).total_seconds() <= MINUTOS_BORRADO * 60
    except Exception:
        return False

def _calcular_alertas(total_ingresos, total_gastos, saldo, t):
    """Devuelve lista de alertas de crisis y nivel de riesgo (bajo/medio/alto)."""
    alertas = []
    if total_ingresos == 0 and total_gastos > 0:
        alertas.append(t["alerta_sin_ingresos"])
    if saldo < 0:
        alertas.append(t["alerta_saldo_negativo"])
    elif total_ingresos > 0 and total_gastos / total_ingresos >= 0.90:
        alertas.append(t["alerta_gastos_altos"])
    if saldo >= 0 and total_gastos > 0 and total_ingresos > 0:
        meses_reserva = saldo / total_gastos if total_gastos else 0
        if 0 < meses_reserva < 0.5:
            alertas.append(t["alerta_saldo_bajo"])
    # Riesgo
    if saldo < 0:
        nivel_riesgo = "alto"
    elif total_ingresos > 0 and total_gastos / total_ingresos >= 0.85:
        nivel_riesgo = "medio"
    elif total_ingresos == 0 and total_gastos > 0:
        nivel_riesgo = "alto"
    elif total_ingresos > 0 and saldo / total_ingresos < 0.1:
        nivel_riesgo = "medio"
    else:
        nivel_riesgo = "bajo"
    return alertas, nivel_riesgo

def _safe_float_pdf(x, default=0.0):
    """Convierte a float de forma segura para PDF (evita fallo con cadenas vacías o NaN)."""
    try:
        if x is None or (isinstance(x, float) and pd.isna(x)):
            return default
        return float(x)
    except (TypeError, ValueError):
        return default

def generar_pdf(df, lang, saldo, texto_arqueo):
    """Genera informe de auditoría contable: situación, alertas, presupuesto real, análisis, riesgo y recomendaciones."""
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    t = TEXTOS.get(lang, TEXTOS["ES"])
    total_ingresos = float(df["ingreso"].sum(skipna=True) or 0) if not df.empty else 0.0
    total_gastos = float(df["gastos"].sum(skipna=True) or 0) if not df.empty else 0.0
    saldo_val = saldo if pd.notna(saldo) else (total_ingresos - total_gastos)
    alertas, nivel_riesgo = _calcular_alertas(total_ingresos, total_gastos, saldo_val, t)
    # Presupuesto por tipo de gasto
    if not df.empty and "tipo_gasto" in df.columns:
        gastos_por_tipo = df[df["gastos"] > 0].groupby("tipo_gasto", dropna=False)["gastos"].sum()
        gastos_por_tipo = gastos_por_tipo[gastos_por_tipo.index.astype(str).str.strip() != ""]
        if gastos_por_tipo.empty:
            gastos_por_tipo = pd.Series({"Otros": total_gastos}) if total_gastos else pd.Series(dtype=float)
    else:
        gastos_por_tipo = pd.Series({"Otros": total_gastos}) if total_gastos else pd.Series(dtype=float)
    # --- Encabezado ---
    y = h - 1.5*cm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, y, t["informe_auditoria"])
    y -= 0.6*cm
    c.setFont("Helvetica", 9)
    c.drawString(2*cm, y, f"{t['fecha_informe']}: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    y -= 0.4*cm
    c.drawString(2*cm, y, t.get("direccion", DIRECCION_IGLESIA))
    y -= 0.5*cm
    c.drawString(2*cm, y, f"{t['arqueo']} (último): {texto_arqueo}" if isinstance(texto_arqueo, str) else f"{t['arqueo']}: {texto_arqueo}")
    y -= 0.8*cm
    # --- 1. SITUACIÓN ACTUAL ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["situacion_actual"])
    y -= 0.5*cm
    c.setFont("Helvetica", 9)
    c.drawString(2*cm, y, f"{t['ingresos']}: ${total_ingresos:,.2f}")
    y -= 0.4*cm
    c.drawString(2*cm, y, f"{t['gastos']}: ${total_gastos:,.2f}")
    y -= 0.4*cm
    c.drawString(2*cm, y, f"{t['saldo']}: ${saldo_val:,.2f}")
    y -= 0.4*cm
    if saldo_val >= 0:
        c.drawString(2*cm, y, f"  → {t['superavit']}: ${saldo_val:,.2f}")
    else:
        c.drawString(2*cm, y, f"  → {t['deficit']}: ${abs(saldo_val):,.2f}")
    y -= 0.8*cm
    # --- 2. ALERTAS DE CRISIS ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["alertas_crisis"])
    y -= 0.45*cm
    c.setFont("Helvetica", 8)
    if not alertas:
        c.drawString(2*cm, y, t["sin_alertas"])
    else:
        for a in alertas:
            c.drawString(2.5*cm, y, f"• {a[:75]}")
            y -= 0.38*cm
    y -= 0.5*cm
    # --- 3. PRESUPUESTO REAL ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["presupuesto_real"])
    y -= 0.45*cm
    c.setFont("Helvetica", 8)
    if total_gastos > 0 and not gastos_por_tipo.empty:
        for tipo, monto in gastos_por_tipo.items():
            pct = (float(monto) / total_gastos * 100) if total_gastos else 0
            if tipo is None or (isinstance(tipo, float) and pd.isna(tipo)) or str(tipo).strip() in ("", "nan"):
                tipo_str = "Otros"
            else:
                tipo_str = str(tipo)[:22]
            c.drawString(2*cm, y, f"  {tipo_str}: ${float(monto):,.2f} ({pct:.0f}%)")
            y -= 0.38*cm
        c.drawString(2*cm, y, f"  {t['total']}: ${total_gastos:,.2f}")
    else:
        c.drawString(2*cm, y, f"  {t['total']} {t['gastos']}: $0.00")
    y -= 0.7*cm
    # --- 4. ANÁLISIS ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["analisis_clinico"])
    y -= 0.45*cm
    c.setFont("Helvetica", 8)
    if lang == "ES":
        if saldo_val < 0:
            parrafo = "Las finanzas están en déficit. Los gastos superan a los ingresos. Se recomienda reducir gastos o aumentar ingresos de forma urgente para evitar mayor endeudamiento."
        elif total_ingresos > 0 and total_gastos / total_ingresos >= 0.85:
            parrafo = "El margen entre ingresos y gastos es estrecho. Cualquier imprevisto puede generar déficit. Conviene aumentar el ahorro y revisar gastos por categoría."
        elif saldo_val >= 0 and total_ingresos > 0:
            parrafo = "Situación estable: hay superávit. Se recomienda mantener control en gastos y destinar parte del excedente a reserva."
        else:
            parrafo = "Sin movimientos suficientes para análisis. Registre ingresos y gastos para obtener un diagnóstico claro."
    else:
        if saldo_val < 0:
            parrafo = "Finances are in deficit. Expenses exceed income. Reduce spending or increase income urgently."
        elif total_ingresos > 0 and total_gastos / total_ingresos >= 0.85:
            parrafo = "Narrow margin between income and expenses. Build savings and review spending by category."
        elif saldo_val >= 0 and total_ingresos > 0:
            parrafo = "Stable situation: surplus. Maintain spending control and allocate part of the surplus to reserves."
        else:
            parrafo = "Insufficient data. Record income and expenses for a clear diagnosis."
    for i in range(0, len(parrafo), 72):
        c.drawString(2*cm, y, parrafo[i:i+72])
        y -= 0.38*cm
    y -= 0.5*cm
    # --- 5. EVALUACIÓN DE RIESGO ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["riesgo"])
    y -= 0.45*cm
    c.setFont("Helvetica", 9)
    riesgo_label = t["riesgo_alto"] if nivel_riesgo == "alto" else (t["riesgo_medio"] if nivel_riesgo == "medio" else t["riesgo_bajo"])
    c.drawString(2*cm, y, f"  → {riesgo_label}")
    y -= 0.7*cm
    # --- 6. RECOMENDACIONES ---
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, t["recomendaciones"])
    y -= 0.45*cm
    c.setFont("Helvetica", 8)
    for key in ["recom_control_gastos", "recom_reserva", "recom_revisar", "recom_evitar_deficit"]:
        c.drawString(2*cm, y, t[key][:78])
        y -= 0.38*cm
    y -= 0.5*cm
    # --- 7. ÚLTIMOS MOVIMIENTOS ---
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y, t["ultimos_movimientos"])
    y -= 0.45*cm
    c.setFont("Helvetica", 7)
    for _, row in df.tail(18).iterrows():
        if y < 2*cm:
            c.showPage()
            y = h - 1.5*cm
        ing = _safe_float_pdf(row.get("ingreso"))
        gas = _safe_float_pdf(row.get("gastos"))
        det = str(row.get("detalle", "") or "")[:20]
        tipo = str(row.get("tipo_gasto", "") or "")[:8]
        linea = f"{str(row.get('fecha',''))[:10]} | {tipo} | {det} | ${ing:.2f} | ${gas:.2f}"
        c.drawString(2*cm, y, linea[:85])
        y -= 0.36*cm
    c.save()
    buf.seek(0)
    return buf

def generar_pdf_hoja_contable(df, lang):
    """Genera PDF de la hoja contable con columnas: ID, Fecha, Detalle, Tipo, Ingreso, Gastos, Totales, Saldo."""
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    t = TEXTOS.get(lang, TEXTOS["ES"])
    c.setFont("Helvetica-Bold", 14)
    c.drawString(2*cm, h - 1.5*cm, t["informe_pdf"] + " - HOJA CONTABLE")
    c.setFont("Helvetica", 9)
    c.drawString(2*cm, h - 2*cm, f"{t['fecha_informe']}: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.drawString(2*cm, h - 2.4*cm, t.get("direccion", DIRECCION_IGLESIA))
    headers = [t["col_id_registro"], t["col_fecha"], t["col_detalle"], t["col_tipo_gasto"], t["col_ingreso"], t["col_gastos"], t["col_total_ingresos"], t["col_total_gastos"], t["col_saldo_actual"]]
    col_widths = [2*cm, 2.2*cm, 3.2*cm, 2.2*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.6*cm]
    x = 1*cm
    c.setFont("Helvetica-Bold", 8)
    for i, (header, cw) in enumerate(zip(headers, col_widths)):
        c.drawString(x, h - 2.9*cm, header[:14])
        x += cw
    y = h - 3.2*cm
    c.setFont("Helvetica", 7)
    for _, row in df.iterrows():
        if y < 1.5*cm:
            c.showPage()
            y = h - 2*cm
            x = 1*cm
            for header, cw in zip(headers, col_widths):
                c.drawString(x, y, header[:14])
                x += cw
            y -= 0.4*cm
        x = 1*cm
        ing = _safe_float_pdf(row.get("ingreso"))
        gas = _safe_float_pdf(row.get("gastos"))
        tot_ing = _safe_float_pdf(row.get("total_ingresos"))
        tot_gas = _safe_float_pdf(row.get("total_gastos"))
        sal = _safe_float_pdf(row.get("saldo_actual"))
        det = str(row.get("detalle", "") or "")[:14]
        tipo = str(row.get("tipo_gasto", "") or "")[:10]
        c.drawString(x, y, str(row.get("id_registro", ""))[:10])
        x += col_widths[0]
        c.drawString(x, y, str(row.get("fecha", ""))[:10])
        x += col_widths[1]
        c.drawString(x, y, det)
        x += col_widths[2]
        c.drawString(x, y, tipo)
        x += col_widths[3]
        c.drawString(x, y, f"${ing:.2f}")
        x += col_widths[4]
        c.drawString(x, y, f"${gas:.2f}")
        x += col_widths[5]
        c.drawString(x, y, f"${tot_ing:.2f}")
        x += col_widths[6]
        c.drawString(x, y, f"${tot_gas:.2f}")
        x += col_widths[7]
        c.drawString(x, y, f"${sal:.2f}")
        y -= 0.4*cm
    c.save()
    buf.seek(0)
    return buf

def _formatear_excel_contador(buf):
    """Ajusta anchos de columna y formato de moneda en el Excel exportado para el contador."""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        buf.seek(0)
        wb = load_workbook(buf)
        ws = wb.active
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            width = min(max(max_len * 1.15, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=5, max_col=9):
            for cell in row:
                if cell.value is not None and isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        buf_out = BytesIO()
        wb.save(buf_out)
        buf_out.seek(0)
        return buf_out.getvalue()
    except Exception:
        buf.seek(0)
        return buf.getvalue()


def generar_pdf_hoja_arqueo(datos_arqueo, lang):
    """Genera PDF compacto de hoja de arqueo para WhatsApp (comprimido, abre en cualquier dispositivo)."""
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4, pageCompression=1)
    w, h = A4
    t = TEXTOS.get(lang, TEXTOS["ES"])
    c.setFont("Helvetica-Bold", 12)
    c.drawString(2*cm, h - 1.2*cm, "HOJA DE ARQUEO - " + (datos_arqueo.get("fecha", "")[:16] or datetime.now().strftime("%Y-%m-%d %H:%M")))
    c.setFont("Helvetica", 8)
    c.drawString(2*cm, h - 1.6*cm, t.get("direccion", DIRECCION_IGLESIA)[:60])
    y = h - 2.2*cm
    desglose = datos_arqueo.get("desglose", {})
    if desglose:
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2*cm, y, "Billetes:")
        y -= 0.4*cm
        c.setFont("Helvetica", 8)
        for denom, key, val in [("$100", "b100", 100), ("$50", "b50", 50), ("$20", "b20", 20), ("$10", "b10", 10), ("$5", "b5", 5)]:
            cant = desglose.get(key, 0) or 0
            if cant:
                c.drawString(2*cm, y, f"  {denom}: {cant} = ${cant * val:.2f}")
                y -= 0.35*cm
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2*cm, y, "Monedas:")
        y -= 0.4*cm
        c.setFont("Helvetica", 8)
        for denom, key, val in [("$2", "m2", 2), ("$1", "m1", 1), ("$0.25", "m025", 0.25), ("$0.10", "m010", 0.10), ("$0.05", "m005", 0.05)]:
            cant = desglose.get(key, 0) or 0
            if cant:
                c.drawString(2*cm, y, f"  {denom}: {cant} = ${cant * val:.2f}")
                y -= 0.35*cm
    total_efectivo = datos_arqueo.get("total_efectivo", 0) or 0
    total_cheques = datos_arqueo.get("total_cheques", 0) or 0
    total_pos = datos_arqueo.get("total_pos", 0) or 0
    sobres_tot = datos_arqueo.get("sobres_tot", 0) or 0
    total = datos_arqueo.get("total", 0) or total_efectivo + total_cheques + total_pos + sobres_tot
    if sobres_tot:
        y -= 0.3*cm
        c.setFont("Helvetica", 8)
        c.drawString(2*cm, y, f"Sobres: {datos_arqueo.get('sobres_cant', 0)} = ${float(sobres_tot):.2f}")
    if total_cheques:
        y -= 0.35*cm
        c.setFont("Helvetica", 8)
        c.drawString(2*cm, y, f"Cheques: {datos_arqueo.get('cheques_cant', 0)} = ${float(total_cheques):.2f}")
    y -= 0.3*cm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(2*cm, y, f"TOTAL: ${float(total):.2f}")
    y -= 0.5*cm
    c.setFont("Helvetica", 8)
    contado = datos_arqueo.get("contado_por", "") or "-"
    verif = datos_arqueo.get("verificado_por", "") or "-"
    c.drawString(2*cm, y, f"Contado por: {contado[:40]}")
    y -= 0.35*cm
    c.drawString(2*cm, y, f"Verificado por: {verif[:40]}")
    c.save()
    buf.seek(0)
    return buf.getvalue()


def generar_excel_hoja_arqueo(datos_arqueo, lang):
    """Genera Excel compacto de hoja de arqueo para WhatsApp (comprimido, abre en cualquier dispositivo)."""
    try:
        t = TEXTOS.get(lang, TEXTOS["ES"])
        filas = [["HOJA DE ARQUEO", ""], ["Fecha", datos_arqueo.get("fecha", datetime.now().strftime("%Y-%m-%d %H:%M"))]]
        desglose = datos_arqueo.get("desglose", {})
        if desglose:
            filas.append(["Billetes", ""])
            for denom, key in [("$100", "b100"), ("$50", "b50"), ("$20", "b20"), ("$10", "b10"), ("$5", "b5")]:
                filas.append([denom, desglose.get(key, 0)])
            filas.append(["Monedas", ""])
            for denom, key in [("$2", "m2"), ("$1", "m1"), ("$0.25", "m025"), ("$0.10", "m010"), ("$0.05", "m005")]:
                filas.append([denom, desglose.get(key, 0)])
        if datos_arqueo.get("sobres_cant") or datos_arqueo.get("sobres_tot"):
            filas.append(["Sobres (cant)", datos_arqueo.get("sobres_cant", 0)])
            filas.append(["Sobres (total $)", datos_arqueo.get("sobres_tot", 0)])
        if datos_arqueo.get("cheques_cant") or datos_arqueo.get("total_cheques"):
            filas.append(["Cheques (cant)", datos_arqueo.get("cheques_cant", 0)])
            filas.append(["Cheques (total $)", datos_arqueo.get("total_cheques", 0)])
        total = datos_arqueo.get("total", 0) or 0
        filas.append(["TOTAL", total])
        filas.append(["Contado por", datos_arqueo.get("contado_por", "")])
        filas.append(["Verificado por", datos_arqueo.get("verificado_por", "")])
        df = pd.DataFrame(filas)
        buf = BytesIO()
        df.to_excel(buf, index=False, header=False, engine="openpyxl")
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return b""


# ============== PÁGINA PRINCIPAL ==============
def main():
    if "idioma" not in st.session_state:
        st.session_state["idioma"] = "ES"
    if "pagina" not in st.session_state:
        st.session_state["pagina"] = "inicio"
    if "usuario_actual" not in st.session_state:
        st.session_state["usuario_actual"] = "admin"
    if "logueado" not in st.session_state:
        st.session_state["logueado"] = False
    if "tema_app" not in st.session_state:
        st.session_state["tema_app"] = "oscuro"

    lang = st.session_state.get("idioma", "ES")
    t = TEXTOS.get(lang, TEXTOS["ES"])

    # ----- PANTALLA DE LOGIN (puerta de acceso): si no está logueado, mostrar solo esto -----
    if not st.session_state.get("logueado"):
        _render_pantalla_login()
        return

    # ----- PRIMERA VEZ admin/admin: obligatorio cambiar contraseña antes de continuar -----
    if st.session_state.get("debe_cambiar_credenciales") and st.session_state.get("usuario_actual") == "admin":
        _render_pantalla_cambiar_credenciales()
        return

    # Mensaje una vez: para que "recordar" funcione en el celular, enlace con token en la URL
    if st.session_state.pop("mostrar_info_recordar_url", False):
        t_ = TEXTOS.get(lang, TEXTOS["ES"])
        token_link = st.session_state.pop("recordar_token_link", None)
        if token_link:
            msg = t_.get("recordar_guardar_pagina_celular", "En el celular: 1) Toca el enlace de abajo. 2) Cuando se abra la app, en el menú del navegador elige «Añadir a pantalla de inicio» o «Guardar en Favoritos». La próxima vez que abras la app desde ese acceso, la contraseña se recordará.")
            st.info(msg)
            # Enlace con URL completa vía JavaScript para que en móvil se guarde bien al añadir a inicio
            token_esc = token_link.replace("\\", "\\\\").replace("'", "\\'")
            st.markdown(f"""
            <p style="margin-top:0.5rem;"><a id="recordar-enlace-celular" href="?r={token_link}" style="font-weight:bold;font-size:1.05rem;">{t_.get('recordar_abrir_enlace', 'Abrir con sesión guardada (toca aquí)')}</a></p>
            <script>
            (function() {{
                var a = document.getElementById('recordar-enlace-celular');
                if (a && window.location) {{
                    var base = window.location.origin + window.location.pathname;
                    a.href = base + (base.indexOf('?') >= 0 ? '&' : '?') + 'r={token_esc}';
                }}
            }})();
            </script>
            """, unsafe_allow_html=True)
        else:
            msg = t_.get("recordar_guardar_pagina", "Para que la sesión se recuerde en este dispositivo (también en el celular), guarda esta página en favoritos o en la pantalla de inicio.")
            st.info(msg)

    # Timeout por inactividad: si pasaron más de MINUTOS_INACTIVIDAD, quitar autorización (salvo si "Recordar sesión")
    now = time.time()
    last = st.session_state.get("last_activity", now)
    if not st.session_state.get("recordar_sesion", False):
        if last and (now - last) > MINUTOS_INACTIVIDAD * 60:
            if st.session_state.get("admin_autorizado"):
                audit_log(st.session_state.get("usuario_actual", "?"), "sesion_expirada_inactividad", "")
            st.session_state["admin_autorizado"] = False
    st.session_state["last_activity"] = now

    # Fix menú lateral: Streamlit ignora initial_sidebar_state si el valor no cambia (issue #4483).
    # Solución AlonSam: mantener "expanded" por defecto para que al pulsar nav haya cambio a "collapsed".
    collapse_requested = st.session_state.pop("sidebar_collapse_requested", False)
    if collapse_requested:
        sidebar_state = "collapsed"
    else:
        sidebar_state = st.session_state.get("sidebar_state", "expanded")
    st.set_page_config(
        page_title=t["titulo"],
        page_icon="⛪",
        layout="wide",
        initial_sidebar_state=sidebar_state
    )

    # Solo resetear a "expanded" cuando NO acabamos de colapsar, para que el próximo clic en nav detecte el cambio.
    # Si acabamos de colapsar, mantener "collapsed" para que la flechita de Streamlit permita reabrir.
    if not collapse_requested:
        st.session_state["sidebar_state"] = "expanded"

    # No forzar width:0 con CSS: ocultaría el botón para volver a abrir el menú.
    # Streamlit colapsa nativamente y deja visible la flechita/hamburguesa para reabrir.

    # PWA: meta tags para que se pueda "instalar" como app (Netflix/Disney style) y compartir link por WhatsApp
    _theme_color = "#1a365d" if st.session_state.get("tema_app", "oscuro") == "claro" else "#0d1b2a"
    st.markdown(f"""
    <script>
    (function() {{
      var meta = document.createElement('meta');
      meta.name = 'viewport';
      meta.content = 'width=device-width, initial-scale=1, maximum-scale=1, user-scalable=no, viewport-fit=cover';
      if (document.head) document.head.appendChild(meta);
      var theme = document.createElement('meta');
      theme.name = 'theme-color';
      theme.content = '{_theme_color}';
      if (document.head) document.head.appendChild(theme);
      var apple = document.createElement('meta');
      apple.name = 'apple-mobile-web-app-capable';
      apple.content = 'yes';
      if (document.head) document.head.appendChild(apple);
      var appleStatus = document.createElement('meta');
      appleStatus.name = 'apple-mobile-web-app-status-bar-style';
      appleStatus.content = 'black-translucent';
      if (document.head) document.head.appendChild(appleStatus);
      var link = document.createElement('link');
      link.rel = 'manifest';
      link.href = '/app/static/manifest.webmanifest';
      if (document.head) document.head.appendChild(link);
    }})();
    </script>
    """, unsafe_allow_html=True)

    # Ocultar logo de Streamlit, botón Deploy y hint "Press Enter to submit form"
    st.markdown("""
    <style>
    a[href*="streamlit.io"], a[href*="streamlit.io"] img,
    [data-testid="stDeployButton"], [data-testid="stToolbar"] a, [data-testid="stToolbar"] img,
    header a[href*="streamlit"], header img[alt*="Streamlit"] { display: none !important; visibility: hidden !important; }
    [data-testid="stFormSubmitButton"] + div { display: none !important; }
    </style>
    """, unsafe_allow_html=True)
    # Ocultar iconos de desarrollador (toolbar, menú, Host, etc.) para todos excepto clave universal
    # IMPORTANTE: cuando el sidebar está colapsado, NO ocultar el botón de expandir (flechita/hamburguesa)
    es_maestro = st.session_state.get("es_acceso_maestro", False)
    sidebar_colapsado = st.session_state.get("sidebar_state") == "collapsed"
    if not es_maestro:
        if sidebar_colapsado:
            # Sidebar colapsado: mostrar solo la flechita para reabrir; ocultar el resto del header
            st.markdown("""
            <style>
            [data-testid="stToolbar"], [data-testid="stStatusWidget"], [data-testid="stDeployButton"] { display: none !important; }
            #MainMenu, footer { visibility: hidden !important; }
            div[data-testid="stToolbar"] { display: none !important; }
            /* Mantener visible el botón/flechita para expandir el sidebar (cortina) */
            header[data-testid="stHeader"] { min-height: 2.5rem !important; }
            button[kind="header"] { display: flex !important; visibility: visible !important; }
            </style>
            """, unsafe_allow_html=True)
        else:
            st.markdown("""
            <style>
            /* Ocultar barra de herramientas, menú y opciones de desarrollador para usuarios normales */
            [data-testid="stToolbar"] { display: none !important; }
            [data-testid="stStatusWidget"] { display: none !important; }
            header[data-testid="stHeader"] { display: none !important; }
            .stApp > header { display: none !important; }
            #MainMenu { visibility: hidden !important; }
            footer { visibility: hidden !important; }
            [data-testid="stDeployButton"] { display: none !important; }
            button[kind="header"] { display: none !important; }
            div[data-testid="stToolbar"] { display: none !important; }
            </style>
            """, unsafe_allow_html=True)

    # Estilo: menú última generación + tema (oscuro/claro)
    tema_app = st.session_state.get("tema_app", "oscuro")
    if tema_app == "claro":
        bg_main = "linear-gradient(180deg, #e8eef5 0%, #d1dce8 50%, #b8c9d9 100%)"
        txt_main = "#1a365d"
        btn_txt = "#1a365d"
        btn_bg = "rgba(26, 54, 93, 0.15)"
        btn_bg_grad = "linear-gradient(180deg, rgba(255,255,255,0.98) 0%, rgba(230,240,255,0.98) 50%, rgba(200,220,245,0.98) 100%)"
        btn_bg_hover = "linear-gradient(180deg, rgba(240,248,255,0.98) 0%, rgba(210,230,250,0.98) 50%, rgba(180,210,240,0.98) 100%)"
        btn_border = "1px solid rgba(26,54,93,0.4)"
        sidebar_btn_bg = "linear-gradient(180deg, rgba(255,255,255,0.95) 0%, rgba(220,235,255,0.95) 50%, rgba(190,215,245,0.95) 100%)"
        sidebar_btn_hover = "linear-gradient(180deg, rgba(240,248,255,0.98) 0%, rgba(200,225,250,0.98) 50%, rgba(170,205,240,0.98) 100%)"
        input_bg = "rgba(255,255,255,0.95)"
        input_color = "#1a365d"
        sidebar_bg = "linear-gradient(180deg, #c5d4e4 0%, #a8bdd4 50%, #8fa9c2 100%)"
        sidebar_txt = "#1a365d"
        sidebar_txt_muted = "rgba(26,54,93,0.7)"
        menu_bg = "rgba(26,54,93,0.12)"
        menu_bg_hover = "rgba(26,54,93,0.2)"
        checkbox_border = "rgba(26,54,93,0.3)"
        checkbox_unchecked_color = "#b91c1c"
        checkbox_unchecked_border = "#dc2626"
        checkbox_checked_color = "#15803d"
        checkbox_checked_border = "#22c55e"
        active_menu_bg = "linear-gradient(135deg, rgba(26,54,93,0.3) 0%, rgba(26,54,93,0.45) 100%)"
        active_menu_txt = "#1a365d"
    else:
        bg_main = f"linear-gradient(180deg, #0a1220 0%, {AZUL_OSCURO} 50%, #132238 100%)"
        txt_main = "#FFFFFF"
        btn_txt = "#FFFFFF"
        btn_bg = "rgba(44, 48, 56, 0.25)"
        btn_bg_grad = "linear-gradient(180deg, rgba(70,78,90,0.98) 0%, rgba(45,52,62,0.98) 50%, rgba(30,36,44,0.98) 100%)"
        btn_bg_hover = "linear-gradient(180deg, rgba(80,88,100,0.98) 0%, rgba(55,62,72,0.98) 50%, rgba(40,46,54,0.98) 100%)"
        btn_border = "1px solid transparent"
        sidebar_btn_bg = "linear-gradient(180deg, rgba(55,65,78,0.9) 0%, rgba(35,42,52,0.9) 50%, rgba(25,30,38,0.9) 100%)"
        sidebar_btn_hover = "linear-gradient(180deg, rgba(65,75,88,0.95) 0%, rgba(45,52,62,0.95) 50%, rgba(35,40,48,0.95) 100%)"
        input_bg = "rgba(30, 40, 55, 0.4)"
        input_color = "#FFFFFF"
        sidebar_bg = f"linear-gradient(180deg, #0a1220 0%, {AZUL_OSCURO} 40%, #0d2238 100%)"
        sidebar_txt = "#FFFFFF"
        sidebar_txt_muted = "rgba(255,255,255,0.7)"
        menu_bg = "rgba(255,255,255,0.08)"
        menu_bg_hover = "rgba(255,255,255,0.15)"
        checkbox_border = "rgba(255,255,255,0.2)"
        checkbox_unchecked_color = "#fca5a5"
        checkbox_unchecked_border = "#ef4444"
        checkbox_checked_color = "#86efac"
        checkbox_checked_border = "#22c55e"
        active_menu_bg = "linear-gradient(135deg, #132238 0%, #1a2942 100%)"
        active_menu_txt = "#FFFFFF"
    st.markdown(f"""
    <style>
    /* Reducir espacio perdido: aprovechar pantalla en móvil y todas las oficinas */
    .stApp, [data-testid="stAppViewContainer"], .main .block-container,
    div[data-testid="stAppViewContainer"] .block-container,
    section[data-testid="stSidebar"] + div .block-container,
    section.main .block-container {{
        padding-top: 0.2rem !important;
        padding-bottom: 0.5rem !important;
        padding-left: 0.75rem !important;
        padding-right: 0.75rem !important;
    }}
    div[data-testid="stVerticalBlock"] > div {{
        padding: 0.2rem 0.15rem !important;
    }}
    .main [data-testid="stVerticalBlock"] > div:first-child {{ padding-top: 0 !important; margin-top: 0 !important; }}
    @media (max-width: 768px) {{
        .stApp, [data-testid="stAppViewContainer"], .main .block-container,
        div[data-testid="stAppViewContainer"] .block-container,
        section.main .block-container {{
            padding-top: 0.05rem !important;
            padding-bottom: 0.35rem !important;
            padding-left: 0.4rem !important;
            padding-right: 0.4rem !important;
        }}
        div[data-testid="stVerticalBlock"] > div {{
            padding: 0.15rem 0.1rem !important;
        }}
        .main hr {{ margin: 0.2rem 0 !important; }}
        [data-testid="column"] {{ padding-left: 0.15rem !important; padding-right: 0.15rem !important; }}
        .main h1, .main h2, .main h3 {{ margin-top: 0.2rem !important; margin-bottom: 0.25rem !important; }}
        .main .stMarkdown {{ margin-bottom: 0.2rem !important; }}
    }}
    .stApp, [data-testid="stAppViewContainer"], .main .block-container {{
        background: {bg_main} !important;
    }}
    .main {{
        background-color: transparent !important;
    }}
    h1, h2, h3, h4, h5, h6, p, .stMarkdown {{
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        color: {txt_main} !important;
    }}
    /* TODOS los botones: metálico azul brillante (#6ba3e0→#2568a0), letras blancas, uniforme en toda la app */
    /* Incluye: Refrescar, Browse files, form submit, download, etc. Gradiente con reflejos luminosos. */
    .stApp [data-testid="stButton"] > button,
    .stApp [data-testid="stFormSubmitButton"] > button,
    .stApp [data-testid="stDownloadButton"] > button,
    .stApp [data-testid="stDownloadButton"] a,
    .stApp [data-testid="stFileUploader"] [data-testid="stButton"] > button,
    .stApp [data-testid="stFileUploader"] button,
    .stApp [data-testid="stFileUploader"] a,
    .stApp [data-testid="stFileUploader"] [role="button"],
    .main .stButton > button,
    .main form [data-testid="stFormSubmitButton"] > button,
    .main [data-testid="stDownloadButton"] > button,
    .main [data-testid="stExpander"] .stButton > button,
    .main [data-testid="stExpander"] [data-testid="stVerticalBlock"] .stButton > button,
    .main [data-testid="stExpander"] [data-testid="stDownloadButton"] > button,
    .main [data-testid="stFileUploader"] [data-testid="stButton"] > button,
    .main [data-testid="stFileUploader"] button,
    .main [data-testid="stFileUploader"] a,
    .main [data-testid="stFileUploader"] [role="button"],
    .main [data-testid="column"] [data-testid="stButton"] > button {{
        width: 100%;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 1.15rem !important;
        padding: 0.85rem 1.5rem !important;
        min-height: 3rem !important;
        background: linear-gradient(180deg, #7eb3f0 0%, #99d0ff 18%, #6ba3e0 35%, #3d7ab8 55%, #2568a0 75%, #2d6a9f 100%) !important;
        color: #ffffff !important;
        text-shadow: 0 1px 2px rgba(0,0,0,0.3), 0 0 8px rgba(255,255,255,0.2) !important;
        border: 1px solid rgba(100,160,220,0.5) !important;
        border-radius: 12px !important;
        font-weight: bold !important;
        letter-spacing: 0.02em !important;
        text-align: center !important;
        box-shadow: inset 0 2px 6px rgba(255,255,255,0.4),
                    inset 0 -3px 6px rgba(0,0,0,0.3),
                    0 4px 12px rgba(0,0,0,0.4) !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease !important;
        text-decoration: none !important;
    }}
    .stApp [data-testid="stButton"] > button:hover,
    .stApp [data-testid="stFormSubmitButton"] > button:hover,
    .stApp [data-testid="stDownloadButton"] > button:hover,
    .stApp [data-testid="stDownloadButton"] a:hover,
    .stApp [data-testid="stFileUploader"] [data-testid="stButton"] > button:hover,
    .stApp [data-testid="stFileUploader"] button:hover,
    .stApp [data-testid="stFileUploader"] a:hover,
    .stApp [data-testid="stFileUploader"] [role="button"]:hover,
    .main .stButton > button:hover,
    .main form [data-testid="stFormSubmitButton"] > button:hover,
    .main [data-testid="stDownloadButton"] > button:hover,
    .main [data-testid="stExpander"] .stButton > button:hover,
    .main [data-testid="stExpander"] [data-testid="stVerticalBlock"] .stButton > button:hover,
    .main [data-testid="stExpander"] [data-testid="stDownloadButton"] > button:hover,
    .main [data-testid="stFileUploader"] [data-testid="stButton"] > button:hover,
    .main [data-testid="stFileUploader"] button:hover,
    .main [data-testid="stFileUploader"] a:hover,
    .main [data-testid="stFileUploader"] [role="button"]:hover,
    .main [data-testid="column"] [data-testid="stButton"] > button:hover {{
        background: linear-gradient(180deg, #8ec5ff 0%, #a8dcff 18%, #7eb3f0 35%, #4a8ac8 55%, #2d78b0 75%, #3a7ab8 100%) !important;
        transform: translateY(-3px) !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.4),
                    inset 0 -2px 4px rgba(0,0,0,0.2),
                    0 6px 20px rgba(0,0,0,0.5),
                    0 0 16px rgba(100,160,220,0.35) !important;
    }}
    .stApp [data-testid="stButton"] > button:active,
    .stApp [data-testid="stFormSubmitButton"] > button:active,
    .stApp [data-testid="stDownloadButton"] > button:active,
    .stApp [data-testid="stDownloadButton"] a:active,
    .stApp [data-testid="stFileUploader"] [data-testid="stButton"] > button:active,
    .stApp [data-testid="stFileUploader"] button:active,
    .stApp [data-testid="stFileUploader"] a:active,
    .stApp [data-testid="stFileUploader"] [role="button"]:active,
    .main .stButton > button:active,
    .main form [data-testid="stFormSubmitButton"] > button:active,
    .main [data-testid="stDownloadButton"] > button:active,
    .main [data-testid="stExpander"] .stButton > button:active,
    .main [data-testid="stExpander"] [data-testid="stVerticalBlock"] .stButton > button:active,
    .main [data-testid="stExpander"] [data-testid="stDownloadButton"] > button:active,
    .main [data-testid="stFileUploader"] [data-testid="stButton"] > button:active,
    .main [data-testid="stFileUploader"] button:active,
    .main [data-testid="stFileUploader"] a:active,
    .main [data-testid="stFileUploader"] [role="button"]:active,
    .main [data-testid="column"] [data-testid="stButton"] > button:active {{
        transform: translateY(1px) !important;
        box-shadow: inset 0 3px 6px rgba(0,0,0,0.35),
                    0 2px 6px rgba(0,0,0,0.4) !important;
    }}
    div[data-testid="stVerticalBlock"] > div {{
        background-color: transparent !important;
        border-radius: 10px !important;
        padding: 0.35rem 0.25rem !important;
        box-shadow: 0 4px 15px rgba(0,0,0,0.15) !important;
    }}
    .big-label {{
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 1.4rem !important;
        font-weight: bold !important;
        color: {txt_main} !important;
        background-color: transparent !important;
        padding: 0.4rem 0.6rem !important;
        border-radius: 8px !important;
    }}
    .total-box {{
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 2rem !important;
        font-weight: bold !important;
        padding: 1rem;
        background-color: transparent !important;
        color: {txt_main} !important;
        border-radius: 8px;
        text-align: center;
        box-shadow: 0 6px 20px rgba(0,0,0,0.2) !important;
    }}
    /* Etiquetas con $ (billetes/monedas): más grandes y centradas */
    div[data-testid="stNumberInput"] label {{
        font-size: 1.7rem !important;
        font-weight: bold !important;
        text-align: center !important;
        display: block !important;
        width: 100% !important;
    }}
    div[data-testid="stNumberInput"] > div {{
        display: flex !important;
        flex-direction: column !important;
        align-items: center !important;
    }}
    /* Campos de monto: sin bordes rojos, borde azul; al pasar mouse verde ganancia y elevación */
    .main input[type="number"],
    .main input[type="text"] {{
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 1.3rem !important;
        padding: 0.6rem !important;
        background-color: {input_bg} !important;
        color: {input_color} !important;
        border: 2px solid {AZUL_OSCURO} !important;
        border-radius: 8px !important;
        outline: none !important;
        box-shadow: 4px 4px 12px rgba(0,0,0,0.4),
                    2px 2px 6px rgba(0,0,0,0.3),
                    inset 0 1px 0 rgba(255,255,255,0.06) !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease, border-color 0.2s ease, background-color 0.2s ease !important;
    }}
    .main input[type="number"]:hover,
    .main input[type="text"]:hover {{
        transform: translateY(-3px) !important;
        border-color: #22c55e !important;
        background-color: rgba(34, 197, 94, 0.15) !important;
        box-shadow: 6px 8px 20px rgba(0,0,0,0.5),
                    3px 4px 10px rgba(0,0,0,0.4),
                    0 0 0 2px rgba(34, 197, 94, 0.3),
                    inset 0 1px 0 rgba(255,255,255,0.08) !important;
    }}
    .main input[type="number"]:focus,
    .main input[type="text"]:focus {{
        outline: none !important;
        border-color: {AZUL_OSCURO} !important;
        box-shadow: 5px 6px 16px rgba(0,0,0,0.45),
                    2px 3px 8px rgba(0,0,0,0.35),
                    inset 0 1px 0 rgba(255,255,255,0.08) !important;
    }}
    .main input[type="number"]:invalid,
    .main input[type="text"]:invalid {{
        border-color: {AZUL_OSCURO} !important;
        outline: none !important;
    }}
    /* Quitar cualquier borde/outline rojo que Streamlit aplique por defecto */
    .main div[data-testid="stNumberInput"] input {{
        border-color: {AZUL_OSCURO} !important;
        outline: none !important;
    }}
    .main div[data-testid="stNumberInput"] input:focus {{
        border-color: {AZUL_OSCURO} !important;
        outline: none !important;
    }}
    [data-testid="stSelectbox"] label, [data-testid="stRadio"] label {{
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        color: {txt_main} !important;
    }}
    div[data-testid="stNumberInput"] label {{
        color: {txt_main} !important;
    }}
    .stSuccess, .stAlert, .stException, .stWarning, [data-testid="stException"] {{
        background-color: transparent !important;
        color: {txt_main} !important;
        font-family: Calibri, 'Segoe UI', sans-serif !important;
    }}
    [data-testid="stExpander"] {{
        background-color: transparent !important;
        border-radius: 8px !important;
    }}
    [data-testid="stExpander"] summary {{
        color: {txt_main} !important;
        font-family: Calibri, 'Segoe UI', sans-serif !important;
    }}
    [data-testid="stVerticalBlock"] {{
        background-color: transparent !important;
    }}
    .main .stCaption, .main [data-testid="stCaption"] {{
        color: {txt_main} !important;
    }}
    .main label {{
        color: {txt_main} !important;
    }}
    /* Métricas (saldo, ingresos, gastos): color según tema */
    div[data-testid="metric-container"], div[data-testid="stMetric"] label, div[data-testid="stMetric"] div {{
        color: {txt_main} !important;
    }}
    /* Dataframe / tabla: texto según tema */
    .main div[data-testid="stDataFrame"] *, .main [data-testid="stDataFrame"] td, .main [data-testid="stDataFrame"] th {{
        color: {txt_main} !important;
    }}
    /* Menú lateral: ordenado, centrado, profesional */
    section[data-testid="stSidebar"] {{
        background: {sidebar_bg} !important;
        box-shadow: 4px 0 24px rgba(0,0,0,0.4) !important;
        padding: 0.5rem 0.6rem !important;
    }}
    section[data-testid="stSidebar"] .block-container {{
        padding: 0.25rem 0.35rem !important;
        max-width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {{
        width: 100% !important;
        max-width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stVerticalBlock"] > div {{
        padding: 0.25rem 0 !important;
        width: 100% !important;
        max-width: 100% !important;
    }}
    section[data-testid="stSidebar"] hr {{
        margin: 0.4rem 0 !important;
        border-color: {sidebar_txt_muted} !important;
        opacity: 0.5;
    }}
    section[data-testid="stSidebar"] div[data-testid="stImage"] {{
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        width: 100% !important;
        padding: 0.4rem 0.25rem !important;
        margin: 0.2rem 0 !important;
        border: 1px solid {sidebar_txt_muted} !important;
        border-radius: 10px !important;
        background: rgba(0,0,0,0.15) !important;
        box-sizing: border-box !important;
    }}
    section[data-testid="stSidebar"] div[data-testid="stImage"] img {{
        max-height: 100px !important;
        max-width: 100% !important;
        width: auto !important;
        height: auto !important;
        object-fit: contain !important;
    }}
    section[data-testid="stSidebar"] .stRadio > label,
    section[data-testid="stSidebar"] .stSelectbox label,
    section[data-testid="stSidebar"] label {{
        font-size: 0.85rem !important;
        color: {sidebar_txt} !important;
        text-align: left !important;
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] > div,
    section[data-testid="stSidebar"] [data-testid="stRadio"] > div {{
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] {{
        width: 100% !important;
        padding: 0.4rem 0.45rem !important;
        margin: 0.25rem 0 !important;
        background: rgba(0,0,0,0.35) !important;
        border-radius: 10px !important;
        box-shadow: inset 0 3px 10px rgba(0,0,0,0.5),
                    inset 0 -2px 4px rgba(0,0,0,0.2),
                    inset 0 1px 0 rgba(255,255,255,0.02) !important;
        border: 1px solid rgba(0,0,0,0.4) !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stSelectbox"] [data-baseweb="select"] {{
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stTextInput"] {{
        padding: 0.35rem 0 !important;
    }}
    section[data-testid="stSidebar"] [data-testid="stTextInput"] input {{
        width: 100% !important;
        box-sizing: border-box !important;
        background: rgba(0,0,0,0.4) !important;
        box-shadow: inset 0 3px 8px rgba(0,0,0,0.5),
                    inset 0 -2px 4px rgba(0,0,0,0.2) !important;
        border: 1px solid rgba(0,0,0,0.4) !important;
        border-radius: 8px !important;
    }}
    .menu-item {{
        display: block;
        width: 100%;
        padding: 0.6rem 0.9rem;
        margin: 0.2rem 0;
        border-radius: 12px;
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 1.1rem !important;
        font-weight: bold;
        color: {sidebar_txt};
        background: {menu_bg};
        border: 1px solid {sidebar_txt_muted};
        text-align: center;
        cursor: pointer;
        transition: all 0.25s ease;
        box-shadow: 0 2px 8px rgba(0,0,0,0.2);
    }}
    .menu-item:hover {{
        background: {menu_bg_hover};
        border-color: {sidebar_txt_muted};
        transform: translateX(4px);
        box-shadow: 0 4px 16px rgba(0,0,0,0.3);
    }}
    .menu-item.active {{
        background: {active_menu_bg};
        color: {active_menu_txt} !important;
        border-color: {sidebar_txt_muted};
        box-shadow: 0 4px 20px rgba(0,0,0,0.35);
    }}
    .menu-seccion {{
        font-size: 0.7rem;
        color: {sidebar_txt_muted};
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin: 0.35rem 0 0.2rem 0;
        padding: 0.35rem 0.5rem;
        text-align: center !important;
        width: 100%;
        max-width: 100%;
        display: block;
        box-sizing: border-box;
        line-height: 1.3;
    }}
    .menu-ministerio {{
        font-size: 1.05rem !important;
        font-weight: 600 !important;
        letter-spacing: 0.14em !important;
        text-align: center !important;
        margin: 0 0 0.3rem 0 !important;
        padding: 0.3rem 0.4rem !important;
        width: 100% !important;
        max-width: 100% !important;
        box-sizing: border-box !important;
        color: #c8d0d8 !important;
        text-shadow: 0 0 8px rgba(192,210,230,0.6),
                     0 0 2px rgba(255,255,255,0.4),
                     0 1px 3px rgba(0,0,0,0.4) !important;
    }}
    section[data-testid="stSidebar"] .stButton > button {{
        width: 100% !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        padding: 0.5rem 0.65rem !important;
        min-height: 2.4rem !important;
        border-radius: 10px !important;
        font-family: Calibri, 'Segoe UI', sans-serif !important;
        font-size: 0.95rem !important;
        font-weight: bold !important;
        color: #ffffff !important;
        text-shadow: 0 1px 2px rgba(0,0,0,0.4) !important;
        border: 1px solid rgba(100,160,220,0.5) !important;
        text-align: center !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease !important;
        background: linear-gradient(180deg, #6ba3e0 0%, #3d7ab8 30%, #2568a0 60%, #2d6a9f 100%) !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.35),
                    inset 0 -3px 6px rgba(0,0,0,0.35),
                    0 3px 8px rgba(0,0,0,0.4) !important;
    }}
    section[data-testid="stSidebar"] .stCaption,
    section[data-testid="stSidebar"] [data-testid="stMarkdown"] {{
        color: {sidebar_txt} !important;
        text-align: center !important;
        width: 100% !important;
    }}
    section[data-testid="stSidebar"] [data-testid="column"] {{
        min-width: 0 !important;
        flex: 1 1 0 !important;
        padding: 0 0.2rem !important;
    }}
    section[data-testid="stSidebar"] .stRadio [role="radiogroup"] {{
        display: flex !important;
        justify-content: center !important;
        align-items: center !important;
        gap: 0.5rem !important;
        padding: 0.45rem 0.55rem !important;
        background: rgba(0,0,0,0.35) !important;
        border-radius: 10px !important;
        box-shadow: inset 0 3px 10px rgba(0,0,0,0.5),
                    inset 0 -2px 5px rgba(0,0,0,0.25),
                    inset 0 1px 0 rgba(255,255,255,0.02) !important;
        border: 1px solid rgba(0,0,0,0.4) !important;
    }}
    section[data-testid="stSidebar"] .stRadio label {{
        margin: 0 !important;
    }}
    section[data-testid="stSidebar"] .stButton > button {{
        text-align: center !important;
        line-height: 1.2 !important;
    }}
    section[data-testid="stSidebar"] .stButton > button:hover {{
        background: linear-gradient(180deg, #7eb3f0 0%, #4a8ac8 30%, #2d78b0 60%, #3a7ab8 100%) !important;
        transform: translateX(3px) translateY(-1px) !important;
        box-shadow: inset 0 2px 4px rgba(255,255,255,0.45),
                    inset 0 -2px 4px rgba(0,0,0,0.25),
                    0 4px 12px rgba(0,0,0,0.45),
                    0 0 12px rgba(100,160,220,0.3) !important;
    }}
    @media (max-width: 768px) {{
        section[data-testid="stSidebar"] {{ padding: 0.25rem 0.35rem !important; }}
        section[data-testid="stSidebar"] div[data-testid="stImage"] {{ padding: 0.3rem 0.2rem !important; }}
        section[data-testid="stSidebar"] div[data-testid="stImage"] img {{ max-height: 85px !important; max-width: 100% !important; }}
        section[data-testid="stSidebar"] .stButton > button {{ padding: 0.45rem 0.5rem !important; min-height: 2.2rem !important; font-size: 0.9rem !important; }}
        section[data-testid="stSidebar"] hr {{ margin: 0.25rem 0 !important; }}
    }}
    </style>
    """, unsafe_allow_html=True)
    if st.session_state.get("tamano_fuente") == "grande":
        st.markdown("""<style>.main .block-container, .main p, .main .stMarkdown, .main label { font-size: 1.12rem !important; }</style>""", unsafe_allow_html=True)
    # Forzar desactivación de "Saved info" / autocompletado del navegador (Chrome, Edge, etc.)
    st.markdown("""
    <script>
    (function(){
      var meta = document.createElement('meta');
      meta.name = 'autocomplete';
      meta.content = 'off';
      if (document.head) document.head.appendChild(meta);
      function forceNoAutofill(el) {
        if (!el || el._autofillDisabled) return;
        el.setAttribute('autocomplete', 'off');
        el.setAttribute('autocapitalize', 'off');
        el.setAttribute('autocorrect', 'off');
        el.setAttribute('data-lpignore', 'true');
        el.setAttribute('data-form-type', 'other');
        el.setAttribute('name', 'field_' + Math.random().toString(36).slice(2));
        el._autofillDisabled = true;
      }
      function desactivarAutofill() {
        try {
          var inputs = document.querySelectorAll('input[type="number"], input[type="text"]');
          inputs.forEach(forceNoAutofill);
        } catch(e) {}
      }
      function run() {
        desactivarAutofill();
        try {
          var root = document.body || document.documentElement;
          var obs = new MutationObserver(function() { setTimeout(desactivarAutofill, 50); });
          obs.observe(root, { childList: true, subtree: true });
        } catch(e) {}
      }
      if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', run);
      } else {
        run();
      }
      [100, 300, 600, 1200, 2500].forEach(function(ms) { setTimeout(desactivarAutofill, ms); });
    })();
    </script>
    """, unsafe_allow_html=True)

    # ----- MENÚ LATERAL (última generación) -----
    usuario_actual = st.session_state.get("usuario_actual", "admin")
    data_permisos = cargar_permisos()
    lista_usuarios = list(data_permisos.get("usuarios", {}).keys()) or ["admin"]

    with st.sidebar:
        # Encabezado compacto (logo desde configuración maestro; posición editable)
        st.markdown(f"<p class='menu-ministerio'>{t['ministerio_finanzas']}</p>", unsafe_allow_html=True)
        ui_cfg_logo = cargar_ui_config() or {}
        logos_sidebar = ui_cfg_logo.get("logos") or {}
        pos_principal = logos_sidebar.get("posicion_principal") or "centro"
        principal_pct = logos_sidebar.get("principal_ancho_pct")
        if principal_pct not in (25, 50, 75, 100):
            principal_pct = 100
        st.markdown(f"<style>[data-testid=\"stSidebar\"] div[data-testid=\"stImage\"] img {{ max-width: {principal_pct}% !important; width: {principal_pct}% !important; height: auto !important; }}</style>", unsafe_allow_html=True)
        ruta_logo_principal = _resolver_ruta_logo(logos_sidebar.get("principal"), LOGO_PRINCIPAL)
        if os.path.isfile(ruta_logo_principal):
            if pos_principal == "izquierda":
                c1, c2, c3 = st.columns([2, 1, 1])
                with c1:
                    st.image(ruta_logo_principal, width="stretch")
            elif pos_principal == "derecha":
                c1, c2, c3 = st.columns([1, 1, 2])
                with c3:
                    st.image(ruta_logo_principal, width="stretch")
            else:
                c1, c2, c3 = st.columns([1, 2, 1])
                with c2:
                    st.image(ruta_logo_principal, width="stretch")
        st.markdown("---")
        # Usuario primero (quién usa la app) — decisión principal
        st.markdown(f"<p class='menu-seccion' style='text-align:center;'>{t['usuario_actual_menu']}</p>", unsafe_allow_html=True)
        sel_usuario = st.selectbox(
            t["quien_usa_app"],
            options=lista_usuarios,
            index=lista_usuarios.index(usuario_actual) if usuario_actual in lista_usuarios else 0,
            format_func=lambda uid: data_permisos["usuarios"].get(uid, {}).get("nombre", uid),
            key="sel_usuario"
        )
        if sel_usuario != usuario_actual:
            if sel_usuario == "admin" and _pin_admin_requerido() and not st.session_state.get("admin_autorizado"):
                pass
            else:
                st.session_state["usuario_actual"] = sel_usuario
                if sel_usuario != "admin":
                    st.session_state["admin_autorizado"] = False
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if sel_usuario == "admin" and _pin_admin_requerido() and not st.session_state.get("admin_autorizado"):
            pin_ingresado = st.text_input(t["pin_ingrese"], type="password", key="pin_admin_input")
            if st.button(t["entrar"], key="btn_pin_entrar") and pin_ingresado:
                if _verificar_pin_admin(pin_ingresado):
                    st.session_state["admin_autorizado"] = True
                    st.session_state["usuario_actual"] = "admin"
                    audit_log("admin", "login_admin", "")
                    st.rerun()
                else:
                    st.error(t["pin_incorrecto"])
        st.markdown("---")
        # Navegación: orden por frecuencia de uso (Inicio → operaciones diarias → reportes → admin)
        # Config maestro: qué botones del menú se ocultan para TODOS (incluido admin)
        ui_cfg_sidebar = cargar_ui_config() or {}
        ocultar_sidebar = ui_cfg_sidebar.get("ocultar_botones") or {}
        def _menu_visible(clave: str) -> bool:
            return not bool(ocultar_sidebar.get(clave))
        es_maestro_sidebar = bool(st.session_state.get("es_acceso_maestro") or ES_PC_MAESTRO)

        st.markdown(f"<p class='menu-seccion' style='text-align:center;'>{t['menu_navegacion']}</p>", unsafe_allow_html=True)
        if _menu_visible("inicio") and tiene_permiso(usuario_actual, "ver_inicio"):
            if st.button(f"🏠 {t['inicio']}", key="btn_inicio", width="stretch"):
                st.session_state["pagina"] = "inicio"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("arqueo_caja") and tiene_permiso(usuario_actual, "ver_arqueo_caja"):
            if st.button(f"📋 {t['arqueo_caja']}", key="btn_arqueo", width="stretch"):
                st.session_state["pagina"] = "arqueo_caja"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("tesoreria") and tiene_permiso(usuario_actual, "ver_tesoreria"):
            if st.button(f"📒 {t['tesoreria']}", key="btn_tesoreria", width="stretch"):
                st.session_state["pagina"] = "tesoreria"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("contabilidad") and tiene_permiso(usuario_actual, "ver_contabilidad"):
            if st.button(f"📊 {t['contabilidad']}", key="btn_contabilidad", width="stretch"):
                st.session_state["pagina"] = "contabilidad"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("presupuesto_metas") and tiene_permiso(usuario_actual, "ver_presupuesto_metas"):
            if st.button(f"🎯 {t['presupuesto_metas']}", key="btn_presupuesto", width="stretch"):
                st.session_state["pagina"] = "presupuesto_metas"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("eventos") and tiene_permiso(usuario_actual, "ver_eventos_inversiones"):
            if st.button(f"📌 {t.get('eventos_inversiones', 'Eventos / Inversiones')}", key="btn_eventos", width="stretch"):
                st.session_state["pagina"] = "eventos"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        if _menu_visible("administracion") and usuario_actual == "admin":
            if st.button(f"⚙️ {t['administracion']}", key="btn_admin", width="stretch"):
                st.session_state["pagina"] = "administracion"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        # Editar (personalización): solo visible con clave maestra de sesión (contraseña universal)
        if st.session_state.get("es_acceso_maestro"):
            if st.button(f"✏️ {t.get('ui_config_menu', 'Editar')}", key="btn_ui_config", width="stretch"):
                st.session_state["pagina"] = "ui_config"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        st.markdown("---")
        # Modo mantenimiento: solo admin puede pausar/reanudar desde aquí (sin ir a Administración)
        if usuario_actual == "admin":
            mant_activo = _mantenimiento_activo()
            st.markdown(f"<p class='menu-seccion' style='text-align:center;'>⏸️ {t.get('mantenimiento_sidebar', 'Pausar sistema')}</p>", unsafe_allow_html=True)
            if mant_activo:
                if st.button(f"▶️ {t.get('mantenimiento_btn_desactivar', 'Reanudar')}", key="sidebar_mant_off", width="stretch"):
                    _set_mantenimiento_activo(False)
                    audit_log(usuario_actual, "mantenimiento_desactivado", "sidebar")
                    st.rerun()
            else:
                if st.button(f"⏸️ {t.get('mantenimiento_btn_activar', 'Pausar (mantenimiento)')}", key="sidebar_mant_on", width="stretch"):
                    _set_mantenimiento_activo(True)
                    audit_log(usuario_actual, "mantenimiento_activado", "sidebar")
                    st.rerun()
            st.markdown("---")
        # Preferencias en fila compacta (idioma + tema)
        st.markdown("<p class='menu-seccion' style='text-align:center;'>Idioma · Tema</p>", unsafe_allow_html=True)
        col_idioma, col_tema = st.columns(2)
        with col_idioma:
            lang = st.radio(t["idioma"], options=["ES", "EN"], format_func=lambda x: "ES" if x == "ES" else "EN",
                           key="idioma", horizontal=True, label_visibility="collapsed")
        t = TEXTOS[lang]
        ministerios = MINISTERIOS if lang == "ES" else MINISTERIOS_EN
        with col_tema:
            tema_sel = st.radio("Tema", [t["tema_oscuro"], t["tema_claro"]], key="sel_tema_app", horizontal=True,
                                index=0 if st.session_state.get("tema_app") == "oscuro" else 1,
                                format_func=lambda x: "🌙" if "oscuro" in x.lower() or "dark" in x.lower() else "☀️",
                                label_visibility="collapsed")
        nuevo_tema = "claro" if tema_sel == t["tema_claro"] else "oscuro"
        if nuevo_tema != st.session_state.get("tema_app"):
            st.session_state["tema_app"] = nuevo_tema
            st.rerun()
        tamano_fuente = st.selectbox(t["tamano_texto"], [t["tamano_normal"], t["tamano_grande"]], key="sel_tamano_fuente",
                                     index=0 if st.session_state.get("tamano_fuente") != "grande" else 1)
        nuevo_tamano = "grande" if tamano_fuente == t["tamano_grande"] else "normal"
        if nuevo_tamano != st.session_state.get("tamano_fuente"):
            st.session_state["tamano_fuente"] = nuevo_tamano
            st.rerun()
        st.markdown("---")
        # Cerrar sesión al final (zona de salida, flujo natural)
        if st.button(f"🚪 {t['cerrar_sesion']}", key="btn_cerrar_sesion", width="stretch"):
            audit_log(usuario_actual, "cerrar_sesion", "")
            st.session_state["logueado"] = False
            st.session_state["admin_autorizado"] = False
            st.session_state["sidebar_state"] = "collapsed"
            st.session_state["sidebar_collapse_requested"] = True
            st.rerun()
        st.caption(f"{t['version']} {VERSION_APP}")

    # ----- AVISO MODO MANTENIMIENTO: no-admin solo ven el aviso (congelados), admin/maestro sigue trabajando -----
    usuario_actual = st.session_state.get("usuario_actual", "")
    es_admin_o_maestro = (usuario_actual == "admin") or st.session_state.get("es_acceso_maestro", False)
    if not es_admin_o_maestro and _mantenimiento_activo():
        st.markdown(f"""
        <div style="background: linear-gradient(135deg, #1e3a5f 0%, #2d4a6f 100%); border: 2px solid #3b82f6; border-radius: 10px; padding: 1.5rem 1.5rem; margin: 2rem 0;">
        <p style="color: #93c5fd; margin: 0; font-weight: bold; font-size: 1.2rem;">⏸️ {t.get('mantenimiento_titulo', 'Sistema en actualización')}</p>
        <p style="color: #bfdbfe; margin: 0.75rem 0 0 0; font-size: 1rem; line-height: 1.5;">{t.get('mantenimiento_aviso', 'El administrador está realizando actualizaciones. Evite ingresar datos nuevos mientras vea este aviso. Si cierra y vuelve a abrir, podrá continuar; su trabajo guardado se mantiene visible.')}</p>
        <p style="color: #93c5fd; margin: 1rem 0 0 0; font-size: 0.9rem;">{t.get('mantenimiento_esperar', 'Cuando el administrador desactive el mantenimiento, recargue la página para continuar. Lo que ya guardó no se pierde.')}</p>
        <p style="color: #93c5fd; margin: 0.75rem 0 0 0; font-size: 0.9rem;">{t.get('mantenimiento_guardado_visible', 'Lo que guardó sigue guardado y lo verá al reanudar el sistema.')}</p>
        </div>
        """, unsafe_allow_html=True)
        return

    # ----- CONTENIDO PRINCIPAL SEGÚN PÁGINA -----
    if st.session_state["pagina"] == "administracion":
        # Estilo: casillas pequeñas, verde = tiene permiso, rojo = no tiene (colores según tema)
        st.markdown(f"""
        <style>
        /* Permisos: 5 columnas por fila, ocupan todo el ancho sin huecos vacíos */
        .main [data-testid="stExpander"] [data-testid="column"] {{ min-width: 0 !important; flex: 1 1 0 !important; padding: 0.2rem !important; }}
        /* Cuadraditos de permisos: marcado verde, desmarcado rojo */
        div[data-testid="stCheckbox"] > label {{
            padding: 0.35rem 0.5rem !important;
            border-radius: 6px !important;
            border: 1px solid {checkbox_border} !important;
            min-height: 2rem !important;
            font-size: 0.85rem !important;
        }}
        div[data-testid="stCheckbox"]:has(input:not(:checked)) > label {{
            background: rgba(239, 68, 68, 0.2) !important;
            color: {checkbox_unchecked_color} !important;
            border-color: {checkbox_unchecked_border} !important;
        }}
        div[data-testid="stCheckbox"]:has(input:checked) > label {{
            background: rgba(34, 197, 94, 0.2) !important;
            color: {checkbox_checked_color} !important;
            border-color: {checkbox_checked_border} !important;
        }}
        div[data-testid="stCheckbox"] input[type="checkbox"] {{
            width: 1.1rem !important; height: 1.1rem !important;
            accent-color: {checkbox_checked_border} !important;
        }}
        @media (max-width: 768px) {{
            .main [data-testid="stExpander"] [data-testid="column"] {{ padding: 0.15rem !important; }}
        }}
        </style>
        """, unsafe_allow_html=True)
        st.markdown(f"## ⚙️ {t['administracion_titulo']}")
        st.markdown(t["admin_instrucciones"])

        # --- Modo mantenimiento (pausar sistema para actualizaciones) ---
        mant_activo = _mantenimiento_activo()
        with st.expander(f"⏸️ {t.get('mantenimiento_expander', 'Pausar sistema (mantenimiento)')}", expanded=mant_activo):
            st.caption(t.get("mantenimiento_ayuda", "Al activar, los demás usuarios verán un aviso de que el sistema está en actualización. Pueden cerrar y volver; su trabajo guardado se mantiene. Desactive cuando termine el mantenimiento."))
            if mant_activo:
                if st.button(f"▶️ {t.get('mantenimiento_btn_desactivar', 'Reanudar sistema')}", key="btn_mant_desactivar"):
                    _set_mantenimiento_activo(False)
                    audit_log(usuario_actual, "mantenimiento_desactivado", "")
                    st.success(t.get("mantenimiento_desactivado_ok", "Sistema reanudado. Los usuarios ya no verán el aviso."))
                    st.rerun()
                st.caption(f"🟠 {t.get('mantenimiento_activo_ahora', 'Modo mantenimiento activo.')}")
            else:
                if st.button(f"⏸️ {t.get('mantenimiento_btn_activar', 'Pausar sistema (mostrar aviso a usuarios)')}", key="btn_mant_activar"):
                    _set_mantenimiento_activo(True)
                    audit_log(usuario_actual, "mantenimiento_activado", "")
                    st.success(t.get("mantenimiento_activado_ok", "Aviso de mantenimiento activado. Los usuarios verán el mensaje al entrar."))
                    st.rerun()

        data_permisos = cargar_permisos()
        usuarios = data_permisos.get("usuarios", {})

        # --- Resumen de permisos (matriz) ---
        with st.expander(f"📊 {t.get('admin_resumen_permisos', 'RESUMEN DE PERMISOS')}", expanded=False):
            permisos_lista_resumen = list(PERMISOS_DISPONIBLES)
            header = "| Usuario |" + "|".join([e[:12] for _, e in permisos_lista_resumen]) + "|"
            sep = "|---|" + "|".join(["---" for _ in permisos_lista_resumen]) + "|"
            filas_md = [header, sep]
            for uid_r, info_r in usuarios.items():
                perms_r = set(info_r.get("permisos", []))
                todo = "*" in perms_r
                nombre_r = info_r.get("nombre", uid_r)[:15]
                celdas = []
                for pk, _ in permisos_lista_resumen:
                    celdas.append("✅" if todo or pk in perms_r else "❌")
                filas_md.append(f"| **{nombre_r}** |" + "|".join(celdas) + "|")
            st.markdown("\n".join(filas_md))

        # --- Añadir usuario ---
        with st.expander(t["admin_anadir_usuario"], expanded=False):
            col_a, col_b = st.columns(2)
            with col_a:
                nuevo_id = st.text_input(t["admin_id_placeholder"], key="nuevo_id_user", max_chars=30).strip().lower().replace(" ", "_")
                nuevo_nombre = st.text_input(t["admin_nombre_placeholder"], key="nuevo_nombre_user", max_chars=50).strip() or nuevo_id
            if st.button(t["admin_btn_anadir"], key="btn_add_user") and nuevo_id and nuevo_id != "admin":
                if nuevo_id in usuarios:
                    st.warning(t["admin_id_ya_existe"])
                else:
                    data_permisos["usuarios"][nuevo_id] = {"nombre": nuevo_nombre or nuevo_id, "permisos": []}
                    if guardar_permisos(data_permisos, t):
                        st.session_state["usuario_recien_anadido"] = nuevo_id
                        st.success(t["admin_usuario_anadido"].format(nombre=nuevo_nombre or nuevo_id))
                        st.rerun()

        # --- Descargar permisos (respaldo JSON) ---
        with st.expander(f"💾 {t.get('admin_descargar_permisos', 'Descargar permisos')}", expanded=False):
            try:
                permisos_json = json.dumps(data_permisos, indent=2, ensure_ascii=False)
                st.download_button(
                    label=t.get("admin_descargar_permisos", "Descargar permisos (JSON)"),
                    data=permisos_json,
                    file_name=f"DB_PERMISOS_respaldo_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                    mime="application/json",
                    key="btn_descargar_permisos"
                )
            except Exception:
                st.caption("Error al preparar respaldo.")

        st.markdown("---")
        data_permisos = cargar_permisos()
        usuarios = data_permisos.get("usuarios", {})
        usuario_recien_anadido = st.session_state.pop("usuario_recien_anadido", None)
        if usuario_recien_anadido and usuario_recien_anadido in usuarios:
            nombre_show = usuarios[usuario_recien_anadido].get("nombre", usuario_recien_anadido)
            st.info("✅ " + t["admin_asignar_permisos"].format(nombre=nombre_show))

        # Perfiles predefinidos para plantillas
        _perfiles_plantilla = {
            "Tesorero": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"],
            "Pastor": ["ver_inicio", "ver_arqueo_caja", "ver_tesoreria", "ver_contabilidad", "ver_presupuesto_metas", "ver_ingresar_bendicion", "ver_registrar_gasto", "ver_hoja_contable", "ver_informe_pdf", "ver_exportar_hoja_pdf"],
            "Asistente": ["ver_inicio", "ver_arqueo_caja"],
            "Ministerio": ["ver_inicio", "ver_arqueo_caja", "ver_hoja_contable", "ver_informe_pdf"],
        }

        for uid, info in usuarios.items():
            nombre = info.get("nombre", uid)
            permisos_usuario = set(info.get("permisos", []))
            es_admin = "*" in permisos_usuario or uid == "admin"
            expandir = not es_admin or uid == usuario_recien_anadido
            # Indicador sin contraseña + última actividad
            _tiene_pwd = bool((info.get("contrasena") or "").strip())
            _ult_fecha, _ult_accion = ultima_actividad_usuario(uid)
            badge = ""
            if not es_admin and not _tiene_pwd:
                badge += f" {t.get('admin_sin_contrasena', '⚠️ Sin contraseña')}"
            if _ult_fecha:
                badge += f" | {t.get('admin_ultima_actividad', 'Última actividad')}: {_ult_fecha[:16]}"
            with st.expander(f"**{nombre}** ({uid})" + (t["admin_expander_admin"] if es_admin else "") + badge, expanded=expandir):
                if es_admin:
                    st.caption(t["admin_caption_admin"])
                else:
                    # Persona responsable
                    resp = info.get("responsable", {}) or {}
                    st.markdown(f"**{t['admin_responsable_titulo']}**")
                    col_r1, col_r2, col_r3 = st.columns([2, 2, 1])
                    with col_r1:
                        resp_nombre = st.text_input(t["admin_responsable_nombre"], value=resp.get("nombre_completo", ""), key=f"resp_nombre_{uid}", max_chars=80, label_visibility="collapsed", placeholder=t["admin_responsable_nombre"])
                    with col_r2:
                        resp_licencia = st.text_input(t["admin_responsable_licencia"], value=resp.get("licencia_or_id", ""), key=f"resp_licencia_{uid}", max_chars=50, label_visibility="collapsed", placeholder=t["admin_responsable_licencia"])
                    with col_r3:
                        btn_resp = st.button(t["admin_guardar_responsable"], key=f"btn_resp_{uid}", width="stretch")
                    if btn_resp:
                        data_p = cargar_permisos()
                        if uid in data_p.get("usuarios", {}):
                            if "responsable" not in data_p["usuarios"][uid]:
                                data_p["usuarios"][uid]["responsable"] = {}
                            data_p["usuarios"][uid]["responsable"]["nombre_completo"] = (resp_nombre or "").strip()
                            data_p["usuarios"][uid]["responsable"]["licencia_or_id"] = (resp_licencia or "").strip()
                            if guardar_permisos(data_p, t):
                                audit_log(usuario_actual, "responsable_actualizado", f"{uid}: {resp_nombre or '-'}")
                                st.success(t["admin_responsable_guardado"])
                                st.rerun()
                    if resp.get("nombre_completo") or resp.get("licencia_or_id"):
                        st.caption(f"✓ {resp.get('nombre_completo', '-')} | {t['admin_responsable_licencia']}: {resp.get('licencia_or_id', '-')}")
                    st.markdown("---")

                    # Plantilla de permisos + copiar de otro usuario
                    col_pl, col_cp = st.columns(2)
                    with col_pl:
                        perfil_sel = st.selectbox(t.get("admin_plantilla", "Aplicar perfil"), options=[""] + list(_perfiles_plantilla.keys()), key=f"perfil_{uid}", help=t.get("admin_plantilla_ayuda", ""))
                        if perfil_sel and st.button(f"✓ {t.get('admin_plantilla', 'Aplicar')}", key=f"btn_perfil_{uid}"):
                            data_p = cargar_permisos()
                            if uid in data_p.get("usuarios", {}):
                                data_p["usuarios"][uid]["permisos"] = list(_perfiles_plantilla[perfil_sel])
                                if guardar_permisos(data_p, t):
                                    audit_log(usuario_actual, "perfil_aplicado", f"{uid} -> {perfil_sel}")
                                    st.rerun()
                    with col_cp:
                        otros = [u for u in usuarios if u != uid and u != "admin"]
                        if otros:
                            copiar_de = st.selectbox(t.get("admin_copiar_permisos", "Copiar de"), options=[""] + otros, format_func=lambda x: usuarios.get(x, {}).get("nombre", x) if x else "—", key=f"copiar_{uid}")
                            if copiar_de and st.button(f"📋 {t.get('admin_copiar_permisos', 'Copiar')}", key=f"btn_copiar_{uid}"):
                                data_p = cargar_permisos()
                                if uid in data_p.get("usuarios", {}) and copiar_de in data_p.get("usuarios", {}):
                                    data_p["usuarios"][uid]["permisos"] = list(data_p["usuarios"][copiar_de].get("permisos", []))
                                    if guardar_permisos(data_p, t):
                                        audit_log(usuario_actual, "permisos_copiados", f"{copiar_de} -> {uid}")
                                        st.success(t.get("admin_copiar_ok", "Copiados.").format(origen=usuarios.get(copiar_de, {}).get("nombre", copiar_de)))
                                        st.rerun()
                    st.markdown("---")

                    # Permisos: 5 columnas
                    permisos_lista = list(PERMISOS_DISPONIBLES)
                    num_cols = 5
                    for fila in range(0, len(permisos_lista), num_cols):
                        cols = st.columns(num_cols)
                        for i, col in enumerate(cols):
                            idx = fila + i
                            if idx >= len(permisos_lista):
                                break
                            permiso_clave, permiso_etiqueta = permisos_lista[idx]
                            tiene = tiene_permiso(uid, permiso_clave)
                            key = f"perm_{uid}_{permiso_clave}"
                            nuevo_valor = st.checkbox(permiso_etiqueta, value=tiene, key=key)
                            if nuevo_valor != tiene:
                                toggle_permiso(uid, permiso_clave, t)
                                st.rerun()
                    st.caption(t["admin_caption_verde_rojo"])

                    # Eliminar usuario (nunca admin)
                    st.markdown("---")
                    if st.session_state.get(f"confirmar_eliminar_{uid}"):
                        st.warning(t.get("admin_confirmar_eliminar", "Escriba ELIMINAR para confirmar."))
                        conf_el = st.text_input("", key=f"input_eliminar_{uid}", placeholder="ELIMINAR" if lang == "ES" else "DELETE")
                        col_el1, col_el2 = st.columns(2)
                        with col_el1:
                            if st.button("✓", key=f"btn_conf_eliminar_{uid}"):
                                palabra = "ELIMINAR" if lang == "ES" else "DELETE"
                                if (conf_el or "").strip().upper() == palabra:
                                    data_p = cargar_permisos()
                                    if uid in data_p.get("usuarios", {}):
                                        del data_p["usuarios"][uid]
                                        if guardar_permisos(data_p, t):
                                            audit_log(usuario_actual, "usuario_eliminado", uid)
                                            st.success(t.get("admin_usuario_eliminado", "Eliminado.").format(nombre=nombre))
                                            st.session_state.pop(f"confirmar_eliminar_{uid}", None)
                                            st.rerun()
                        with col_el2:
                            if st.button("✕", key=f"btn_cancel_eliminar_{uid}"):
                                st.session_state.pop(f"confirmar_eliminar_{uid}", None)
                                st.rerun()
                    else:
                        if st.button(t.get("admin_eliminar_usuario", "🗑️ Eliminar"), key=f"btn_eliminar_{uid}", width="stretch"):
                            st.session_state[f"confirmar_eliminar_{uid}"] = True
                            st.rerun()

                # Contraseña (todos los usuarios)
                pwd_nueva = st.text_input(t["admin_contrasena"], value="", type="password", key=f"pwd_{uid}", placeholder=t["admin_contrasena_placeholder"])
                col_pwd1, col_pwd2 = st.columns(2)
                with col_pwd1:
                    if st.button(t["admin_guardar_contrasena"], key=f"btn_pwd_{uid}"):
                        data_p = cargar_permisos()
                        if uid in data_p.get("usuarios", {}):
                            pwd_plain = (pwd_nueva or "").strip()
                            if pwd_plain:
                                ok_pol, msg_pol = _validar_politica_contrasena(pwd_plain)
                                if not ok_pol:
                                    st.warning(msg_pol)
                                else:
                                    data_p["usuarios"][uid]["contrasena"] = _hash_contrasena(pwd_plain)
                                    if guardar_permisos(data_p, t):
                                        audit_log(usuario_actual, "contrasena_actualizada", uid)
                                        st.success(t["admin_contrasena_guardada"])
                                        st.rerun()
                            else:
                                data_p["usuarios"][uid]["contrasena"] = ""
                                if guardar_permisos(data_p, t):
                                    audit_log(usuario_actual, "contrasena_actualizada", uid)
                                    st.success(t["admin_contrasena_guardada"])
                                    st.rerun()
                with col_pwd2:
                    if st.button(t["admin_reset_contrasena"], key=f"btn_reset_pwd_{uid}"):
                        data_p = cargar_permisos()
                        if uid in data_p.get("usuarios", {}):
                            data_p["usuarios"][uid]["contrasena"] = ""
                            if guardar_permisos(data_p, t):
                                audit_log(usuario_actual, "contrasena_reseteada", uid)
                                st.success("Contraseña restablecida a predeterminada." if lang == "ES" else "Password reset to default.")
                                st.rerun()
                # Pregunta de seguridad para recuperar contraseña
                st.caption(t.get("admin_pregunta_seguridad_titulo", "Pregunta de seguridad (recuperar contraseña)"))
                preguntas_opciones = ["recuperar_pregunta_1", "recuperar_pregunta_2", "recuperar_pregunta_3", "recuperar_pregunta_4", "recuperar_pregunta_5", "recuperar_pregunta_6"]
                pregunta_actual = data_permisos.get("usuarios", {}).get(uid, {}).get("pregunta_seguridad", "")
                idx_sel = preguntas_opciones.index(pregunta_actual) if pregunta_actual in preguntas_opciones else 0
                pregunta_sel = st.selectbox(
                    t.get("admin_pregunta_seguridad", "Pregunta"),
                    options=preguntas_opciones,
                    format_func=lambda k: t.get(k, k),
                    index=idx_sel if pregunta_actual in preguntas_opciones else 0,
                    key=f"pregunta_sel_{uid}",
                )
                respuesta_seg = st.text_input(
                    t.get("recuperar_respuesta", "Respuesta (para recuperar contraseña)"),
                    value="",
                    type="password",
                    key=f"respuesta_seg_{uid}",
                    placeholder=t.get("admin_respuesta_placeholder", "Solo el admin la conoce"),
                )
                if st.button(t.get("admin_guardar_pregunta", "Guardar pregunta y respuesta"), key=f"btn_guardar_pregunta_{uid}"):
                    if respuesta_seg and respuesta_seg.strip():
                        data_p = cargar_permisos()
                        if uid in data_p.get("usuarios", {}):
                            data_p["usuarios"][uid]["pregunta_seguridad"] = pregunta_sel
                            data_p["usuarios"][uid]["respuesta_seguridad_hash"] = _hash_contrasena(respuesta_seg.strip())
                            if guardar_permisos(data_p, t):
                                audit_log(usuario_actual, "pregunta_seguridad_actualizada", uid)
                                st.success(t.get("admin_pregunta_guardada", "Pregunta de seguridad guardada."))
                                st.rerun()
                    else:
                        st.warning(t.get("admin_respuesta_requerida", "Escriba la respuesta para guardar."))
        st.markdown("---")
        if usuario_actual == "admin" and ES_PC_MAESTRO:
            with st.expander(f"🔄 {t['reiniciar_tesoreria']}", expanded=False):
                st.caption(t["reiniciar_explicacion"])
                confirmacion = st.text_input(f"{t['confirmar_escribir']} **{CONFIRMACION_REINICIO}** {t['para_confirmar']}", key="confirm_reinicio")
                if st.button(t["admin_btn_reiniciar"], key="btn_reiniciar_tesoreria", type="secondary"):
                    if (confirmacion or "").strip().upper() == CONFIRMACION_REINICIO:
                        if reiniciar_tesoreria_master():
                            audit_log(usuario_actual, "reinicio_tesoreria", "")
                            st.success(t["reinicio_ok"])
                            st.rerun()
                        else:
                            st.error(t["admin_error_reinicio"])
                    else:
                        st.warning(t["admin_debe_escribir"].format(palabra=CONFIRMACION_REINICIO))
            with st.expander(f"⏪ {t['retroceder_historial']}", expanded=False):
                st.caption(t["retroceder_explicacion"])
                respaldos = listar_respaldos()
                if not respaldos:
                    st.info(t["sin_respaldos"])
                else:
                    def _fmt_respaldo(i):
                        try:
                            return respaldos[i][1] if i < len(respaldos) and len(respaldos[i]) > 1 else str(i)
                        except (IndexError, TypeError):
                            return str(i)
                    sel = st.selectbox(t["seleccionar_respaldo"], options=range(len(respaldos)), format_func=_fmt_respaldo, key="sel_respaldo")
                    if st.button(t["restaurar"], key="btn_restaurar_respaldo"):
                        ruta_rest = respaldos[sel][0] if 0 <= sel < len(respaldos) and respaldos[sel] else None
                        if ruta_rest and restaurar_respaldo(ruta_rest):
                            audit_log(usuario_actual, "restaurar_respaldo", ruta_rest)
                            st.success(t["restaurado_ok"])
                            st.rerun()
                        else:
                            st.error(t["admin_error_restaurar"])
        # Acceso maestro: restablecer admin a admin/admin (solo cuando se ingresó con contraseña universal)
        if st.session_state.get("es_acceso_maestro") and usuario_actual == "admin":
            with st.expander(t["acceso_maestro"], expanded=False):
                st.caption(t["acceso_maestro_info"])
                if st.button(t["restablecer_admin_admin"], key="btn_restablecer_admin_admin"):
                    data = cargar_permisos()
                    if "admin" in data.get("usuarios", {}):
                        data["usuarios"]["admin"]["contrasena"] = ""
                        if guardar_permisos(data, t):
                            audit_log("admin", "admin_restablecido_admin_admin", "por acceso maestro")
                            st.success("Admin restablecido a admin/admin. Cierre sesión y pruebe." if lang == "ES" else "Admin reset to admin/admin. Log out and try.")
                            st.rerun()
        # PIN de administrador: establecer o cambiar
        with st.expander(t["cambiar_pin"] if _pin_admin_requerido() else t["establecer_pin"], expanded=False):
            pin_actual = st.text_input(t["pin_actual"], type="password", key="admin_pin_actual")
            pin_nuevo = st.text_input(t["pin_nuevo"], type="password", key="admin_pin_nuevo")
            if st.button("Guardar PIN", key="btn_guardar_pin") and pin_nuevo:
                if _pin_admin_requerido() and not _verificar_pin_admin(pin_actual):
                    st.error(t["pin_incorrecto"])
                else:
                    try:
                        data = cargar_permisos()
                        data["admin_pin_hash"] = _hash_pin(pin_nuevo)
                        with open(DB_PERMISOS, "w", encoding="utf-8") as f:
                            json.dump(data, f, indent=2, ensure_ascii=False)
                        st.session_state["admin_autorizado"] = True
                        st.success(t["pin_guardado"])
                        audit_log("admin", "pin_actualizado", "")
                    except Exception as e:
                        st.error(str(e))
        # --- Rastreo de movimientos por usuario (solo clave maestra/universal) ---
        if st.session_state.get("es_acceso_maestro"):
            st.markdown("---")
            with st.expander(f"🔍 {t.get('admin_rastreo_titulo', 'RASTREO DE MOVIMIENTOS')}", expanded=False):
                st.caption(t.get("admin_rastreo_sub", "Registro de acciones por usuario (solo clave maestra)."))
                movimientos = movimientos_por_usuario(limite_por_usuario=30)
                if not movimientos:
                    st.info(t.get("admin_sin_movimientos", "Sin movimientos registrados."))
                else:
                    usuarios_mov = list(movimientos.keys())
                    filtro_usr = st.selectbox("Usuario", options=["— Todos —"] + usuarios_mov, key="filtro_rastreo_usr")
                    if filtro_usr == "— Todos —":
                        uids_mostrar = usuarios_mov
                    else:
                        uids_mostrar = [filtro_usr]
                    for uid_m in uids_mostrar:
                        nombre_m = usuarios.get(uid_m, {}).get("nombre", uid_m)
                        st.markdown(f"### {nombre_m} (`{uid_m}`)")
                        registros = movimientos.get(uid_m, [])
                        if not registros:
                            st.caption(t.get("admin_sin_movimientos", "Sin movimientos."))
                        else:
                            filas_tabla = []
                            for fecha_m, accion_m, detalle_m in registros:
                                filas_tabla.append({"Fecha": fecha_m, "Acción": accion_m, "Detalle": (detalle_m or "")[:80]})
                            st.dataframe(pd.DataFrame(filas_tabla), width="stretch", hide_index=True)
        st.markdown("---")
        if st.button(t["volver_inicio"]):
            st.session_state["pagina"] = "inicio"
            st.rerun()
        return

    if st.session_state["pagina"] == "inicio":
        # Logo de pantalla de Inicio (editable en Editar; posición y % de pantalla)
        ui_cfg_inicio_logo = cargar_ui_config() or {}
        logos_inicio = ui_cfg_inicio_logo.get("logos") or {}
        pos_inicio = logos_inicio.get("posicion_inicio") or "centro"
        inicio_pct = logos_inicio.get("inicio_ancho_pct")
        if inicio_pct not in (25, 50, 75, 100):
            inicio_pct = 100
        ruta_imagen = _resolver_ruta_logo(logos_inicio.get("inicio"), IMAGEN_INICIO_ES)
        if not os.path.isfile(ruta_imagen):
            ruta_imagen = _resolver_ruta_logo(logos_inicio.get("principal"), LOGO_PRINCIPAL)
        if not os.path.isfile(ruta_imagen):
            ruta_imagen = IMAGEN_INICIO_ES if os.path.isfile(IMAGEN_INICIO_ES) else (IMAGEN_INICIO_FALLBACK if os.path.isfile(IMAGEN_INICIO_FALLBACK) else None)
        if ruta_imagen and os.path.isfile(ruta_imagen):
            st.markdown(
                f"<style>.main .block-container div[data-testid=\"stImage\"] img {{ max-width: {inicio_pct}% !important; width: {inicio_pct}% !important; height: auto !important; object-fit: contain !important; }}</style>",
                unsafe_allow_html=True,
            )
            if pos_inicio == "izquierda":
                col_izq, col_c, col_d = st.columns([2, 1, 1])
                with col_izq:
                    st.image(ruta_imagen, width="stretch")
            elif pos_inicio == "derecha":
                col_izq, col_c, col_d = st.columns([1, 1, 2])
                with col_d:
                    st.image(ruta_imagen, width="stretch")
            else:
                col_izq, col_logo, col_d = st.columns([1, 1, 1])
                with col_logo:
                    st.image(ruta_imagen, width="stretch")
        # Estilo: sin bordes, imagen flotando; botones respetan tema
        st.markdown(f"""
        <style>
        .stApp, [data-testid="stAppViewContainer"], .main, .main .block-container {{
            border: none !important; outline: none !important; box-shadow: none !important;
            padding-left: 0.35rem !important; padding-right: 0.35rem !important; padding-top: 0.05rem !important; padding-bottom: 0.35rem !important;
            max-width: 100% !important;
        }}
        .main .block-container {{ padding: 0.1rem 0.35rem 0.5rem 0.35rem !important; }}
        .main div[data-testid="stImage"] img, .main .stImage img {{
            border: none !important;
            outline: none !important;
            box-shadow: 0 12px 40px rgba(8, 20, 45, 0.85),
                        0 6px 20px rgba(10, 25, 55, 0.75),
                        0 3px 12px rgba(5, 15, 40, 0.8),
                        0 0 0 1px rgba(15, 35, 70, 0.3) !important;
        }}
        .main div[data-testid="stImage"], .main [data-testid="stImage"] {{
            margin: 0 !important;
            padding: 0 !important;
            margin-top: 0 !important;
            border: none !important;
            outline: none !important;
            box-shadow: none !important;
        }}
        .main hr {{ display: none !important; }}
        .main .stButton > button {{
            padding: 1.25rem 1.5rem !important;
            min-height: 100px !important;
            font-size: 1.15rem !important;
            font-weight: bold !important;
            color: #ffffff !important;
            text-shadow: 0 1px 2px rgba(0,0,0,0.3) !important;
            border: 1px solid rgba(100,160,220,0.5) !important;
            border-radius: 12px !important;
            background: linear-gradient(180deg, #6ba3e0 0%, #3d7ab8 30%, #2568a0 60%, #2d6a9f 100%) !important;
            box-shadow: inset 0 2px 4px rgba(255,255,255,0.3),
                        inset 0 -3px 6px rgba(0,0,0,0.3),
                        0 4px 12px rgba(0,0,0,0.4) !important;
        }}
        .main .stButton > button:hover {{
            background: linear-gradient(180deg, #7eb3f0 0%, #4a8ac8 30%, #2d78b0 60%, #3a7ab8 100%) !important;
            transform: translateY(-3px) !important;
            box-shadow: inset 0 2px 4px rgba(255,255,255,0.4),
                        inset 0 -2px 4px rgba(0,0,0,0.2),
                        0 6px 20px rgba(0,0,0,0.5),
                        0 0 16px rgba(100,160,220,0.35) !important;
        }}
        .main .stButton > button:active {{
            transform: translateY(1px) !important;
            box-shadow: inset 0 3px 6px rgba(0,0,0,0.35), 0 2px 6px rgba(0,0,0,0.4) !important;
        }}
        @media (max-width: 768px) {{
            .main .block-container {{ padding: 0.05rem 0.25rem 0.35rem 0.25rem !important; }}
            .main div[data-testid="stImage"] img, .main .stImage img {{
                max-height: 42vh !important;
                max-width: 100% !important;
                width: auto !important;
                object-fit: contain !important;
                margin-bottom: 0.25rem !important;
            }}
            .main div[data-testid="stImage"], .main [data-testid="stImage"] {{ margin: 0 0 0.15rem 0 !important; }}
            .main .stButton > button {{
                min-height: 48px !important;
                padding: 0.6rem 0.8rem !important;
                font-size: 1rem !important;
            }}
            .main [data-testid="column"] {{ min-width: 0 !important; padding: 0.15rem !important; }}
        }}
        </style>
        """, unsafe_allow_html=True)
        # Debajo de la imagen: tres botones que al clicar muestran Misión, Visión u Objetivo Supremo
        cartel_abierto = st.session_state.get("cartel_abierto_inicio")
        c1, c2, c3 = st.columns(3)
        with c1:
            if st.button(f"**{t['mision']}** — {t['ver_mas']}", key="btn_cartel_mision", width="stretch"):
                st.session_state["cartel_abierto_inicio"] = "mision" if cartel_abierto != "mision" else None
                st.rerun()
        with c2:
            if st.button(f"**{t['vision']}** — {t['ver_mas']}", key="btn_cartel_vision", width="stretch"):
                st.session_state["cartel_abierto_inicio"] = "vision" if cartel_abierto != "vision" else None
                st.rerun()
        with c3:
            if st.button(f"**{t['objetivo_supremo']}** — {t['ver_mas']}", key="btn_cartel_objetivo", width="stretch"):
                st.session_state["cartel_abierto_inicio"] = "objetivo" if cartel_abierto != "objetivo" else None
                st.rerun()
        # Panel desplegable: texto de Misión, Visión u Objetivo
        if cartel_abierto in ("mision", "vision", "objetivo"):
            titulo_cartel = t["mision"] if cartel_abierto == "mision" else (t["vision"] if cartel_abierto == "vision" else t["objetivo_supremo"])
            texto_cartel = t["mision_texto"] if cartel_abierto == "mision" else (t["vision_texto"] if cartel_abierto == "vision" else t["objetivo_texto"])
            st.markdown("---")
            st.markdown(f"### {titulo_cartel}")
            st.info(texto_cartel)
            if st.button(t["cerrar"], key="btn_cerrar_cartel"):
                st.session_state["cartel_abierto_inicio"] = None
                st.rerun()
            st.markdown("---")
        # Accesos rápidos: cada oficina en su columna (Arqueo, Tesorería, Contabilidad, Presupuesto, Eventos)
        ui_cfg_inicio = cargar_ui_config() or {}
        ocultar_inicio = ui_cfg_inicio.get("ocultar_botones") or {}

        col_arqueo, col_tesoreria, col_contab, col_presup, col_eventos = st.columns(5)
        with col_arqueo:
            if not ocultar_inicio.get("arqueo_caja") and tiene_permiso(usuario_actual, "ver_arqueo_caja") and st.button(f"📋 {t['arqueo_caja']}", key="btn_ir_arqueo", width="stretch"):
                st.session_state["pagina"] = "arqueo_caja"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        with col_tesoreria:
            if not ocultar_inicio.get("tesoreria") and tiene_permiso(usuario_actual, "ver_tesoreria") and st.button(f"📒 {t['tesoreria']}", key="btn_ir_tesoreria", width="stretch"):
                st.session_state["pagina"] = "tesoreria"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        with col_contab:
            if not ocultar_inicio.get("contabilidad") and tiene_permiso(usuario_actual, "ver_contabilidad") and st.button(f"📊 {t['contabilidad']}", key="btn_ir_contabilidad", width="stretch"):
                st.session_state["pagina"] = "contabilidad"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        with col_presup:
            if not ocultar_inicio.get("presupuesto_metas") and tiene_permiso(usuario_actual, "ver_presupuesto_metas") and st.button(f"🎯 {t['presupuesto_metas']}", key="btn_ir_presupuesto", width="stretch"):
                st.session_state["pagina"] = "presupuesto_metas"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        with col_eventos:
            if not ocultar_inicio.get("eventos") and tiene_permiso(usuario_actual, "ver_eventos_inversiones") and st.button(f"📌 {t.get('eventos_inversiones', 'Eventos / Inversiones')}", key="btn_ir_eventos", width="stretch"):
                st.session_state["pagina"] = "eventos"
                st.session_state["sidebar_state"] = "collapsed"
                st.session_state["sidebar_collapse_requested"] = True
                st.rerun()
        with st.expander(f"❓ {t['primera_vez']}", expanded=False):
            st.caption(t["ayuda_rapida"])
        with st.expander(f"📲 {t['compartir_app']}", expanded=False):
            st.caption(t["compartir_app_instrucciones"])
        return

    # Página de edición de diseño (solo clave maestra)
    if st.session_state["pagina"] == "ui_config":
        _render_ui_config_page()
        return

    # ----- OFICINAS: ARQUEO, TESORERÍA, CONTABILIDAD, PRESUPUESTO -----
    pagina_act = st.session_state.get("pagina", "inicio")
    if pagina_act == "ministerio_finanzas":
        st.session_state["pagina"] = "contabilidad"
        st.rerun()
    if pagina_act not in ("arqueo_caja", "tesoreria", "contabilidad", "presupuesto_metas", "eventos"):
        ui_cfg_bienv = cargar_ui_config() or {}
        textos_bienv = ui_cfg_bienv.get("textos") or {}
        mensaje_bienvenida = textos_bienv.get("bienvenida_texto") or t["bienvenida_texto"]
        st.info(mensaje_bienvenida)
        return

    with st.spinner(t.get("cargando", "Cargando datos...")):
        df = cargar_db()

    # ---------- ARQUEO DE CAJA (Cierre Diario) ----------
    if pagina_act == "arqueo_caja":
        st.markdown(f"## 📋 {t['arqueo_caja']}")
        st.caption(t["arqueo_caja_sub"])
        st.markdown("""<style>
        @media (max-width: 768px) {
        .main input[type="number"],
        .main [data-testid="stNumberInput"] input {
          font-size: 1.9rem !important;
          font-weight: bold !important;
          padding: 0.9rem 0.6rem !important;
          text-align: center !important;
          min-height: 2.8rem !important;
        }
        .main [data-testid="stNumberInput"] label { font-size: 1.1rem !important; }
        }
        </style>""", unsafe_allow_html=True)
        # Resumen del día
        hoy_str = datetime.now().strftime("%Y-%m-%d")
        df_hoy = df[df["fecha"].astype(str).str[:10] == hoy_str] if not df.empty and "fecha" in df.columns else pd.DataFrame()
        ing_hoy = float(df_hoy["ingreso"].sum()) if not df_hoy.empty else 0.0
        gas_hoy = float(df_hoy["gastos"].sum()) if not df_hoy.empty else 0.0
        with st.expander(f"📊 {t['resumen_dia']}", expanded=not df_hoy.empty):
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric(t["ingresos_mes"] + " " + hoy_str, f"${ing_hoy:,.2f}")
            with c2:
                st.metric(t["gastos_mes"], f"${gas_hoy:,.2f}")
            with c3:
                st.metric(t["saldo_actual"], f"${ing_hoy - gas_hoy:,.2f}")
        if tiene_permiso(usuario_actual, "ver_ingresar_bendicion"):
            if st.session_state.get("limpiar_arqueo"):
                for key in ["b100", "b50", "b20", "b10", "b5", "m2", "m1", "m025", "m010", "m005"]:
                    st.session_state[key] = 0
                st.session_state["limpiar_arqueo"] = False
            with st.expander(f"💰 {t['ingresar_bendicion']}", expanded=True):
                def _aplica_mayusculas_contado():
                    v = st.session_state.get("contado_por_arqueo", "")
                    if v and v != v.upper():
                        st.session_state["contado_por_arqueo"] = v.upper()

                def _aplica_mayusculas_verificado():
                    v = st.session_state.get("verificado_por_arqueo", "")
                    if v and v != v.upper():
                        st.session_state["verificado_por_arqueo"] = v.upper()

                st.markdown("""<style>
        .main [data-testid="stTextInput"]:nth-of-type(1) input,
        .main [data-testid="stTextInput"]:nth-of-type(2) input { text-transform: uppercase !important; }
        /* Refrescar: metálico azul igual que todos los botones */
        .main [data-testid="stExpander"] .stButton > button {
            background: linear-gradient(180deg, #7eb3f0 0%, #99d0ff 18%, #6ba3e0 35%, #3d7ab8 55%, #2568a0 75%, #2d6a9f 100%) !important;
            color: #ffffff !important;
            box-shadow: inset 0 2px 6px rgba(255,255,255,0.4), inset 0 -3px 6px rgba(0,0,0,0.3), 0 4px 12px rgba(0,0,0,0.4) !important;
        }
        .main [data-testid="stExpander"] .stButton > button:hover {
            background: linear-gradient(180deg, #8ec5ff 0%, #a8dcff 20%, #7eb3f0 40%, #4a8ac8 60%, #2d78b0 80%, #3a7ab8 100%) !important;
            transform: translateY(-3px) !important;
        }
        .main [data-testid="stExpander"] .stButton > button:active {
            transform: translateY(1px) !important;
        }
        </style>""", unsafe_allow_html=True)
                meta_arq = cargar_arqueo_meta()
                nombres_contado, nombres_verificado = _nombres_arqueo_desde_meta(meta_arq)
                if "contado_por_arqueo" not in st.session_state:
                    st.session_state["contado_por_arqueo"] = usuario_actual
                if "verificado_por_arqueo" not in st.session_state:
                    st.session_state["verificado_por_arqueo"] = ""

                contado_por_raw = st.text_input(t["contado_por"] + " *", key="contado_por_arqueo", max_chars=80, placeholder="Escriba el nombre (se guarda en MAYÚSCULAS)", help=t["contado_por_ayuda"], on_change=_aplica_mayusculas_contado)
                if nombres_contado:
                    st.caption(t["arqueo_sugerencias"] + ": ")
                    cols_c = st.columns(min(6, len(nombres_contado) + 1))[:6]
                    for i, nom in enumerate(nombres_contado[:5]):
                        with cols_c[i]:
                            if st.button(nom, key=f"contado_sug_{i}", width="stretch"):
                                st.session_state["contado_por_arqueo"] = nom
                                st.rerun()
                contado_por = (contado_por_raw or "").strip().upper()
                if contado_por and nombres_contado:
                    contado_por = _normalizar_y_coincidir(contado_por, nombres_contado) or contado_por
                    if len((contado_por_raw or "").strip()) == 1 and len(contado_por) > 1:
                        st.session_state["contado_por_arqueo"] = contado_por
                        st.rerun()
                ok_contado, msg_contado = _validar_nombre_arqueo(contado_por)
                if contado_por and not ok_contado:
                    st.warning(t.get(msg_contado, msg_contado))

                verificado_habilitado = bool(contado_por)
                verificado_por_raw = st.text_input(t["verificado_por"] + " *", key="verificado_por_arqueo", max_chars=80, placeholder="Escriba el nombre (se guarda en MAYÚSCULAS)", disabled=not verificado_habilitado, help=t["verificado_por_ayuda"], on_change=_aplica_mayusculas_verificado)
                if verificado_habilitado and nombres_verificado:
                    st.caption(t["arqueo_sugerencias"] + ": ")
                    cols_v = st.columns(min(6, len(nombres_verificado) + 1))[:6]
                    for i, nom in enumerate(nombres_verificado[:5]):
                        with cols_v[i]:
                            if st.button(nom, key=f"verificado_sug_{i}", width="stretch"):
                                st.session_state["verificado_por_arqueo"] = nom
                                st.rerun()
                verificado_por = (verificado_por_raw or "").strip().upper()
                if verificado_por and nombres_verificado:
                    verificado_por = _normalizar_y_coincidir(verificado_por, nombres_verificado) or verificado_por
                    if len((verificado_por_raw or "").strip()) == 1 and len(verificado_por) > 1:
                        st.session_state["verificado_por_arqueo"] = verificado_por
                        st.rerun()
                ok_verif, msg_verif = _validar_nombre_arqueo(verificado_por)
                if verificado_por and not ok_verif:
                    st.warning(t.get(msg_verif, msg_verif))

                ambos_ok = bool(contado_por and verificado_por and ok_contado and ok_verif)
                if not verificado_habilitado:
                    st.info(t["arqueo_llenar_ambos"] if lang == "ES" else t["arqueo_llenar_ambos"])
                elif not ambos_ok and (contado_por or verificado_por):
                    st.warning(t["arqueo_llenar_ambos"])
                st.markdown("---")

                campos_habilitados = ambos_ok
                fondo_caja_val = st.number_input(t["fondo_caja"], min_value=0.0, value=0.0, step=10.0, key="fondo_caja_arqueo", disabled=not campos_habilitados)
                tipo_ingreso_opciones = TIPOS_INGRESO_ES if lang == "ES" else TIPOS_INGRESO_EN
                tipo_ingreso_sel = st.selectbox(t["tipo_ingreso"], tipo_ingreso_opciones, key="tipo_ingreso_bendicion", disabled=not campos_habilitados)
                medios_opciones = MEDIOS_PAGO_ES if lang == "ES" else MEDIOS_PAGO_EN
                medio_efectivo = MEDIO_EFECTIVO_ES if lang == "ES" else MEDIO_EFECTIVO_EN
                medio_cheques_es = "Cheques"
                medio_cheques_en = "Checks"
                medio_pago_sel = st.selectbox(t["medio_pago"], medios_opciones, key="medio_pago_bendicion", disabled=not campos_habilitados)
                es_efectivo = (medio_pago_sel == medio_efectivo)
                es_cheques = (medio_pago_sel == medio_cheques_es or medio_pago_sel == medio_cheques_en)

                if es_cheques:
                    cheques_cant = st.number_input(t["cheques_cantidad"], min_value=0, value=0, step=1, key="cheques_cant_arqueo", disabled=not campos_habilitados)
                    cheques_tot = st.number_input(t["cheques_total"], min_value=0.0, value=0.0, step=10.0, key="cheques_tot_arqueo", disabled=not campos_habilitados)
                    st.caption("Marque los cheques como «solo para depósito» antes de guardarlos." if lang == "ES" else "Stamp checks 'for deposit only' before storing.")
                elif es_efectivo:
                    ministerio = st.selectbox(t["ministerio"], ministerios, key="min_bendicion", disabled=not campos_habilitados)
                    st.markdown(f"<p class='big-label'>{t['billetes']}</p>", unsafe_allow_html=True)
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        b100 = st.number_input("$100", min_value=0, value=0, step=1, key="b100", disabled=not campos_habilitados)
                    with col2:
                        b50 = st.number_input("$50", min_value=0, value=0, step=1, key="b50", disabled=not campos_habilitados)
                    with col3:
                        b20 = st.number_input("$20", min_value=0, value=0, step=1, key="b20", disabled=not campos_habilitados)
                    with col4:
                        b10 = st.number_input("$10", min_value=0, value=0, step=1, key="b10", disabled=not campos_habilitados)
                    with col5:
                        b5 = st.number_input("$5", min_value=0, value=0, step=1, key="b5", disabled=not campos_habilitados)
                    st.markdown(f"<p class='big-label'>{t['monedas']}</p>", unsafe_allow_html=True)
                    c1, c2, c3, c4, c5 = st.columns(5)
                    with c1:
                        m2 = st.number_input("$2", min_value=0, value=0, step=1, key="m2", disabled=not campos_habilitados)
                    with c2:
                        m1 = st.number_input("$1", min_value=0, value=0, step=1, key="m1", disabled=not campos_habilitados)
                    with c3:
                        m025 = st.number_input("$0.25", min_value=0, value=0, step=1, key="m025", disabled=not campos_habilitados)
                    with c4:
                        m010 = st.number_input("$0.10", min_value=0, value=0, step=1, key="m010", disabled=not campos_habilitados)
                    with c5:
                        m005 = st.number_input("$0.05", min_value=0, value=0, step=1, key="m005", disabled=not campos_habilitados)
                    total_arqueo = (
                        b100 * 100 + b50 * 50 + b20 * 20 + b10 * 10 + b5 * 5 +
                        m2 * 2 + m1 * 1 + m025 * 0.25 + m010 * 0.10 + m005 * 0.05
                    )
                    if "_last_total_arqueo" not in st.session_state or abs(st.session_state.get("_last_total_arqueo", 0) - total_arqueo) > 0.01:
                        st.session_state["total_suelto_arqueo"] = round(total_arqueo, 2)
                        st.session_state["_last_total_arqueo"] = total_arqueo
                    if "total_suelto_arqueo" not in st.session_state:
                        st.session_state["total_suelto_arqueo"] = round(total_arqueo, 2)
                    st.caption(t.get("arqueo_teclado_movil", ""))
                    st.markdown(f"**{t.get('arqueo_total_calculado', 'Total billetes + monedas')}:** ${total_arqueo:.2f}")
                    sobres_cant = st.number_input(t["sobres_cantidad"], min_value=0, value=0, step=1, key="sobres_cant_arqueo", disabled=not campos_habilitados)
                    sobres_tot = st.number_input(t["sobres_total"], min_value=0.0, value=0.0, step=10.0, key="sobres_tot_arqueo", disabled=not campos_habilitados)
                    total_suelto = st.number_input(t["total_suelto"], min_value=0.0, step=10.0, key="total_suelto_arqueo", disabled=not campos_habilitados)
                    st.caption("Verifique que el contenido de cada sobre coincida con lo escrito." if lang == "ES" else "Verify envelope contents match written amounts.")
                    st.markdown(f"<div class='total-box'>{t['total']}: ${total_arqueo:.2f}</div>", unsafe_allow_html=True)
                elif not es_cheques:
                    monto_no_efectivo = st.number_input(t["monto_pos_tarjeta"], min_value=0.0, value=0.0, step=10.0, key="monto_ingreso_pos", disabled=not campos_habilitados)
                    ref_opcional = st.text_input(t["referencia_opcional"], key="ref_pos", max_chars=20, placeholder=t.get("referencia_placeholder", ""), disabled=not campos_habilitados)

                _arqueo_keys = (
                    "contado_por_arqueo", "verificado_por_arqueo", "tipo_ingreso_bendicion", "medio_pago_bendicion",
                    "min_bendicion", "fondo_caja_arqueo", "cheques_cant_arqueo", "cheques_tot_arqueo",
                    "b100", "b50", "b20", "b10", "b5", "m2", "m1", "m025", "m010", "m005",
                    "sobres_cant_arqueo", "sobres_tot_arqueo", "total_suelto_arqueo", "monto_ingreso_pos", "ref_pos",
                    "_last_total_arqueo", "limpiar_arqueo"
                )
                col_reg, col_sp, col_ref = st.columns([1, 2, 1])
                with col_reg:
                    if st.button(t["registrar"], key="btn_bendicion", disabled=not ambos_ok):
                        if es_cheques:
                            monto_ingreso = round(float(cheques_tot), 2) if cheques_tot else 0.0
                            if monto_ingreso <= 0:
                                st.warning(t["arqueo_cero"])
                            else:
                                rid = generar_id_arqueo()
                                fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                detalle = f"Arqueo - Cheques ({cheques_cant} cheques)"
                                if lang == "ES":
                                    tipo_guardar_ing = tipo_ingreso_sel
                                else:
                                    try:
                                        idx_ing = TIPOS_INGRESO_EN.index(tipo_ingreso_sel)
                                        tipo_guardar_ing = TIPOS_INGRESO_ES[idx_ing]
                                    except (ValueError, IndexError):
                                        tipo_guardar_ing = DEFAULT_TIPO_INGRESO
                                nueva = pd.DataFrame([{
                                    "id_registro": rid,
                                    "fecha": fecha_ahora,
                                    "detalle": detalle,
                                    "tipo_gasto": tipo_guardar_ing,
                                    "ingreso": monto_ingreso,
                                    "gastos": 0,
                                    "total_ingresos": 0,
                                    "total_gastos": 0,
                                    "saldo_actual": 0
                                }])
                                df = pd.concat([df, nueva], ignore_index=True)
                                if guardar_db(df, t):
                                    meta = cargar_arqueo_meta()
                                    meta[rid] = {
                                        "contado_por": (contado_por or "").strip() or usuario_actual,
                                        "verificado_por": (verificado_por or "").strip(),
                                        "ip_dispositivo": _get_client_ip(),
                                        "desglose": {},
                                        "total_cheques": monto_ingreso,
                                        "cheques_cant": cheques_cant,
                                        "fecha": fecha_ahora,
                                    }
                                    guardar_arqueo_meta(meta)
                                    audit_log(usuario_actual, "ingreso_registrado", f"{rid} {detalle} ${monto_ingreso:.2f}")
                                    st.session_state["limpiar_arqueo"] = True
                                    st.success(t["bendicion_registrada"])
                                    st.rerun()
                        elif es_efectivo:
                            monto_ingreso = round(float(total_suelto or 0) + float(sobres_tot or 0), 2)
                            if monto_ingreso <= 0:
                                st.warning(t["arqueo_cero"])
                            else:
                                rid = generar_id_arqueo()
                                fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                detalle = f"Arqueo - {ministerio}"
                                if lang == "ES":
                                    tipo_guardar_ing = tipo_ingreso_sel
                                else:
                                    try:
                                        idx_ing = TIPOS_INGRESO_EN.index(tipo_ingreso_sel)
                                        tipo_guardar_ing = TIPOS_INGRESO_ES[idx_ing]
                                    except (ValueError, IndexError):
                                        tipo_guardar_ing = DEFAULT_TIPO_INGRESO
                                nueva = pd.DataFrame([{
                                    "id_registro": rid,
                                    "fecha": fecha_ahora,
                                    "detalle": detalle,
                                    "tipo_gasto": tipo_guardar_ing,
                                    "ingreso": monto_ingreso,
                                    "gastos": 0,
                                    "total_ingresos": 0,
                                    "total_gastos": 0,
                                    "saldo_actual": 0
                                }])
                                df = pd.concat([df, nueva], ignore_index=True)
                                if guardar_db(df, t):
                                    meta = cargar_arqueo_meta()
                                    meta[rid] = {
                                        "contado_por": (contado_por or "").strip() or usuario_actual,
                                        "verificado_por": (verificado_por or "").strip(),
                                        "ip_dispositivo": _get_client_ip(),
                                        "desglose": {"b100": b100, "b50": b50, "b20": b20, "b10": b10, "b5": b5,
                                                     "m2": m2, "m1": m1, "m025": m025, "m010": m010, "m005": m005},
                                        "total_efectivo": float(total_arqueo),
                                        "total": monto_ingreso,
                                        "sobres_cant": sobres_cant,
                                        "sobres_tot": float(sobres_tot or 0),
                                        "total_suelto": float(total_suelto or 0),
                                        "fondo_caja": float(fondo_caja_val or 0),
                                        "fecha": fecha_ahora,
                                    }
                                    guardar_arqueo_meta(meta)
                                    audit_log(usuario_actual, "ingreso_registrado", f"{rid} {detalle} ${monto_ingreso:.2f}")
                                    st.session_state["limpiar_arqueo"] = True
                                    st.success(t["bendicion_registrada"])
                                    st.rerun()
                        else:
                            monto_ingreso = round(float(monto_no_efectivo), 2) if monto_no_efectivo else 0.0
                            if monto_ingreso <= 0:
                                st.warning(t["arqueo_cero"])
                            else:
                                rid = generar_id_arqueo()
                                fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                if lang == "ES":
                                    tipo_guardar_ing = tipo_ingreso_sel
                                else:
                                    try:
                                        idx_ing = TIPOS_INGRESO_EN.index(tipo_ingreso_sel)
                                        tipo_guardar_ing = TIPOS_INGRESO_ES[idx_ing]
                                    except (ValueError, IndexError):
                                        tipo_guardar_ing = DEFAULT_TIPO_INGRESO
                                if medio_pago_sel == (MEDIOS_PAGO_ES[1] if lang == "ES" else MEDIOS_PAGO_EN[1]):
                                    sufijo = "POS"
                                elif medio_pago_sel == (MEDIOS_PAGO_ES[2] if lang == "ES" else MEDIOS_PAGO_EN[2]):
                                    sufijo = "Tarjeta débito" if lang == "ES" else "Debit card"
                                else:
                                    sufijo = "Transferencia" if lang == "ES" else "Transfer"
                                ref = (ref_opcional or "").strip()[:20]
                                if ref:
                                    detalle = f"{tipo_guardar_ing} ({sufijo} ****{ref})"
                                else:
                                    detalle = f"{tipo_guardar_ing} ({sufijo})"
                                nueva = pd.DataFrame([{
                                    "id_registro": rid,
                                    "fecha": fecha_ahora,
                                    "detalle": detalle,
                                    "tipo_gasto": tipo_guardar_ing,
                                    "ingreso": monto_ingreso,
                                    "gastos": 0,
                                    "total_ingresos": 0,
                                    "total_gastos": 0,
                                    "saldo_actual": 0
                                }])
                                df = pd.concat([df, nueva], ignore_index=True)
                                if guardar_db(df, t):
                                    meta = cargar_arqueo_meta()
                                    meta[rid] = {
                                        "contado_por": (contado_por or "").strip() or usuario_actual,
                                        "verificado_por": (verificado_por or "").strip(),
                                        "ip_dispositivo": _get_client_ip(),
                                        "total_pos": monto_ingreso,
                                        "fecha": fecha_ahora,
                                    }
                                    guardar_arqueo_meta(meta)
                                    audit_log(usuario_actual, "ingreso_registrado", f"{rid} {detalle} ${monto_ingreso:.2f}")
                                    st.session_state["limpiar_arqueo"] = True
                                    st.success(t["bendicion_registrada"])
                                    st.rerun()
                with col_ref:
                    if st.button(t.get("arqueo_refrescar", "Refrescar"), key="btn_refrescar_arqueo", help=t.get("arqueo_refrescar_ayuda", "")):
                        for k in _arqueo_keys:
                            if k in st.session_state:
                                del st.session_state[k]
                        st.rerun()
        with st.expander(f"📋 {t['conciliar']}", expanded=False):
            st.caption(t["conciliar_ayuda"])
            fecha_conciliar = st.date_input(t["fecha_conciliar"], value=datetime.now().date(), key="fecha_conciliar_arqueo")
            fecha_str = fecha_conciliar.strftime("%Y-%m-%d") if fecha_conciliar else datetime.now().strftime("%Y-%m-%d")
            df_dash = df.copy()
            try:
                df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str).str[:10], errors="coerce")
                df_dia = df_dash[df_dash["fecha"].astype(str).str[:10] == fecha_str]
            except Exception:
                df_dia = pd.DataFrame()
            ing_dia = float(df_dia["ingreso"].sum()) if not df_dia.empty else 0.0
            st.metric(t["ingresos_mes"] + f" ({fecha_str})", f"${ing_dia:,.2f}")
            lo_contado = st.number_input(t["lo_contado_caja"], min_value=0.0, value=0.0, step=10.0, key="conciliar_contado")
            if lo_contado > 0:
                diff = lo_contado - ing_dia
                if abs(diff) < 0.02:
                    st.success(t["coincide_registrado"])
                else:
                    st.caption(f"Diferencia: ${diff:+,.2f}")
            meta_conc = cargar_arqueo_meta()
            fechas_cerradas = meta_conc.get("_fechas_cerradas", []) or []
            if fecha_str in fechas_cerradas:
                st.success(f"✓ {t['arqueo_cerrado']} ({fecha_str})")
            elif st.button(t["cerrar_arqueo"], key="btn_cerrar_arqueo"):
                fechas_cerradas = list(set(fechas_cerradas + [fecha_str]))
                meta_conc["_fechas_cerradas"] = fechas_cerradas
                guardar_arqueo_meta(meta_conc)
                audit_log(usuario_actual, "arqueo_cerrado", fecha_str)
                st.success(t["arqueo_cerrado"])
                st.rerun()
        with st.expander(f"📤 {t['descargar_hoja_arqueo']}", expanded=False):
            st.caption(t["descargar_hoja_arqueo_ayuda"])
            arqueos_ac = df[df["id_registro"].astype(str).str.startswith("AC-")] if not df.empty else pd.DataFrame()
            if arqueos_ac.empty:
                st.info(t["sin_movimientos"])
            else:
                arqueos_lista = arqueos_ac.iloc[-20:].iloc[::-1]
                opciones_rid = [str(r.get("id_registro", "")) for _, r in arqueos_lista.iterrows()]
                rid_a_label = {str(r.get("id_registro", "")): f"{str(r.get('fecha',''))[:16]} — ${float(r.get('ingreso',0)):,.2f}" for _, r in arqueos_lista.iterrows()}
                if not opciones_rid:
                    st.info(t["sin_movimientos"])
                else:
                    rid_sel = st.selectbox(
                        t["seleccionar_arqueo"],
                        options=opciones_rid,
                        format_func=lambda x: rid_a_label.get(x, str(x)),
                        key="sel_arqueo_descarga"
                    )
                    df_filt = arqueos_ac[arqueos_ac["id_registro"].astype(str) == str(rid_sel or "")] if rid_sel else pd.DataFrame()
                    if df_filt.empty:
                        fila_sel = arqueos_lista.iloc[0]
                    else:
                        fila_sel = df_filt.iloc[0]
                    rid_ult = str(fila_sel.get("id_registro", ""))
                    meta = cargar_arqueo_meta()
                    datos = meta.get(rid_ult, {})
                    datos["fecha"] = datos.get("fecha") or str(fila_sel.get("fecha", ""))
                    datos["total"] = float(fila_sel.get("ingreso", 0))
                    datos["total_efectivo"] = datos.get("total_efectivo") or datos.get("total", 0)
                    col_pdf_a, col_xlsx_a = st.columns(2)
                    with col_pdf_a:
                        try:
                            pdf_arq = generar_pdf_hoja_arqueo(datos, lang)
                            if pdf_arq:
                                st.download_button(
                                    label=t["hoja_arqueo_pdf"],
                                    data=pdf_arq,
                                    file_name=f"Hoja_Arqueo_{rid_ult.replace(':', '-')[:30]}.pdf",
                                    mime="application/pdf",
                                    key="download_hoja_arqueo_pdf"
                                )
                        except Exception:
                            st.caption("PDF no disponible" if lang == "ES" else "PDF not available")
                    with col_xlsx_a:
                        try:
                            xlsx_arq = generar_excel_hoja_arqueo(datos, lang)
                            if xlsx_arq:
                                st.download_button(
                                    label=t["hoja_arqueo_excel"],
                                    data=xlsx_arq,
                                    file_name=f"Hoja_Arqueo_{rid_ult.replace(':', '-')[:30]}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key="download_hoja_arqueo_excel"
                                )
                        except Exception:
                            st.caption("Excel no disponible" if lang == "ES" else "Excel not available")
        return

    # ---------- TESORERÍA (Libro de Registros) ----------
    if pagina_act == "tesoreria":
        st.markdown(f"## 📒 {t['tesoreria']}")
        st.caption(t["tesoreria_sub"])
        if tiene_permiso(usuario_actual, "ver_registrar_gasto"):
            MAX_CARACTERES_DESCRIPCION_T = 1000
            _gasto_keys = ["desc_g", "monto_g", "aprobado_por_g", "tipo_gasto_sel", "suministro_sel", "gasto_frecuente_sel", "modo_escaneo_rapido"]
            if st.session_state.get("limpiar_gasto"):
                for key in ["desc_g", "aprobado_por_g"]:
                    if key in st.session_state:
                        st.session_state[key] = ""
                if "monto_g" in st.session_state:
                    st.session_state["monto_g"] = 0.0
                for k in ("monto_sugerido", "descripcion_sugerida", "factura_detectada_actual", "ocr_text_actual", "hash_foto_actual"):
                    st.session_state.pop(k, None)
                tipo_opciones = TIPOS_GASTO_ES if st.session_state.get("idioma", "ES") == "ES" else TIPOS_GASTO_EN
                st.session_state["tipo_gasto_sel"] = tipo_opciones[0]
                st.session_state["suministro_sel"] = ""
                st.session_state["gasto_frecuente_sel"] = ""
                try:
                    for k in list(st.session_state.keys()):
                        if k == "foto_gasto" or (isinstance(k, str) and k.startswith("foto_gasto")):
                            del st.session_state[k]
                except Exception:
                    pass
                st.session_state["limpiar_gasto"] = False
            modo_rapido = st.checkbox(t.get("modo_escaneo_rapido", "Modo escaneo rápido"), key="modo_escaneo_rapido", value=st.session_state.get("modo_escaneo_rapido", False))
            with st.expander(f"📤 {t['registrar_gasto']}", expanded=True):
                tipo_opciones = TIPOS_GASTO_ES if lang == "ES" else TIPOS_GASTO_EN
                tipo_seleccionado = st.selectbox(t["tipo_gasto"], options=tipo_opciones, key="tipo_gasto_sel")
                es_recurrentes = tipo_seleccionado in ("Recurrentes", "Recurring")
                es_operativo = tipo_seleccionado in ("Operativo", "Operational")
                suministro_sel = ""
                gasto_frec_sel = ""
                if es_recurrentes:
                    suministros_lista = cargar_suministros()
                    suministro_sel = st.selectbox(t["suministros"], options=[""] + suministros_lista, key="suministro_sel")
                    if suministro_sel and not st.session_state.get("descripcion_sugerida"):
                        if "desc_g" not in st.session_state or not st.session_state.get("desc_g"):
                            st.session_state["desc_g"] = suministro_sel
                elif es_operativo:
                    gastos_frec = _gastos_frecuentes_desde_df(df)
                    opciones_gf = [""] + gastos_frec[:15]
                    gasto_frec_sel = st.selectbox(t["gastos_frecuentes"], options=opciones_gf, key="gasto_frecuente_sel")
                    if gasto_frec_sel and not st.session_state.get("descripcion_sugerida"):
                        if "desc_g" not in st.session_state or not st.session_state.get("desc_g"):
                            st.session_state["desc_g"] = gasto_frec_sel
                with st.form("form_gasto", clear_on_submit=True):
                    monto_sugerido = st.session_state.get("monto_sugerido", "")
                    try:
                        monto_val_inicial = float(str(monto_sugerido).replace(",", ".")) if monto_sugerido else 0.0
                    except (ValueError, TypeError):
                        monto_val_inicial = 0.0
                    st.number_input(t["monto"], min_value=0.0, value=monto_val_inicial, step=0.01, format="%.2f", key="monto_g", help=t.get("gasto_teclado_movil", ""))
                    col_btn_reg, col_btn_sp, col_btn_ref = st.columns([1, 2, 1])
                    with col_btn_reg:
                        enviado = st.form_submit_button(t["registrar"])
                col_btn_reg, col_btn_sp, col_btn_ref = st.columns([1, 2, 1])
                with col_btn_ref:
                    if st.button(t.get("gasto_refrescar", "Refrescar"), key="btn_refrescar_gasto", help=t.get("gasto_refrescar_ayuda", ""), width="stretch"):
                        st.session_state["limpiar_gasto"] = True
                        st.rerun()
                if not modo_rapido:
                    desc_sugerida = st.session_state.get("descripcion_sugerida", "")
                    desc_base = desc_sugerida or (suministro_sel if es_recurrentes else (gasto_frec_sel if es_operativo else ""))
                    desc_gasto = st.text_input(t["descripcion"], value=st.session_state.get("desc_g", desc_base) or desc_base, max_chars=MAX_CARACTERES_DESCRIPCION_T, key="desc_g")
                    n_car = len(desc_gasto or "")
                    st.markdown(
                        f"<p style='font-size:0.75rem; color: {sidebar_txt_muted}; margin-top: -0.5rem;'>{n_car} / {MAX_CARACTERES_DESCRIPCION_T} {t['caracteres']}</p>",
                        unsafe_allow_html=True
                    )
                    nombres_aprobado = _nombres_aprobado_desde_df(df)
                    def _aplica_mayusculas_aprobado():
                        v = st.session_state.get("aprobado_por_g", "")
                        if v and v != v.upper():
                            st.session_state["aprobado_por_g"] = v.upper()
                    aprobado_por_raw = st.text_input(t["aprobado_por"] + f" {UMBRAL_GASTO_APROBACION:.0f})", key="aprobado_por_g", max_chars=100, placeholder="Nombre de quien aprueba", on_change=_aplica_mayusculas_aprobado)
                    if nombres_aprobado:
                        st.caption(t.get("gasto_sugerencias", "Sugerencias") + ": ")
                        cols_ap = st.columns(min(6, len(nombres_aprobado) + 1))[:6]
                        for i, nom in enumerate(nombres_aprobado[:5]):
                            with cols_ap[i]:
                                if st.button(nom, key=f"aprobado_sug_{i}", width="stretch"):
                                    st.session_state["aprobado_por_g"] = nom
                                    st.rerun()
                    aprobado_por_gasto = (aprobado_por_raw or "").strip().upper()
                    if aprobado_por_gasto and nombres_aprobado:
                        aprobado_por_gasto = _normalizar_y_coincidir(aprobado_por_gasto, nombres_aprobado) or aprobado_por_gasto
                    ok_aprob, msg_aprob = _validar_nombre_arqueo(aprobado_por_gasto)
                    if aprobado_por_gasto and not ok_aprob:
                        st.warning(t.get(msg_aprob, msg_aprob))
                else:
                    desc_gasto = st.session_state.get("desc_g", "") or "Gasto (modo rápido)"
                    aprobado_por_gasto = ""
                    ok_aprob = True
                st.markdown(f"<p class='big-label'>{t['tomar_foto']}</p>", unsafe_allow_html=True)
                st.caption(t.get("foto_multiples_ayuda", "Puede subir varias fotos (frente, reverso, anexos)."))
                foto_subidas = st.file_uploader(" ", type=["jpg", "jpeg", "png"], key="foto_gasto", accept_multiple_files=True)
                foto_subida = foto_subidas[0] if foto_subidas else None

                if foto_subida:
                    bytes_foto = foto_subida.getvalue()
                    hash_foto = _hash_imagen(bytes_foto)
                    if imagen_ya_subida(hash_foto):
                        st.warning(t["imagen_ya_subida_aviso"])
                    if st.session_state.get("hash_foto_actual") != hash_foto:
                        ocr_text = _ocr_imagen(bytes_foto)
                        datos = _extraer_datos_factura(ocr_text)
                        st.session_state["factura_detectada_actual"] = datos
                        st.session_state["ocr_text_actual"] = ocr_text
                        st.session_state["hash_foto_actual"] = hash_foto
                        st.session_state["monto_sugerido"] = f"{datos['total']:.2f}" if datos.get("total") is not None else ""
                        st.session_state["descripcion_sugerida"] = (datos.get("comercio") or "")[:MAX_CARACTERES_DESCRIPCION_T]
                        st.rerun()
                    datos_show = st.session_state.get("factura_detectada_actual") or {}
                    tiene_datos_ocr = datos_show and (datos_show.get("total") is not None or datos_show.get("comercio"))
                    with st.expander(f"📋 {t['factura_detectada']}", expanded=True):
                        if st.button(t.get("reintentar_ocr", "Reintentar OCR"), key="btn_reintentar_ocr"):
                            ocr_text = _ocr_imagen(bytes_foto)
                            datos = _extraer_datos_factura(ocr_text)
                            st.session_state["factura_detectada_actual"] = datos
                            st.session_state["ocr_text_actual"] = ocr_text
                            st.session_state["monto_sugerido"] = f"{datos['total']:.2f}" if datos.get("total") is not None else ""
                            st.session_state["descripcion_sugerida"] = (datos.get("comercio") or "")[:MAX_CARACTERES_DESCRIPCION_T]
                            st.rerun()
                        st.markdown(f"**{t.get('vista_previa_factura', 'Vista previa')}:**")
                        for idx_f, fup in enumerate(foto_subidas):
                            lbl = (t.get("foto_frente", "Frente"), t.get("foto_reverso", "Reverso"), t.get("foto_anexo", "Anexo"))[min(idx_f, 2)] if idx_f < 3 else f"#{idx_f + 1}"
                            st.caption(lbl if len(foto_subidas) > 1 else "")
                            st.image(fup.getvalue(), width="stretch")
                        if datos_show.get("total") is not None:
                            st.markdown(f"**{t['total_detectado']}:** ${datos_show['total']:.2f}")
                        if datos_show.get("impuesto") is not None:
                            st.markdown(f"**{t['impuesto_detectado']}:** ${datos_show['impuesto']:.2f}")
                        if datos_show.get("comercio"):
                            st.markdown(f"**{t['comercio_detectado']}:** {datos_show['comercio'][:100]}")
                        monto_manual = 0.0
                        try:
                            mg = st.session_state.get("monto_g")
                            if mg is not None:
                                monto_manual = round(float(mg), 2) if isinstance(mg, (int, float)) else round(float((str(mg) or "").strip().replace(",", ".") or 0), 2)
                        except (ValueError, TypeError):
                            pass
                        if datos_show.get("total") is not None and monto_manual > 0 and abs(monto_manual - datos_show["total"]) > 0.01:
                            st.warning(t.get("ocr_diferencia_aviso", "").format(manual=monto_manual, ocr=datos_show["total"]))
                        if not tiene_datos_ocr:
                            st.info(t.get("ocr_sin_datos", "No se detectó información. Use «Reintentar OCR» o ingrese los datos manualmente."))

                if enviado:
                    rid = generar_id_gasto()
                    monto_raw = st.session_state.get("monto_g")
                    try:
                        monto_final = round(float(monto_raw), 2) if monto_raw is not None else 0.0
                    except (ValueError, TypeError):
                        try:
                            monto_parse = (str(monto_raw or "") or "").strip().replace(",", ".")
                            monto_final = round(float(monto_parse) if monto_parse else 0.0, 2)
                        except (ValueError, TypeError):
                            monto_final = 0.0
                    if monto_final < 0:
                        monto_final = 0.0
                    monto_final = round(float(monto_final), 2)
                    if monto_final <= 0:
                        st.warning(t["gasto_cero"])
                    elif monto_final >= UMBRAL_GASTO_APROBACION and (not aprobado_por_gasto or not ok_aprob):
                        st.warning(t.get("gasto_aprobado_requerido", "Gastos ≥ $500 requieren nombre válido en Aprobado por.") if lang == "ES" else t.get("gasto_aprobado_requerido", "Expenses ≥ $500 require valid name in Approved by."))
                    else:
                        foto_ruta = ""
                        fotos_guardadas = []
                        if foto_subidas:
                            carpeta_fotos = "fotos_gastos"
                            os.makedirs(carpeta_fotos, exist_ok=True)
                            for idx, fup in enumerate(foto_subidas):
                                bytes_orig = fup.getvalue()
                                bytes_guardar = _comprimir_imagen(bytes_orig)
                                base_name = os.path.splitext(os.path.basename(getattr(fup, "name", "foto")))[0].replace("\\", "_").replace("/", "_")[:60]
                                nombre_foto = f"{rid}_{idx}_{base_name}.jpg" if idx > 0 else f"{rid}_{base_name}.jpg"
                                ruta = os.path.join(carpeta_fotos, nombre_foto)
                                try:
                                    with open(ruta, "wb") as f:
                                        f.write(bytes_guardar)
                                    if idx == 0:
                                        foto_ruta = ruta
                                    fotos_guardadas.append(ruta)
                                except Exception as e:
                                    st.warning(t["gasto_foto_no"].format(e=str(e)))
                        fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        detalle_gasto = (desc_gasto or "Gasto")[:MAX_CARACTERES_DESCRIPCION_T]
                        if monto_final >= UMBRAL_GASTO_APROBACION and (aprobado_por_gasto or "").strip():
                            detalle_gasto = (detalle_gasto + " (Aprobado por: " + (aprobado_por_gasto or "").strip()[:80] + ")")[:MAX_CARACTERES_DESCRIPCION_T]
                        if lang == "ES":
                            tipo_guardar = tipo_seleccionado
                        else:
                            try:
                                idx = TIPOS_GASTO_EN.index(tipo_seleccionado)
                                tipo_guardar = TIPOS_GASTO_ES[idx]
                            except (ValueError, IndexError):
                                tipo_guardar = DEFAULT_TIPO_GASTO
                        nueva = pd.DataFrame([{
                            "id_registro": rid,
                            "fecha": fecha_ahora,
                            "detalle": detalle_gasto,
                            "tipo_gasto": tipo_guardar,
                            "ingreso": 0,
                            "gastos": monto_final,
                            "total_ingresos": 0,
                            "total_gastos": 0,
                            "saldo_actual": 0
                        }])
                        df = pd.concat([df, nueva], ignore_index=True)
                        if guardar_db(df, t):
                            meta_g = cargar_arqueo_meta()
                            if rid not in meta_g:
                                meta_g[rid] = {}
                            meta_g[rid]["ip_dispositivo"] = _get_client_ip()
                            meta_g[rid]["fecha"] = fecha_ahora
                            if len(fotos_guardadas) > 1:
                                meta_g[rid]["fotos_adicionales"] = fotos_guardadas[1:]
                            guardar_arqueo_meta(meta_g)
                            audit_log(usuario_actual, "gasto_registrado", f"{rid} ${monto_final:.2f} {detalle_gasto[:50]}")
                            if foto_subidas:
                                bytes_foto = foto_subidas[0].getvalue()
                                hash_foto = _hash_imagen(bytes_foto)
                                ocr_text = _ocr_imagen(bytes_foto)
                                datos_f = _extraer_datos_factura(ocr_text)
                                guardar_factura({
                                    "id_registro": rid,
                                    "hash_imagen": hash_foto,
                                    "ocr_text": ocr_text,
                                    "total": monto_final,
                                    "impuesto": datos_f.get("impuesto"),
                                    "subtotal": datos_f.get("subtotal"),
                                    "comercio": detalle_gasto,
                                    "tipo_gasto": tipo_guardar,
                                    "items": datos_f.get("items", []),
                                    "fecha_factura": datos_f.get("fecha_texto", ""),
                                    "fecha_registro": fecha_ahora,
                                })
                            for key in ("monto_sugerido", "descripcion_sugerida", "factura_detectada_actual", "ocr_text_actual", "hash_foto_actual"):
                                st.session_state.pop(key, None)
                            st.session_state["limpiar_gasto"] = True
                            st.success(f"{t['gasto_registrado']} ${monto_final:.2f}")
                            st.rerun()
            with st.expander(f"📥 {t.get('importar_gastos', 'Importar gastos')}", expanded=False):
                st.caption(t.get("importar_gastos_ayuda", ""))
                archivo_import = st.file_uploader(" ", type=["csv", "xlsx", "xls"], key="import_gastos_file")
                if archivo_import:
                    try:
                        if archivo_import.name.lower().endswith(".csv"):
                            df_imp = pd.read_csv(archivo_import, encoding="utf-8-sig")
                        else:
                            df_imp = pd.read_excel(archivo_import)
                        cols_necesarias = ["fecha", "detalle", "gastos"]
                        if not all(c in df_imp.columns for c in cols_necesarias):
                            st.warning("El archivo debe tener columnas: fecha, detalle, gastos (y opcional: tipo_gasto)")
                        else:
                            tipo_col = "tipo_gasto" if "tipo_gasto" in df_imp.columns else None
                            importados = 0
                            for _, row in df_imp.iterrows():
                                try:
                                    fecha_val = str(row.get("fecha", ""))[:19]
                                    det = str(row.get("detalle", "Gasto"))[:MAX_CARACTERES_DESCRIPCION_T]
                                    gas = float(row.get("gastos", 0) or 0)
                                    if gas <= 0:
                                        continue
                                    tipo_imp = str(row.get(tipo_col, DEFAULT_TIPO_GASTO))[:50] if tipo_col else DEFAULT_TIPO_GASTO
                                    rid = generar_id_gasto()
                                    nueva = pd.DataFrame([{"id_registro": rid, "fecha": fecha_val, "detalle": det, "tipo_gasto": tipo_imp, "ingreso": 0, "gastos": gas, "total_ingresos": 0, "total_gastos": 0, "saldo_actual": 0}])
                                    df = pd.concat([df, nueva], ignore_index=True)
                                    importados += 1
                                except Exception:
                                    continue
                            if importados > 0 and guardar_db(df, t):
                                df = _recalcular_totales_ledger(df)
                                guardar_db(df, t)
                                audit_log(usuario_actual, "gastos_importados", f"{importados} registros")
                                st.success(f"Importados {importados} gastos correctamente.")
                                st.rerun()
                            elif importados == 0:
                                st.info("No se encontraron filas válidas para importar.")
                    except Exception as ex:
                        st.error(f"Error al importar: {ex}")
            with st.expander(f"📋 {t.get('rendicion_cuentas_titulo', 'Rendición de cuentas (cuadre rápido)')}", expanded=False):
                st.caption(t.get("rendicion_cuentas_ayuda", "Fondo entregado para insumos (ej. $200). Registre cada gasto (harina $50, azúcar $20, taxi $40). La devolución ($90) se registra como ingreso. Cuadre: Fondo − Gastado = Devolución."))
                responsable_rend = st.text_input(t.get("rendicion_responsable", "Responsable (ej. Hermanas)"), key="rendicion_responsable", max_chars=80, placeholder="Hermanas")
                fondo_entregado = st.number_input(t.get("rendicion_fondo_entregado", "Fondo entregado ($)"), min_value=0.0, value=0.0, step=10.0, key="rendicion_fondo")
                st.markdown(t.get("rendicion_lineas_gastos", "**Gastos (concepto y monto)**"))
                lineas_rend = []
                for i in range(8):
                    c1, c2 = st.columns([2, 1])
                    with c1:
                        concepto_rend = st.text_input(t.get("concepto", "Concepto"), key=f"rendicion_conc_{i}", placeholder="Harina, Azúcar, Taxi...")
                    with c2:
                        monto_rend = st.number_input("$", min_value=0.0, value=0.0, step=5.0, key=f"rendicion_monto_{i}")
                    if concepto_rend or monto_rend > 0:
                        lineas_rend.append((concepto_rend.strip() or f"Item {i+1}", round(float(monto_rend), 2)))
                total_gastado = sum(m for _, m in lineas_rend)
                devolucion_rend = st.number_input(t.get("rendicion_devolucion", "Devolución ($)"), min_value=0.0, value=0.0, step=5.0, key="rendicion_devolucion")
                esperado = round(fondo_entregado - total_gastado, 2)
                cuadre_ok = abs(devolucion_rend - esperado) < 0.02
                if fondo_entregado > 0:
                    if cuadre_ok:
                        st.success(f"✓ {t.get('rendicion_cuadre_ok', 'Cuadre correcto')}: {fondo_entregado:.2f} − {total_gastado:.2f} = {devolucion_rend:.2f}")
                    else:
                        st.warning(f"⚠ {t.get('rendicion_cuadre_revisar', 'Revisar cuadre')}: Fondo {fondo_entregado:.2f} − Gastado {total_gastado:.2f} = {esperado:.2f} (devolución ingresada: {devolucion_rend:.2f})")
                if st.button(t.get("rendicion_registrar_btn", "Registrar rendición"), key="btn_rendicion"):
                    if not lineas_rend and devolucion_rend <= 0:
                        st.warning(t.get("rendicion_sin_datos", "Ingrese al menos un gasto o la devolución."))
                    else:
                        fecha_ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        resp_label = (responsable_rend or "Rendición").strip()[:80]
                        tipo_operativo = "Operativo" if lang == "ES" else "Operational"
                        for concepto, monto in lineas_rend:
                            if monto <= 0:
                                continue
                            rid = generar_id_gasto()
                            det = f"Rendición {resp_label}: {concepto}"[:200]
                            nueva = pd.DataFrame([{"id_registro": rid, "fecha": fecha_ahora, "detalle": det, "tipo_gasto": tipo_operativo, "ingreso": 0, "gastos": monto, "total_ingresos": 0, "total_gastos": 0, "saldo_actual": 0}])
                            df = pd.concat([df, nueva], ignore_index=True)
                        if devolucion_rend > 0:
                            rid_ing = generar_id()
                            tipo_dev_es = "Devolución de fondo"
                            tipo_dev_en = "Fund return"
                            tipo_dev = tipo_dev_es if lang == "ES" else tipo_dev_en
                            det_ing = f"Devolución de fondo — {resp_label}" if lang == "ES" else f"Fund return — {resp_label}"
                            nueva_ing = pd.DataFrame([{"id_registro": rid_ing, "fecha": fecha_ahora, "detalle": det_ing[:200], "tipo_gasto": tipo_dev_es, "ingreso": round(devolucion_rend, 2), "gastos": 0, "total_ingresos": 0, "total_gastos": 0, "saldo_actual": 0}])
                            df = pd.concat([df, nueva_ing], ignore_index=True)
                        df = _recalcular_totales_ledger(df)
                        if guardar_db(df, t):
                            audit_log(usuario_actual, "rendicion_cuentas", f"{resp_label} gastos={len(lineas_rend)} dev={devolucion_rend:.2f}")
                            st.success(t.get("rendicion_registrada_ok", "Rendición registrada. Gastos y devolución guardados."))
                            st.rerun()
                        else:
                            st.error(t.get("error_guardar", "Error al guardar."))
            with st.expander(f"🔔 {t.get('recordatorios_recurrentes', 'Recordatorios')}", expanded=False):
                suministros_pend = cargar_suministros()
                mes_actual = datetime.now().strftime("%Y-%m")
                if not df.empty:
                    try:
                        df_r = df.copy()
                        df_r["_fecha"] = pd.to_datetime(df_r["fecha"].astype(str).str[:10], errors="coerce")
                        df_r["_mes"] = df_r["_fecha"].dt.strftime("%Y-%m")
                        df_mes = df_r[df_r["_mes"] == mes_actual]
                        detalle_mes = set((df_mes["detalle"].fillna("").astype(str).str[:50]).str.upper())
                        pendientes = [s for s in suministros_pend if s.upper()[:30] not in " ".join(detalle_mes)[:500]]
                        if pendientes:
                            st.caption(t.get("recordatorios_pendientes", "Suministros que podrían estar pendientes este mes:"))
                            for p in pendientes[:10]:
                                st.markdown(f"- {p}")
                        else:
                            st.info(t.get("recordatorios_todos_ok", "Todos los suministros habituales parecen registrados este mes."))
                    except Exception:
                        st.caption("No se pudo calcular." if lang == "ES" else "Could not calculate.")
                else:
                    st.caption(t.get("recordatorios_registre", "Registre gastos para ver recordatorios."))
            st.markdown("""<style>
            .main [data-testid="stExpander"] div[data-testid="stVerticalBlock"] .stButton > button,
            .main [data-testid="stExpander"] [data-testid="stFormSubmitButton"] > button,
            .main .block-container [data-testid="stExpander"] .stButton > button { min-height: 3.2rem !important; padding: 0.9rem 1.6rem !important; font-size: 1.15rem !important; }
            /* Refrescar y todos los botones en expander: metálico azul */
            .main [data-testid="stExpander"] .stButton > button {
                background: linear-gradient(180deg, #7eb3f0 0%, #99d0ff 18%, #6ba3e0 35%, #3d7ab8 55%, #2568a0 75%, #2d6a9f 100%) !important;
                color: #ffffff !important;
                box-shadow: inset 0 2px 6px rgba(255,255,255,0.4), inset 0 -3px 6px rgba(0,0,0,0.3), 0 4px 12px rgba(0,0,0,0.4) !important;
            }
            .main [data-testid="stExpander"] .stButton > button:hover {
                background: linear-gradient(180deg, #8ec5ff 0%, #a8dcff 20%, #7eb3f0 40%, #4a8ac8 60%, #2d78b0 80%, #3a7ab8 100%) !important;
                transform: translateY(-3px) !important;
            }
            .main [data-testid="stExpander"] .stButton > button:active { transform: translateY(1px) !important; }
            @media (max-width: 768px) {
            .main .stButton > button, .main [data-testid="stFormSubmitButton"] > button { min-height: 3.5rem !important; padding: 1rem 1.8rem !important; font-size: 1.15rem !important; }
            }
            </style>""", unsafe_allow_html=True)
        return

    # ---------- PRESUPUESTO Y METAS (Visión) ----------
    if pagina_act == "presupuesto_metas":
        st.markdown(f"## 🎯 {t['presupuesto_metas']}")
        st.caption(t["presupuesto_metas_sub"])
        presup = cargar_presupuesto()
        anio_sel = presup.get("anio", datetime.now().year)
        tipos_es = TIPOS_GASTO_ES
        tipos_en = TIPOS_GASTO_EN
        por_tipo = presup.get("por_tipo") or {}
        meta_ingresos = float(presup.get("meta_ingresos") or 0)

        anio_str = str(anio_sel)
        df_p = df.copy() if not df.empty else pd.DataFrame()
        if not df_p.empty and "fecha" in df_p.columns and "tipo_gasto" in df_p.columns and "gastos" in df_p.columns:
            try:
                df_p["_fecha"] = pd.to_datetime(df_p["fecha"].astype(str).str[:10], errors="coerce")
                df_p["_anio"] = df_p["_fecha"].dt.year.astype(str)
                df_p = df_p[df_p["_anio"] == anio_str]
                gastos_num = pd.to_numeric(df_p["gastos"], errors="coerce").fillna(0)
                real_por_tipo = df_p.assign(gastos_num=gastos_num).groupby("tipo_gasto")["gastos_num"].sum().to_dict()
                ingreso_num = pd.to_numeric(df_p["ingreso"], errors="coerce").fillna(0)
                ingresos_reales = float(ingreso_num.sum())
            except Exception:
                real_por_tipo = {}
                ingresos_reales = 0.0
        else:
            real_por_tipo = {}
            ingresos_reales = 0.0

        total_presup = sum(float(por_tipo.get(t, 0) or 0) for t in tipos_es)
        total_real_gastos = sum(float(v) for v in real_por_tipo.values())
        pct_ing = (ingresos_reales / meta_ingresos * 100) if meta_ingresos > 0 else 0.0
        m1, m2, m3 = st.columns(3)
        with m1:
            st.metric(t.get("presupuesto_total_gastos", "Presupuesto gastos (año)"), f"${total_presup:,.2f}")
        with m2:
            st.metric(t.get("presupuesto_real_gastos", "Gastos reales (año)"), f"${total_real_gastos:,.2f}")
        with m3:
            st.metric(t.get("meta_ingresos_label", "Meta ingresos"), f"${ingresos_reales:,.2f} / ${meta_ingresos:,.2f}" if meta_ingresos > 0 else f"${ingresos_reales:,.2f}")

        with st.expander(f"✏️ {t.get('presupuesto_por_tipo', 'Presupuesto por tipo de gasto')} — {t.get('presupuesto_anio', 'Año')}", expanded=False):
            anio_input = st.number_input(t.get("presupuesto_anio", "Año"), min_value=datetime.now().year - 2, max_value=datetime.now().year + 1, value=anio_sel, step=1, key="presup_anio_input")
            tipos_label = tipos_es if lang == "ES" else tipos_en
            nuevos = {}
            cols = st.columns(2)
            for i, tipo in enumerate(tipos_es):
                with cols[i % 2]:
                    val = por_tipo.get(tipo, 0)
                    num = st.number_input(
                        tipos_label[i],
                        min_value=0.0, value=float(val) if val else 0.0, step=100.0,
                        key=f"presup_{tipo}_{anio_sel}",
                    )
                    nuevos[tipo] = num
            meta_new = st.number_input(t.get("meta_ingresos_label", "Meta de ingresos ($)"), min_value=0.0, value=meta_ingresos, step=500.0, key="presup_meta_ingresos")
            if st.button(t.get("presupuesto_guardar", "Guardar presupuesto"), key="btn_guardar_presupuesto"):
                presup["por_tipo"] = nuevos
                presup["meta_ingresos"] = meta_new
                presup["anio"] = int(anio_input)
                if guardar_presupuesto(presup):
                    st.success(t.get("presupuesto_guardado_ok", "Presupuesto guardado."))
                    st.rerun()

        st.markdown(f"**{t.get('presupuesto_vs_real', 'Presupuesto vs real')}** — {t.get('presupuesto_real_titulo', 'Real (este año)')} ({anio_sel})")
        filas = []
        for i, tipo in enumerate(tipos_es):
            presup_val = float(por_tipo.get(tipo) or 0)
            real_val = float(real_por_tipo.get(tipo) or 0)
            pct = (real_val / presup_val * 100) if presup_val > 0 else (100.0 if real_val > 0 else 0.0)
            estado = "🔴" if pct > 100 else ("🟡" if pct > 80 else "🟢")
            nombre_tipo = tipos_en[i] if lang == "EN" else tipo
            filas.append({"tipo": nombre_tipo, "presupuesto": presup_val, "real": real_val, "pct": pct, "estado": estado})
        otros_tipos = [k for k in real_por_tipo if k not in tipos_es and str(k).strip()]
        for ot in otros_tipos:
            real_val = float(real_por_tipo.get(ot) or 0)
            if real_val > 0:
                nombre_ot = ot if isinstance(ot, str) else str(ot)
                filas.append({"tipo": nombre_ot, "presupuesto": 0.0, "real": real_val, "pct": 0.0, "estado": "⚪"})
        if filas:
            df_tabla = pd.DataFrame(filas)
            st.dataframe(
                df_tabla.assign(
                    presupuesto=df_tabla["presupuesto"].apply(lambda x: f"${x:,.2f}"),
                    real=df_tabla["real"].apply(lambda x: f"${x:,.2f}"),
                    pct=df_tabla["pct"].apply(lambda x: f"{x:.1f}%" if x > 0 else "—"),
                )[["estado", "tipo", "presupuesto", "real", "pct"]],
                width="stretch", hide_index=True,
            )
            for r in filas:
                if r["pct"] > 100:
                    st.warning(f"⚠️ {r['tipo']}: {t.get('presupuesto_alerta_superado', 'Superó el presupuesto')} ({r['pct']:.1f}%)")
        else:
            st.caption(t.get("presupuesto_sin_datos", "No hay movimientos en el período para comparar."))

        st.markdown("**" + t.get("meta_ingresos_label", "Meta de ingresos ($)") + "**")
        if meta_ingresos > 0:
            st.progress(min(1.0, ingresos_reales / meta_ingresos))
            st.caption(f"${ingresos_reales:,.2f} / ${meta_ingresos:,.2f} ({pct_ing:.1f}%)")
            if ingresos_reales >= meta_ingresos:
                st.success(t.get("presupuesto_meta_alcanzada", "¡Meta de ingresos alcanzada!"))
        else:
            st.caption(f"${ingresos_reales:,.2f} " + t.get("presupuesto_defina_meta", "(defina meta arriba)"))

        try:
            buf_p = BytesIO()
            resum = pd.DataFrame(filas) if filas else pd.DataFrame(columns=["tipo", "presupuesto", "real", "pct"])
            if not resum.empty:
                resum["meta_ingresos"] = meta_ingresos
                resum["ingresos_reales"] = ingresos_reales
            else:
                resum = pd.DataFrame([{"tipo": "—", "presupuesto": total_presup, "real": total_real_gastos, "pct": 0, "meta_ingresos": meta_ingresos, "ingresos_reales": ingresos_reales}])
            resum.to_csv(buf_p, index=False, encoding="utf-8-sig")
            buf_p.seek(0)
            st.download_button(
                label=t.get("presupuesto_exportar", "Exportar presupuesto vs real"),
                data=buf_p.getvalue(),
                file_name=f"presupuesto_vs_real_{anio_sel}.csv",
                mime="text/csv",
                key="btn_export_presupuesto",
            )
        except Exception:
            pass
        return

    # ---------- EVENTOS / INVERSIONES (cruze gastos–ventas–margen, rentabilidad, informe) ----------
    if pagina_act == "eventos":
        st.markdown(f"## 📌 {t.get('eventos_inversiones', 'Eventos / Inversiones')}")
        st.caption(t.get("eventos_inversiones_sub", "Cruze de gastos, ventas y margen por evento. Rentabilidad, donaciones, mano de obra, informe y recomendación."))
        data_ev = cargar_eventos()
        lista_ev = data_ev.get("eventos") or []
        buscar_ev = st.text_input(t.get("eventos_buscar", "Buscar por nombre de evento o inversión"), key="eventos_buscar_input", placeholder=t.get("eventos_buscar_placeholder", "Ej: Venta comida marzo"))
        if buscar_ev and buscar_ev.strip():
            lista_ev = [e for e in lista_ev if (e.get("nombre") or "").lower().find(buscar_ev.strip().lower()) >= 0]
        with st.expander(f"➕ {t.get('eventos_nuevo', 'Nuevo evento / inversión')}", expanded=False):
            nombre_ev = st.text_input(t.get("eventos_nombre", "Nombre del evento o inversión"), key="ev_nombre", max_chars=120, placeholder="Venta comida hermanas Marzo 2025")
            fecha_ev = st.date_input(t.get("eventos_fecha", "Fecha"), value=datetime.now().date(), key="ev_fecha")
            gastos_ev = st.number_input(t.get("eventos_gastos", "Gastos ($)"), min_value=0.0, value=0.0, step=10.0, key="ev_gastos")
            ventas_ev = st.number_input(t.get("eventos_ventas", "Ventas / Ingresos ($)"), min_value=0.0, value=0.0, step=10.0, key="ev_ventas")
            margen_ev = round(ventas_ev - gastos_ev, 2)
            rentable_ev = margen_ev > 0
            st.metric(t.get("eventos_margen", "Margen"), f"${margen_ev:,.2f}", f"{t.get('eventos_rentable', 'Rentable') if rentable_ev else t.get('eventos_no_rentable', 'No rentable')}")
            donaciones_ev = st.number_input(t.get("eventos_donaciones", "Donaciones ($)"), min_value=0.0, value=0.0, step=5.0, key="ev_donaciones")
            perdidas_ev = st.number_input(t.get("eventos_perdidas", "Pérdidas ($)"), min_value=0.0, value=0.0, step=5.0, key="ev_perdidas")
            mano_obra_pagada_ev = st.number_input(t.get("eventos_mano_obra_pagada", "Mano de obra pagada ($)"), min_value=0.0, value=0.0, step=5.0, key="ev_mano_obra_pagada")
            mano_obra_donada_ev = st.radio(t.get("eventos_mano_obra_donada", "Mano de obra donada"), options=["no", "si"], format_func=lambda x: t.get("eventos_si", "Sí") if x == "si" else t.get("eventos_no", "No"), key="ev_mano_obra_donada", horizontal=True)
            quien_tipo_ev = st.selectbox(t.get("eventos_mano_obra_por", "Mano de obra donada por"), options=["", "hermanas", "miembros", "ambos"], format_func=lambda x: {"": "—", "hermanas": t.get("eventos_por_hermanas", "Hermanas"), "miembros": t.get("eventos_por_miembros", "Miembros"), "ambos": t.get("eventos_por_ambos", "Hermanas y miembros")}.get(x, x), key="ev_quien_tipo")
            quien_dono_ev = st.text_input(t.get("eventos_quien_dono", "Quiénes donaron mano de obra (nombres)"), key="ev_quien_dono", placeholder="María, Ana, Juan")
            recomendacion_ev = st.radio(t.get("eventos_recomendacion", "¿Se recomienda repetir?"), options=["si", "no", "tal_vez"], format_func=lambda x: {"si": t.get("eventos_recom_si", "Sí, recomendado"), "no": t.get("eventos_recom_no", "No"), "tal_vez": t.get("eventos_recom_tal_vez", "Tal vez")}.get(x, x), key="ev_recomendacion", horizontal=True)
            nota_ev = st.text_area(t.get("eventos_nota", "Nota"), key="ev_nota", max_chars=500, placeholder=t.get("eventos_nota_placeholder", "Detalles del evento"))
            if st.button(t.get("eventos_guardar_btn", "Guardar evento"), key="btn_guardar_evento"):
                if not nombre_ev or not nombre_ev.strip():
                    st.warning(t.get("eventos_nombre_requerido", "Indique el nombre del evento."))
                else:
                    ev_id = f"EV-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}"
                    nuevo = {
                        "id": ev_id,
                        "nombre": nombre_ev.strip(),
                        "fecha": fecha_ev.strftime("%Y-%m-%d"),
                        "gastos": round(float(gastos_ev), 2),
                        "ventas": round(float(ventas_ev), 2),
                        "margen": margen_ev,
                        "rentable": rentable_ev,
                        "donaciones": round(float(donaciones_ev), 2),
                        "perdidas": round(float(perdidas_ev), 2),
                        "mano_obra_pagada": round(float(mano_obra_pagada_ev), 2),
                        "mano_obra_donada": mano_obra_donada_ev == "si",
                        "mano_obra_por": (quien_tipo_ev or "").strip() or None,
                        "quien_dono_mano_obra": (quien_dono_ev or "").strip()[:200],
                        "recomendacion": recomendacion_ev,
                        "nota": (nota_ev or "").strip()[:500],
                    }
                    lista_ev = data_ev.get("eventos") or []
                    lista_ev.append(nuevo)
                    data_ev["eventos"] = lista_ev
                    if guardar_eventos(data_ev):
                        audit_log(usuario_actual, "evento_creado", ev_id)
                        st.success(t.get("eventos_guardado_ok", "Evento guardado."))
                        st.rerun()
                    else:
                        st.error(t.get("error_guardar", "Error al guardar."))
        st.markdown(f"**{t.get('eventos_listado', 'Listado de eventos')}**")
        if not lista_ev:
            st.info(t.get("eventos_sin_eventos", "No hay eventos. Cree uno en «Nuevo evento / inversión»."))
        else:
            for ev in sorted(lista_ev, key=lambda x: x.get("fecha") or "", reverse=True):
                with st.expander(f"{'✅' if ev.get('rentable') else '⚠️'} {ev.get('nombre', '—')} — {ev.get('fecha', '')} | ${ev.get('margen', 0):,.2f}", expanded=False):
                    gastos_v = ev.get("gastos") or 0
                    ventas_v = ev.get("ventas") or 0
                    margen_v = ev.get("margen", ventas_v - gastos_v)
                    st.markdown(f"### {t.get('eventos_informe_titulo', 'Informe del evento')}")
                    informe_lines = [
                        f"**{t.get('eventos_cruce', 'Cruze')}:** {t.get('eventos_gaste', 'Gasté')} ${gastos_v:,.2f} · {t.get('eventos_vendi', 'Vendí')} ${ventas_v:,.2f} · **{t.get('eventos_margen', 'Margen')} ${margen_v:,.2f}**",
                        f"**{t.get('eventos_rentable', 'Rentable')}:** " + (t.get("eventos_rentable", "Rentable") if ev.get("rentable") else t.get("eventos_no_rentable", "No rentable")),
                        f"**{t.get('eventos_donaciones', 'Donaciones')}:** ${ev.get('donaciones', 0):,.2f} · **{t.get('eventos_perdidas', 'Pérdidas')}:** ${ev.get('perdidas', 0):,.2f}",
                        f"**{t.get('eventos_mano_obra_pagada', 'Mano de obra pagada')}:** ${ev.get('mano_obra_pagada', 0):,.2f}",
                        f"**{t.get('eventos_mano_obra_donada', 'Mano de obra donada')}:** " + (t.get("eventos_si", "Sí") if ev.get("mano_obra_donada") else t.get("eventos_no", "No")),
                    ]
                    por_tipo = ev.get("mano_obra_por")
                    if por_tipo:
                        por_label = {"hermanas": t.get("eventos_por_hermanas", "Hermanas"), "miembros": t.get("eventos_por_miembros", "Miembros"), "ambos": t.get("eventos_por_ambos", "Hermanas y miembros")}.get(por_tipo, por_tipo)
                        informe_lines.append(f"**{t.get('eventos_mano_obra_por', 'Donada por')}:** {por_label}")
                    if ev.get("quien_dono_mano_obra"):
                        informe_lines.append(f"**{t.get('eventos_quien_dono', 'Quiénes donaron mano de obra')}:** {ev['quien_dono_mano_obra']}")
                    rec = ev.get("recomendacion", "")
                    rec_label = {"si": t.get("eventos_recom_si", "Sí, recomendado"), "no": t.get("eventos_recom_no", "No"), "tal_vez": t.get("eventos_recom_tal_vez", "Tal vez")}.get(rec, rec)
                    informe_lines.append(f"**{t.get('eventos_recomendacion', '¿Se recomienda repetir?')}:** {rec_label}")
                    if ev.get("nota"):
                        informe_lines.append(f"**{t.get('eventos_nota', 'Nota')}:** {ev['nota']}")
                    for line in informe_lines:
                        st.markdown(line)
                    informe_texto = "\n".join([ev.get("nombre", "—"), ev.get("fecha", ""), ""] + [l.replace("**", "").replace("*", "") for l in informe_lines])
                    buf_inf = BytesIO()
                    buf_inf.write(informe_texto.encode("utf-8"))
                    buf_inf.seek(0)
                    st.download_button(
                        label=t.get("eventos_descargar_informe", "Descargar informe"),
                        data=buf_inf.getvalue(),
                        file_name=f"informe_evento_{(ev.get('nombre') or 'evento')[:40].replace('/', '-')}_{ev.get('fecha', '')}.txt",
                        mime="text/plain",
                        key=f"ev_informe_{ev.get('id', '')}",
                    )
                    st.markdown("---")
                    st.markdown(t.get("eventos_botones_registrar", "**Registrar en libro:**"))
                    col_r1, col_r2 = st.columns(2)
                    with col_r1:
                        if tiene_permiso(usuario_actual, "ver_ingresar_bendicion") and st.button(t.get("eventos_btn_registrar_ingreso", "Registrar ingreso"), key=f"ev_ing_{ev.get('id', '')}"):
                            st.session_state["pagina"] = "arqueo_caja"
                            st.session_state["sidebar_state"] = "collapsed"
                            st.session_state["sidebar_collapse_requested"] = True
                            st.rerun()
                    with col_r2:
                        if tiene_permiso(usuario_actual, "ver_registrar_gasto") and st.button(t.get("eventos_btn_registrar_gasto", "Registrar gasto"), key=f"ev_gas_{ev.get('id', '')}"):
                            st.session_state["pagina"] = "tesoreria"
                            st.session_state["sidebar_state"] = "collapsed"
                            st.session_state["sidebar_collapse_requested"] = True
                            st.rerun()
        return

    # ---------- CONTABILIDAD (Bóveda Histórica) ----------
    st.markdown(f"## 📊 {t['contabilidad']}")
    st.caption(t["contabilidad_sub"])
    if not df.empty:
        ok_int, msg_int = verificar_integridad_ledger(df)
        if not ok_int:
            _tit = (t.get("integridad_banner_titulo", "Error de integridad detectado")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
            _des = (t.get("integridad_banner_deshabilitado", "El borrado está deshabilitado hasta corregir o restaurar.")).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
            st.markdown(f"""<div style="background: linear-gradient(135deg, #4a0e0e 0%, #7a1a1a 100%); border: 2px solid #ff5252; border-radius: 8px; padding: 1rem 1.25rem; margin-bottom: 1rem;">
            <p style="color: #ffcccc; margin: 0; font-weight: bold;">⚠️ {_tit}</p>
            <p style="color: #ffb3b3; margin: 0.5rem 0 0 0; font-size: 0.95rem;">{_des}</p>
            </div>""", unsafe_allow_html=True)
            st.error(f"**{t.get('integridad_error_titulo', t['integridad_aviso'])}** — {t['integridad_aviso']} ({msg_int or ''})")
            st.caption("No se recomienda borrar registros hasta restaurar o corregir la integridad." if lang == "ES" else "Do not delete records until integrity is restored or corrected.")

    # ---------- DASHBOARD RESUMEN (saldo, ingresos/gastos del mes) ----------
    if tiene_permiso(usuario_actual, "ver_hoja_contable") and not df.empty:
        mes_actual = datetime.now().strftime("%Y-%m")
        df_dash = df.copy()
        try:
            df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str).str[:10], errors="coerce")
            df_mes = df_dash[df_dash["_fecha"].dt.strftime("%Y-%m") == mes_actual]
        except Exception:
            df_mes = pd.DataFrame()
        if not df_mes.empty:
            ing_mes = float(df_mes["ingreso"].sum())
            gas_mes = float(df_mes["gastos"].sum())
        else:
            ing_mes = gas_mes = 0.0
        saldo_act = float(df["saldo_actual"].iloc[-1]) if "saldo_actual" in df.columns and len(df) else 0.0
        c1, c2, c3 = st.columns(3)
        with c1:
            st.metric(t["saldo_actual"], f"${saldo_act:,.2f}")
        with c2:
            st.metric(t["ingresos_mes"], f"${ing_mes:,.2f}")
        with c3:
            st.metric(t["gastos_mes"], f"${gas_mes:,.2f}")

        # ---------- Resumen por período (mes / trimestre / año) ----------
        with st.expander(f"📅 {t.get('resumen_periodo', 'Resumen por período')}", expanded=False):
            periodo_sel = st.radio(
                t.get("resumen_periodo", "Período"),
                options=["mes", "trimestre", "ano"],
                format_func=lambda x: t.get(f"periodo_{x}", x),
                key="contabilidad_periodo_sel",
                horizontal=True,
            )
            hoy = datetime.now().date()
            if periodo_sel == "mes":
                inicio = hoy.replace(day=1)
                fin = hoy
                prev_fin = inicio - timedelta(days=1)
                prev_inicio = prev_fin.replace(day=1)
            elif periodo_sel == "trimestre":
                q = (hoy.month - 1) // 3 + 1
                inicio = hoy.replace(month=(q - 1) * 3 + 1, day=1)
                fin = hoy
                prev_fin = inicio - timedelta(days=1)
                prev_inicio = prev_fin.replace(month=10, day=1) if q == 1 else prev_fin.replace(month=(q - 2) * 3 + 1, day=1)
            else:
                inicio = hoy.replace(month=1, day=1)
                fin = hoy
                prev_fin = inicio - timedelta(days=1)
                prev_inicio = prev_fin.replace(month=1, day=1)
            try:
                df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str).str[:10], errors="coerce")
                mask = (df_dash["_fecha"].dt.date >= inicio) & (df_dash["_fecha"].dt.date <= fin)
                df_per = df_dash.loc[mask]
                mask_prev = (df_dash["_fecha"].dt.date >= prev_inicio) & (df_dash["_fecha"].dt.date <= prev_fin)
                df_prev = df_dash.loc[mask_prev]
            except Exception:
                df_per = df_dash
                df_prev = pd.DataFrame()
            ing_per = float(df_per["ingreso"].sum()) if not df_per.empty else 0.0
            gas_per = float(df_per["gastos"].sum()) if not df_per.empty else 0.0
            res_per = ing_per - gas_per
            saldo_cierre_per = float(df_per["saldo_actual"].iloc[-1]) if not df_per.empty and "saldo_actual" in df_per.columns else saldo_act
            ing_prev = float(df_prev["ingreso"].sum()) if not df_prev.empty else 0.0
            gas_prev = float(df_prev["gastos"].sum()) if not df_prev.empty else 0.0
            res_prev = ing_prev - gas_prev
            pc1, pc2, pc3, pc4 = st.columns(4)
            with pc1:
                st.metric(t.get("ingresos_periodo", "Ingresos"), f"${ing_per:,.2f}")
            with pc2:
                st.metric(t.get("gastos_periodo", "Gastos"), f"${gas_per:,.2f}")
            with pc3:
                st.metric(t.get("resultado_periodo", "Resultado"), f"${res_per:,.2f}")
            with pc4:
                st.metric(t.get("saldo_cierre", "Saldo al cierre"), f"${saldo_cierre_per:,.2f}")
            st.caption(t.get("comparar_periodo_anterior", "Comparar con período anterior"))
            diff_ing = ing_per - ing_prev
            diff_gas = gas_per - gas_prev
            pct_ing = (diff_ing / ing_prev * 100) if ing_prev else 0
            pct_gas = (diff_gas / gas_prev * 100) if gas_prev else 0
            comp1, comp2 = st.columns(2)
            with comp1:
                st.metric(f"{t.get('ingresos_periodo', 'Ingresos')} {t.get('vs_anterior', 'vs anterior')}", f"${diff_ing:+,.2f}", f"{pct_ing:+.1f}%")
            with comp2:
                st.metric(f"{t.get('gastos_periodo', 'Gastos')} {t.get('vs_anterior', 'vs anterior')}", f"${diff_gas:+,.2f}", f"{pct_gas:+.1f}%")

        # ---------- Exportar resumen para contador (totales por mes) ----------
        with st.expander(f"📤 {t.get('exportar_resumen_contador', 'Exportar resumen para contador')}", expanded=False):
            st.caption(t.get("exportar_resumen_ayuda", ""))
            try:
                df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str).str[:10], errors="coerce")
                df_dash["_mes"] = df_dash["_fecha"].dt.to_period("M").astype(str)
                resumen_mes = df_dash.groupby("_mes").agg(
                    ingresos=("ingreso", "sum"),
                    gastos=("gastos", "sum"),
                ).reset_index()
                resumen_mes["resultado"] = resumen_mes["ingresos"] - resumen_mes["gastos"]
                resumen_mes = resumen_mes.rename(columns={"_mes": "mes"})
                buf = BytesIO()
                resumen_mes.to_csv(buf, index=False, encoding="utf-8-sig")
                buf.seek(0)
                st.download_button(
                    label=t.get("exportar_btn_csv_mes", "CSV (totales por mes)"),
                    data=buf.getvalue(),
                    file_name=f"resumen_contable_{datetime.now().strftime('%Y-%m-%d')}.csv",
                    mime="text/csv",
                    key="btn_exportar_resumen_csv",
                )
                if "tipo_gasto" in df_dash.columns:
                    try:
                        df_gt_export = df_dash.copy()
                        df_gt_export["gastos_num"] = pd.to_numeric(df_gt_export["gastos"], errors="coerce").fillna(0)
                        df_gt_export = df_gt_export[df_gt_export["gastos_num"] > 0]
                        if not df_gt_export.empty:
                            resumen_tipo = df_gt_export.groupby(["_mes", "tipo_gasto"]).agg(gastos=("gastos_num", "sum")).reset_index()
                            resumen_tipo = resumen_tipo.rename(columns={"_mes": "mes"})
                            buf_xlsx = BytesIO()
                            with pd.ExcelWriter(buf_xlsx, engine="openpyxl") as writer:
                                resumen_mes.to_excel(writer, index=False, sheet_name="Por mes")
                                resumen_tipo.to_excel(writer, index=False, sheet_name="Gastos por mes y tipo")
                            buf_xlsx.seek(0)
                            st.download_button(
                                label=t.get("exportar_btn_excel_mes_tipo", "Excel (por mes y tipo de gasto)"),
                                data=buf_xlsx.getvalue(),
                                file_name=f"resumen_contable_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="btn_exportar_resumen_excel",
                            )
                    except Exception:
                        pass
            except Exception as e:
                st.caption(f"Error al generar resumen: {e}")

        # ---------- Respaldo de hoy (solo maestro/admin) ----------
        es_maestro_contab = (st.session_state.get("es_acceso_maestro") or ES_PC_MAESTRO) and usuario_actual == "admin"
        if es_maestro_contab and not df.empty:
            with st.expander(f"💾 {t.get('respaldo_hoy', 'Descargar respaldo de hoy')}", expanded=False):
                st.caption(t.get("respaldo_hoy_ayuda", ""))
                hoy_nombre = datetime.now().strftime("%Y-%m-%d")
                buf_resp = BytesIO()
                df.to_csv(buf_resp, index=False, encoding="utf-8-sig")
                buf_resp.seek(0)
                st.download_button(
                    label=f"CSV respaldo {hoy_nombre}",
                    data=buf_resp.getvalue(),
                    file_name=f"respaldo_contable_{hoy_nombre}.csv",
                    mime="text/csv",
                    key="btn_respaldo_hoy_csv",
                )

        # ---------- Alertas (saldo bajo, gastos > ingresos, sin movimientos) ----------
        with st.expander(f"⚠️ {t.get('alertas_contabilidad', 'Alertas')}", expanded=False):
            alertas_list = []
            if saldo_act < 0:
                alertas_list.append(t["alerta_saldo_negativo"])
            if ing_mes > 0 and gas_mes >= ing_mes:
                alertas_list.append(t["alerta_gastos_altos"])
            if saldo_act >= 0 and (saldo_act < 500 or (gas_mes > 0 and saldo_act < gas_mes)):
                alertas_list.append(t["alerta_saldo_bajo"])
            try:
                df_ult = df.copy()
                df_ult["_f"] = pd.to_datetime(df_ult["fecha"].astype(str).str[:10], errors="coerce")
                df_ult = df_ult.dropna(subset=["_f"]).sort_values("_f")
                if not df_ult.empty:
                    ultima_fecha = df_ult["_f"].iloc[-1].date()
                    dias_sin = (datetime.now().date() - ultima_fecha).days
                    if dias_sin > 7:
                        alertas_list.append(t.get("alerta_sin_movimientos_dias", "Sin movimientos recientes.").format(d=dias_sin))
            except Exception:
                pass
            if not alertas_list:
                st.success(t["sin_alertas"])
            else:
                for a in alertas_list:
                    st.warning(a)

        # Gráfico trazabilidad: saldo en el tiempo (línea continua, verde=alza, rojo=baja) — estilo bolsa
        try:
            df_dash = df_dash.copy()
            df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str), errors="coerce")
            df_dash = df_dash.dropna(subset=["_fecha"]).sort_values("_fecha").reset_index(drop=True)
            if not df_dash.empty and "saldo_actual" in df_dash.columns and _PLOTLY_DISPONIBLE:
                saldos = pd.to_numeric(df_dash["saldo_actual"], errors="coerce").fillna(0).tolist()
                fechas = df_dash["_fecha"].tolist()
                tema_oscuro = st.session_state.get("tema_app", "oscuro") == "oscuro"
                # Estilo bolsa: fondo azul oscuro, rejilla neon, porcentajes celeste/turquesa
                if tema_oscuro:
                    paper_bg = "rgba(5,15,35,0.98)"
                    plot_bg = "rgba(8,25,55,0.95)"
                    grid_color = "rgba(0,200,255,0.4)"
                    grid_color_sec = "rgba(0,180,230,0.25)"
                    font_color = "rgba(150,220,255,0.95)"
                    pct_color = "rgba(100,210,255,0.9)"
                else:
                    paper_bg = "rgba(240,248,255,0.98)"
                    plot_bg = "rgba(230,245,255,0.95)"
                    grid_color = "rgba(0,150,200,0.35)"
                    grid_color_sec = "rgba(0,120,180,0.2)"
                    font_color = "#1a365d"
                    pct_color = "rgba(0,120,180,0.9)"
                # Porcentaje de variación (referencia: saldo inicial)
                ref = saldos[0] if saldos else 0
                pct_inicial = 0.0
                pct_final = ((saldos[-1] - ref) / abs(ref) * 100) if ref != 0 else 0.0
                pct_max = ((max(saldos) - ref) / abs(ref) * 100) if ref != 0 else 0.0
                pct_min = ((min(saldos) - ref) / abs(ref) * 100) if ref != 0 else 0.0
                # Segmentos: verde=alza, rojo=baja (línea continua)
                fig = go.Figure()
                if len(saldos) == 1:
                    fig.add_trace(go.Scatter(
                        x=fechas, y=saldos, mode="markers",
                        marker=dict(size=14, color="#00d4ff", line=dict(width=2, color="#00a8cc"), symbol="diamond"),
                        name=t["grafico_saldo"],
                    ))
                else:
                    for i in range(len(saldos) - 1):
                        color_seg = "#00e676" if saldos[i + 1] >= saldos[i] else "#ff5252"
                        fig.add_trace(go.Scatter(
                            x=[fechas[i], fechas[i + 1]], y=[saldos[i], saldos[i + 1]],
                            mode="lines",
                            line=dict(color=color_seg, width=3.5),
                            showlegend=False,
                            hovertemplate="%{x|%Y-%m-%d %H:%M:%S}<br>$%{y:,.2f}<extra></extra>",
                        ))
                    cmin, cmax = min(saldos), max(saldos)
                    if cmin == cmax:
                        cmin, cmax = cmin - 1, cmax + 1
                    fig.add_trace(go.Scatter(
                        x=fechas, y=saldos, mode="markers",
                        marker=dict(size=7, color=saldos, colorscale=[[0, "#ff5252"], [0.5, "#00d4ff"], [1, "#00e676"]],
                                    cmin=cmin, cmax=cmax, showscale=False, line=dict(width=1.5, color="rgba(255,255,255,0.7)")),
                        name=t["grafico_saldo"],
                        hovertemplate="%{x|%Y-%m-%d %H:%M:%S}<br>$%{y:,.2f}<extra></extra>",
                    ))
                # Anotaciones de porcentaje (celeste/turquesa, significado claro)
                annotations = []
                if len(saldos) > 1:
                    annotations.append(dict(x=fechas[0], y=saldos[0], text=f"{pct_inicial:+.1f}%", showarrow=False,
                                           xanchor="right", font=dict(size=12, color=pct_color, family="Calibri"),
                                           bgcolor="rgba(0,0,0,0.3)", borderpad=4))
                    annotations.append(dict(x=fechas[-1], y=saldos[-1], text=f"{pct_final:+.1f}%", showarrow=False,
                                           xanchor="left", font=dict(size=12, color=pct_color, family="Calibri"),
                                           bgcolor="rgba(0,0,0,0.3)", borderpad=4))
                    if len(saldos) > 2 and (pct_max != pct_final or pct_min != pct_final):
                        idx_max = saldos.index(max(saldos))
                        idx_min = saldos.index(min(saldos))
                        if idx_max not in (0, len(saldos) - 1):
                            annotations.append(dict(x=fechas[idx_max], y=saldos[idx_max], text=f"{pct_max:+.1f}%", showarrow=False,
                                                   font=dict(size=11, color=pct_color), bgcolor="rgba(0,0,0,0.25)", borderpad=3))
                        if idx_min not in (0, len(saldos) - 1) and idx_min != idx_max:
                            annotations.append(dict(x=fechas[idx_min], y=saldos[idx_min], text=f"{pct_min:+.1f}%", showarrow=False,
                                                   font=dict(size=11, color=pct_color), bgcolor="rgba(0,0,0,0.25)", borderpad=3))
                fig.update_layout(
                    title=dict(text=f"📈 {t['grafico_trazabilidad']} — {t['grafico_var']} %", font=dict(size=18, color=font_color)),
                    paper_bgcolor=paper_bg,
                    plot_bgcolor=plot_bg,
                    font=dict(family="Calibri, 'Segoe UI', sans-serif", color=font_color, size=12),
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5, bgcolor="rgba(0,0,0,0)", font=dict(size=11, color=font_color)),
                    margin=dict(t=80, b=80, l=70, r=50),
                    hovermode="x unified",
                    annotations=annotations,
                    xaxis=dict(
                        showgrid=True, gridcolor=grid_color, gridwidth=1.5, zeroline=True, zerolinecolor=grid_color_sec,
                        tickfont=dict(size=11, color=font_color), tickformat="%d/%m %H:%M",
                        rangeslider=dict(visible=True, bgcolor="rgba(0,100,150,0.2)", bordercolor=grid_color, thickness=0.05),
                        type="date",
                    ),
                    yaxis=dict(
                        title=dict(text="$", font=dict(color=font_color)),
                        showgrid=True, gridcolor=grid_color, gridwidth=1.5, zeroline=True, zerolinecolor=grid_color_sec,
                        tickfont=dict(size=11, color=font_color), tickformat="$,.0f",
                        dtick=None,
                    ),
                )
                fig.update_xaxes(showgrid=True, gridwidth=1.5, gridcolor=grid_color, zerolinewidth=1)
                fig.update_yaxes(showgrid=True, gridwidth=1.5, gridcolor=grid_color, zerolinewidth=1)
                st.plotly_chart(fig, width="stretch", config={"displayModeBar": True, "displaylogo": False, "scrollZoom": True})
                st.caption(f"🟢 {t['grafico_alza']}  ·  🔴 {t['grafico_baja']} — {t['grafico_trazabilidad']}")

                # Gráfico gastos por tipo en el tiempo (barras apiladas por mes)
                if "tipo_gasto" in df_dash.columns and _PLOTLY_DISPONIBLE:
                    try:
                        df_gt = df_dash.copy()
                        df_gt["_mes"] = df_gt["_fecha"].dt.to_period("M").astype(str)
                        df_gt["gastos_num"] = pd.to_numeric(df_gt["gastos"], errors="coerce").fillna(0)
                        df_gt = df_gt[df_gt["gastos_num"] > 0]
                        if not df_gt.empty:
                            pivot_gt = df_gt.pivot_table(index="_mes", columns="tipo_gasto", values="gastos_num", aggfunc="sum", fill_value=0)
                            pivot_gt = pivot_gt.reindex(sorted(pivot_gt.columns), axis=1) if len(pivot_gt.columns) > 1 else pivot_gt
                            fig_gt = go.Figure()
                            for col in pivot_gt.columns:
                                fig_gt.add_trace(go.Bar(name=str(col) or "(sin tipo)", x=pivot_gt.index, y=pivot_gt[col].tolist()))
                            fig_gt.update_layout(
                                barmode="stack",
                                title=dict(text=f"📊 {t.get('grafico_por_tipo_gasto', 'Gastos por tipo en el tiempo')}", font=dict(size=14, color=font_color)),
                                paper_bgcolor=paper_bg, plot_bgcolor=plot_bg,
                                font=dict(color=font_color, size=11),
                                xaxis=dict(tickfont=dict(color=font_color)),
                                yaxis=dict(title="$", tickfont=dict(color=font_color), gridcolor=grid_color),
                                legend=dict(orientation="h", yanchor="bottom", font=dict(color=font_color)),
                            )
                            st.plotly_chart(fig_gt, width="stretch", config={"displayModeBar": True, "displaylogo": False})
                    except Exception:
                        pass

                # Cuadro Ingresos & Gastos: solo visible con contraseña universal + clic
                es_maestro = st.session_state.get("es_acceso_maestro") and usuario_actual == "admin"
                if es_maestro:
                    ver_ing_gas = st.checkbox(f"📊 {t['grafico_ver_ingresos_gastos']}", key="maestro_ver_grafico_ingresos_gastos", value=False)
                    if ver_ing_gas:
                        df_dash["_mes"] = df_dash["_fecha"].dt.to_period("M").astype(str)
                        ultimos_meses = sorted(df_dash["_mes"].dropna().unique())[-6:]
                        chart_data = []
                        for m in ultimos_meses:
                            sub = df_dash[df_dash["_mes"] == m]
                            chart_data.append({"mes": m, "Ingresos": float(sub["ingreso"].sum()), "Gastos": float(sub["gastos"].sum())})
                        if chart_data:
                            df_chart = pd.DataFrame(chart_data)
                            fig2 = go.Figure()
                            fig2.add_trace(go.Bar(name=t["ingresos_mes"], x=df_chart["mes"], y=df_chart["Ingresos"],
                                                  marker_color="rgba(0,230,118,0.65)", marker_line_color="rgba(0,200,100,0.5)", marker_line_width=0.5, width=0.25))
                            fig2.add_trace(go.Bar(name=t["gastos_mes"], x=df_chart["mes"], y=df_chart["Gastos"],
                                                  marker_color="rgba(255,82,82,0.65)", marker_line_color="rgba(220,50,50,0.5)", marker_line_width=0.5, width=0.25))
                            fig2.update_layout(barmode="group", bargap=0.5, bargroupgap=0.15,
                                              title=dict(text=f"📊 {t['grafico_ingresos_gastos']}", font=dict(size=14, color=font_color)),
                                              paper_bgcolor=paper_bg, plot_bgcolor=plot_bg, font=dict(color=font_color),
                                              margin=dict(t=50, b=40), xaxis=dict(gridcolor=grid_color, gridwidth=1.5),
                                              yaxis=dict(gridcolor=grid_color, gridwidth=1.5, tickformat="$,.0f"))
                            st.plotly_chart(fig2, width="stretch", config={"displayModeBar": False})
            else:
                es_maestro_else = st.session_state.get("es_acceso_maestro") and usuario_actual == "admin"
                if es_maestro_else:
                    ver_ing_gas_else = st.checkbox(f"📊 {t['grafico_ver_ingresos_gastos']}", key="maestro_ver_grafico_ingresos_gastos", value=False)
                    if ver_ing_gas_else:
                        df_dash["_fecha"] = pd.to_datetime(df_dash["fecha"].astype(str).str[:10], errors="coerce")
                        df_dash["_mes"] = df_dash["_fecha"].dt.to_period("M").astype(str)
                        ultimos_meses = sorted(df_dash["_mes"].dropna().unique())[-6:]
                        chart_data = [{"mes": m, "Ingresos": float(df_dash[df_dash["_mes"] == m]["ingreso"].sum()), "Gastos": float(df_dash[df_dash["_mes"] == m]["gastos"].sum())} for m in ultimos_meses]
                        if chart_data:
                            df_chart = pd.DataFrame(chart_data).set_index("mes")
                            st.bar_chart(df_chart[["Ingresos", "Gastos"]], width="stretch")
        except Exception:
            pass
        with st.expander(f"📋 {t['conciliar']}", expanded=False):
            st.caption(t["conciliar_ayuda"])
            st.metric(t["ingresos_mes"] + " (registrado)", f"${ing_mes:,.2f}")
            lo_contado = st.number_input(t["lo_contado_caja"], min_value=0.0, value=0.0, step=10.0, key="conciliar_contado")
            if lo_contado > 0:
                diff = lo_contado - ing_mes
                if abs(diff) < 0.02:
                    st.success(t["coincide_registrado"])
                else:
                    st.caption(f"Diferencia: ${diff:+,.2f}")
        st.markdown("---")

    # ---------- BUSCAR FACTURAS (POR COMERCIO, FECHA, TOTAL, TEXTO OCR) ----------
    if tiene_permiso(usuario_actual, "ver_hoja_contable"):
        with st.expander(f"🔍 {t['buscar_facturas_titulo']}", expanded=False):
            st.caption(t["buscar_facturas_ayuda"])
            data_f = cargar_facturas()
            facturas = data_f.get("facturas", [])
            if not facturas:
                st.info(t["sin_facturas"])
            else:
                col_b1, col_b2, col_b3 = st.columns(3)
                with col_b1:
                    filtro_comercio = st.text_input(t["buscar_por_comercio"], key="buscar_comercio", placeholder="")
                    filtro_texto_ocr = st.text_input(t["buscar_por_texto_ocr"], key="buscar_ocr", placeholder="")
                with col_b2:
                    filtro_fecha_desde = st.date_input(t["buscar_por_fecha_desde"], value=None, key="buscar_fecha_desde")
                    filtro_fecha_hasta = st.date_input(t["buscar_por_fecha_hasta"], value=None, key="buscar_fecha_hasta")
                with col_b3:
                    filtro_total_min = st.number_input(t["buscar_por_total_min"], min_value=0.0, value=0.0, key="buscar_total_min", step=10.0)
                    filtro_total_max = st.number_input(t["buscar_por_total_max"], min_value=0.0, value=0.0, key="buscar_total_max", step=50.0)
                tipos_en_facturas = sorted(set(str(fa.get("tipo_gasto") or "").strip() for fa in facturas if (fa.get("tipo_gasto") or "").strip()))
                opciones_tipo_f = [t["todos_los_tipos"]] + tipos_en_facturas
                filtro_tipo_gasto = st.selectbox(t["buscar_por_tipo_gasto"], options=opciones_tipo_f, key="buscar_tipo_gasto")
                col_imp1, col_imp2, _ = st.columns(3)
                with col_imp1:
                    filtro_impuesto_min = st.number_input(t["buscar_por_impuesto_desde"], min_value=0.0, value=0.0, key="buscar_impuesto_min", step=5.0)
                with col_imp2:
                    filtro_impuesto_max = st.number_input(t["buscar_por_impuesto_hasta"], min_value=0.0, value=0.0, key="buscar_impuesto_max", step=10.0)
                filtered = []
                for fa in facturas:
                    comercio = (fa.get("comercio") or "").upper()
                    ocr = (fa.get("ocr_text") or "").upper()
                    try:
                        total = float(fa.get("total") or 0)
                    except (TypeError, ValueError):
                        total = 0.0
                    try:
                        impuesto = float(fa.get("impuesto") or 0)
                    except (TypeError, ValueError):
                        impuesto = 0.0
                    fecha_reg = (fa.get("fecha_registro") or "")[:10]
                    try:
                        dt_reg = datetime.strptime(fecha_reg, "%Y-%m-%d").date() if len(fecha_reg) >= 10 else None
                    except ValueError:
                        dt_reg = None
                    tipo_fa = (fa.get("tipo_gasto") or "").strip()
                    if filtro_comercio and filtro_comercio.upper() not in comercio:
                        continue
                    if filtro_texto_ocr and filtro_texto_ocr.upper() not in ocr:
                        continue
                    if filtro_total_min and filtro_total_min > 0 and total < filtro_total_min:
                        continue
                    if filtro_total_max and filtro_total_max > 0 and total > filtro_total_max:
                        continue
                    if filtro_fecha_desde and (dt_reg is None or dt_reg < filtro_fecha_desde):
                        continue
                    if filtro_fecha_hasta and (dt_reg is None or dt_reg > filtro_fecha_hasta):
                        continue
                    if filtro_tipo_gasto and filtro_tipo_gasto != t["todos_los_tipos"] and tipo_fa != filtro_tipo_gasto:
                        continue
                    if filtro_impuesto_min and filtro_impuesto_min > 0 and impuesto < filtro_impuesto_min:
                        continue
                    if filtro_impuesto_max and filtro_impuesto_max > 0 and impuesto > filtro_impuesto_max:
                        continue
                    filtered.append(fa)
                if not filtered:
                    st.warning(t["sin_resultados_facturas"])
                else:
                    st.markdown(f"**{t['resultados_facturas']}:** {len(filtered)}")
                    for fa in filtered[:50]:
                        with st.expander(f"{fa.get('id_registro', '')} — {(fa.get('comercio') or '')[:40]} — ${(fa.get('total') or 0):.2f}"):
                            st.caption(f"**{t['col_fecha']}:** {fa.get('fecha_registro', '')}")
                            if fa.get("tipo_gasto"):
                                st.caption(f"**{t['tipo_gasto']}:** {fa.get('tipo_gasto', '')}")
                            if fa.get("impuesto") is not None:
                                st.caption(f"**{t['impuesto_detectado']}:** ${fa['impuesto']:.2f}")
                            if fa.get("ocr_text"):
                                st.text_area(t["texto_ocr"], value=fa.get("ocr_text", "")[:3000], height=120, disabled=True, key=f"ocr_{fa.get('id_registro','')}_{hash(fa.get('ocr_text','')) % 100000}")
            st.markdown("---")
            zip_bytes, zip_nombre = _crear_zip_fotos_gastos()
            if zip_bytes:
                st.download_button(label=t["descargar_fotos_zip"], data=zip_bytes, file_name=zip_nombre, mime="application/zip", key="btn_zip_fotos")

    # ---------- HOJA CONTABLE (LISTADO, FILTROS, BORRADO Y LIMPIEZA) ----------
    if tiene_permiso(usuario_actual, "ver_hoja_contable"):
        st.markdown(f"**{t['borrar_solo_30min']}**")
        df = cargar_db()
        if df.empty:
            st.info(t["sin_movimientos"])
        else:
            if "fecha" in df.columns:
                df = df.sort_values("fecha", ascending=True).reset_index(drop=True)
                df = _recalcular_totales_ledger(df)
            df_show = df.copy()
            ok_int_hoja, _ = verificar_integridad_ledger(df)
            df_show["puede_borrar"] = df_show["fecha"].apply(puede_borrar)
            meta_borrar = cargar_arqueo_meta()
            fechas_cerradas = set(meta_borrar.get("_fechas_cerradas", []) or [])
            df_show["fecha_cerrada"] = df_show["fecha"].astype(str).str[:10].isin(fechas_cerradas)
            # --- Filtros ---
            st.markdown(f"**{t['filtros']}**")
            col_f1, col_f2, col_f3, col_f4, col_f5 = st.columns(5)
            with col_f1:
                fecha_desde = st.date_input(
                    t["fecha_desde"], value=None, key="filtro_fecha_desde", help=t["help_fecha_desde"]
                )
            with col_f2:
                fecha_hasta = st.date_input(
                    t["fecha_hasta"], value=None, key="filtro_fecha_hasta", help=t["help_fecha_hasta"]
                )
            with col_f3:
                monto_min = st.number_input(
                    t["monto_min"], min_value=0.0, value=0.0, step=10.0, key="filtro_monto_min",
                    help=t["help_monto_min"]
                )
            with col_f4:
                monto_max = st.number_input(
                    t["monto_max"], min_value=0.0, value=0.0, step=50.0, key="filtro_monto_max",
                    help=t["help_monto_max"]
                )
            with col_f5:
                tipos_unicos = sorted(set(
                    str(x).strip() for x in df_show["tipo_gasto"].dropna().unique()
                    if str(x).strip() and str(x).lower() != "nan"
                ))
                opciones_tipo = [t["todos_los_tipos"]] + tipos_unicos
                filtrar_tipo_sel = st.selectbox(
                    t["filtrar_tipo"], options=opciones_tipo, key="filtro_tipo", help=t["help_filtrar_tipo"]
                )
            col_f6, col_f7, col_f8 = st.columns(3)
            with col_f6:
                tipo_mov_sel = st.radio(
                    t.get("filtro_tipo_movimiento", "Tipo de movimiento"),
                    options=["todos", "ingresos", "gastos"],
                    format_func=lambda x: t.get("todos_movimientos", "Todos") if x == "todos" else (t.get("solo_ingresos", "Solo ingresos") if x == "ingresos" else t.get("solo_gastos", "Solo gastos")),
                    key="filtro_tipo_mov_hoja",
                    horizontal=True,
                )
            with col_f7:
                orden_sel = st.radio(
                    t.get("orden_hoja", "Orden"),
                    options=["recientes", "antiguos"],
                    format_func=lambda x: t.get("orden_recientes", "Más recientes primero") if x == "recientes" else t.get("orden_antiguos", "Más antiguos primero"),
                    key="filtro_orden_hoja",
                    horizontal=True,
                )
            with col_f8:
                buscar_detalle = st.text_input(
                    t.get("buscar_en_detalle", "Buscar en detalle"),
                    placeholder=t.get("buscar_en_detalle_placeholder", ""),
                    key="filtro_buscar_detalle_hoja",
                )
            st.caption(t["filtros_ayuda"])
            # Aplicar máscara de filtros
            mask = pd.Series(True, index=df_show.index)
            if "fecha" in df_show.columns:
                try:
                    def _parse_fecha(s):
                        s = (s or "").strip()[:10]
                        if not s or len(s) < 10 or not s[:4].isdigit():
                            return None
                        try:
                            return datetime.strptime(s, "%Y-%m-%d").date()
                        except ValueError:
                            return None
                    fechas_parsed = df_show["fecha"].astype(str).str[:10].apply(_parse_fecha)
                    if fecha_desde is not None:
                        mask &= (fechas_parsed >= fecha_desde).fillna(False)
                    if fecha_hasta is not None:
                        mask &= (fechas_parsed <= fecha_hasta).fillna(False)
                except Exception:
                    pass
            ing_num = pd.to_numeric(df_show["ingreso"], errors="coerce").fillna(0)
            gas_num = pd.to_numeric(df_show["gastos"], errors="coerce").fillna(0)
            monto_fila = ing_num + gas_num
            if monto_min is not None and monto_min > 0:
                mask &= (monto_fila >= monto_min)
            if monto_max is not None and monto_max > 0:
                mask &= (monto_fila <= monto_max)
            if filtrar_tipo_sel and filtrar_tipo_sel != t["todos_los_tipos"]:
                tipo_norm = df_show["tipo_gasto"].fillna("").astype(str).str.strip()
                mask &= (tipo_norm == filtrar_tipo_sel)
            if tipo_mov_sel == "ingresos":
                mask &= (ing_num > 0)
            elif tipo_mov_sel == "gastos":
                mask &= (gas_num > 0)
            texto_buscar = str(buscar_detalle).strip() if buscar_detalle else ""
            if texto_buscar:
                detalle_str = df_show["detalle"].fillna("").astype(str).str.lower()
                mask &= detalle_str.str.contains(re.escape(texto_buscar.lower()), na=False)
            filtered = df_show[mask]
            if "fecha" in filtered.columns:
                filtered = filtered.sort_values("fecha", ascending=(orden_sel == "antiguos")).reset_index(drop=True)
            if filtered.empty:
                st.info(t["sin_resultados_filtro"])
            else:
                columnas_ver = [c for c in COLUMNAS_LEDGER if c in filtered.columns]
                display_df = filtered[columnas_ver].copy()
                es_maestro = (st.session_state.get("es_acceso_maestro") or ES_PC_MAESTRO) and usuario_actual == "admin"
                if es_maestro:
                    meta = cargar_arqueo_meta()
                    display_df["ip_dispositivo"] = display_df["id_registro"].apply(
                        lambda rid: meta.get(str(rid), {}).get("ip_dispositivo", "—")
                    )
                    columnas_ver = list(display_df.columns)
                nombres_columnas = {
                    "id_registro": t["col_id_registro"],
                    "fecha": t["col_fecha"],
                    "detalle": t["col_detalle"],
                    "tipo_gasto": t["col_tipo_gasto"],
                    "ingreso": t["col_ingreso"],
                    "gastos": t["col_gastos"],
                    "total_ingresos": t["col_total_ingresos"],
                    "total_gastos": t["col_total_gastos"],
                    "saldo_actual": t["col_saldo_actual"],
                    "ip_dispositivo": t["col_ip"],
                }
                def _fmt_monto_celda(x):
                    try:
                        if pd.isna(x) or str(x).strip() == "":
                            return ""
                        return f"${float(x):.2f}"
                    except (TypeError, ValueError):
                        return ""
                for col in ["ingreso", "gastos", "total_ingresos", "total_gastos", "saldo_actual"]:
                    if col in display_df.columns:
                        display_df[col] = display_df[col].apply(_fmt_monto_celda)
                display_df = display_df.rename(columns=nombres_columnas)
                # Paginación
                total = len(display_df)
                n_pag = max(1, (total + REGISTROS_POR_PAGINA - 1) // REGISTROS_POR_PAGINA)
                pag_actual = st.session_state.get("pagina_hoja", 0)
                pag_actual = min(max(0, pag_actual), n_pag - 1)
                inicio = pag_actual * REGISTROS_POR_PAGINA
                fin = min(inicio + REGISTROS_POR_PAGINA, total)
                display_pag = display_df.iloc[inicio:fin]
                st.dataframe(display_pag, width="stretch", hide_index=True)
                if n_pag > 1:
                    col_prev, col_info, col_next = st.columns([1, 2, 1])
                    with col_prev:
                        if st.button(t["anterior"], key="btn_prev_pag", disabled=(pag_actual == 0)):
                            st.session_state["pagina_hoja"] = pag_actual - 1
                            st.rerun()
                    with col_info:
                        st.caption(f"{t['pagina']} {pag_actual + 1} {t['de']} {n_pag} ({total} {t['registros']})")
                    with col_next:
                        if st.button(t["siguiente"], key="btn_next_pag", disabled=(pag_actual >= n_pag - 1)):
                            st.session_state["pagina_hoja"] = pag_actual + 1
                            st.rerun()
                es_maestro = st.session_state.get("es_acceso_maestro") and usuario_actual == "admin"
                if (tiene_permiso(usuario_actual, "ver_eliminar_registros") or es_maestro) and ok_int_hoja:
                    titulo_limpiar = t["limpiar_fila"] + (f" ({t['maestro_sin_limite']})" if es_maestro else f" ({t['solo_30min']})")
                    st.markdown(f"**{titulo_limpiar}:**")
                    modo_borrar = st.session_state.get("modo_borrar_registros", False)
                    if st.button(f"🗑️ {t['borrar_btn']}", key="btn_activar_borrar"):
                        st.session_state["modo_borrar_registros"] = not modo_borrar
                        st.rerun()
                    if modo_borrar:
                        st.caption(t["maestro_seleccionar_masa"])
                        indices_pagina = [idx for idx in display_pag.index if (filtered.loc[idx, "puede_borrar"] or es_maestro) and not filtered.loc[idx, "fecha_cerrada"]]
                        if st.button(t["seleccionar_todos"], key="btn_seleccionar_todos"):
                            for idx in indices_pagina:
                                st.session_state[f"borrar_cb_{idx}"] = True
                            st.rerun()
                        st.markdown("""<style>
                        div[data-testid='stCheckbox'] { transform: scale(0.88); }
                        div[data-testid='stCheckbox'] input:checked + label { color: #22c55e !important; }
                        </style>""", unsafe_allow_html=True)
                        n_cols = 4
                        for j in range(0, len(indices_pagina), n_cols):
                            cols = st.columns(n_cols)
                            for k in range(n_cols):
                                idx_pos = j + k
                                if idx_pos < len(indices_pagina):
                                    idx = indices_pagina[idx_pos]
                                    row = filtered.loc[idx]
                                    id_reg = row.get("id_registro", idx)
                                    det = (row.get("detalle") or "")[:15]
                                    fecha = str(row.get("fecha", ""))[:10]
                                    with cols[k]:
                                        st.checkbox(f"✓ {id_reg}", key=f"borrar_cb_{idx}")
                                        st.caption(f"{fecha} {det}", help=f"{id_reg}")
                        col_sel, col_tod = st.columns(2)
                        with col_sel:
                            indices_sel = [idx for idx in filtered.index if (filtered.loc[idx, "puede_borrar"] or es_maestro) and not filtered.loc[idx, "fecha_cerrada"] and st.session_state.get(f"borrar_cb_{idx}", False)]
                            n_sel = len(indices_sel)
                            if st.button(f"🗑️ {t['borrar_seleccionados']} ({n_sel})", key="btn_borrar_sel", disabled=(n_sel == 0)):
                                df = df.drop(index=indices_sel).reset_index(drop=True)
                                for idx in indices_sel:
                                    st.session_state.pop(f"borrar_cb_{idx}", None)
                                if guardar_db(df, t):
                                    audit_log(usuario_actual, "registro_eliminado_masa", str(n_sel))
                                    st.session_state["modo_borrar_registros"] = False
                                    st.rerun()
                        with col_tod:
                            if st.button(t["borrar_todos"], key="btn_borrar_todos"):
                                st.session_state["confirmar_borrar_todos"] = True
                                st.rerun()
                        if st.session_state.get("confirmar_borrar_todos"):
                            st.warning(t["confirmar_limpiar_todo"])
                            conf = st.text_input("", key="input_confirmar_borrar_todos", placeholder=CONFIRMACION_LIMPIAR_TODO)
                            if st.button(t["maestro_borrar_todo"], key="btn_confirmar_borrar_todos"):
                                if (conf or "").strip().upper() == CONFIRMACION_LIMPIAR_TODO:
                                    idx_a_borrar = filtered.index[~filtered["fecha_cerrada"]]
                                    df = df.drop(index=idx_a_borrar).reset_index(drop=True)
                                    if guardar_db(df, t):
                                        audit_log(usuario_actual, "registro_eliminado_masa", str(len(idx_a_borrar)))
                                        st.session_state.pop("confirmar_borrar_todos", None)
                                        st.session_state["modo_borrar_registros"] = False
                                        for idx in filtered.index:
                                            st.session_state.pop(f"borrar_cb_{idx}", None)
                                        st.rerun()
                                else:
                                    st.warning(t["confirmar_limpiar_todo"])
                            if st.button("✕ Cancelar", key="btn_cancelar_borrar_todos"):
                                st.session_state.pop("confirmar_borrar_todos", None)
                                st.rerun()
            if st.session_state.get("es_acceso_maestro") and usuario_actual == "admin":
                if not ok_int_hoja:
                    st.warning(t.get("integridad_aviso", "") + " " + (t.get("integridad_error_titulo", "") or ""))
                    st.caption("El borrado de registros está deshabilitado hasta corregir la integridad del libro.")
                elif ok_int_hoja:
                    st.markdown("---")
                    st.markdown(f"**{t['maestro_borrar_titulo']}**")
                with st.expander(t["maestro_borrar_todo"], expanded=False):
                    st.caption(t["confirmar_limpiar_todo"])
                    confirmar_limpiar = st.text_input("", key="input_limpiar_todo", placeholder=CONFIRMACION_LIMPIAR_TODO)
                    if st.button(t["maestro_borrar_todo"], key="btn_limpiar_todo"):
                        if (confirmar_limpiar or "").strip().upper() == CONFIRMACION_LIMPIAR_TODO:
                            if reiniciar_tesoreria_master():
                                audit_log(usuario_actual, "limpiar_tablero_completo", "")
                                st.success(t["reinicio_ok"])
                                st.rerun()
                            else:
                                st.error(t["error_limpiar_tablero"])
                        else:
                            st.warning(t["confirmar_limpiar_todo"])
                with st.expander(t["maestro_borrar_detalle"], expanded=False):
                    st.caption(t["maestro_seleccionar_detalle"])
                    texto_detalle = st.text_input("", key="input_maestro_detalle", placeholder=t.get("descripcion", "Descripción"))
                    if texto_detalle and texto_detalle.strip():
                        col_detalle = "detalle" if "detalle" in df.columns else "descripcion"
                        mask_det = df[col_detalle].astype(str).str.upper().str.contains(texto_detalle.strip().upper(), na=False)
                        n_coincidencias = mask_det.sum()
                        if n_coincidencias > 0:
                            st.caption(t["maestro_coinciden_registros"].format(n=n_coincidencias))
                            if st.button(f"🗑️ {t['maestro_borrar_detalle']} ({n_coincidencias})", key="btn_borrar_detalle"):
                                df = df[~mask_det].reset_index(drop=True)
                                if guardar_db(df, t):
                                    audit_log(usuario_actual, "registro_eliminado_detalle", texto_detalle.strip())
                                    st.rerun()
                        else:
                            st.caption(t["sin_resultados_filtro"])
            elif ES_PC_MAESTRO:
                st.markdown("---")
                st.markdown(f"**{t['limpiar_todo_tablero']}**")
                st.caption(t["confirmar_limpiar_todo"])
                confirmar_limpiar = st.text_input("", key="input_limpiar_todo", placeholder=CONFIRMACION_LIMPIAR_TODO)
                if st.button(t["limpiar_todo_tablero"], key="btn_limpiar_todo"):
                    if (confirmar_limpiar or "").strip().upper() == CONFIRMACION_LIMPIAR_TODO:
                        if reiniciar_tesoreria_master():
                            audit_log(usuario_actual, "limpiar_tablero_completo", "")
                            st.success(t["reinicio_ok"])
                            st.rerun()
                        else:
                            st.error(t["error_limpiar_tablero"])
                    else:
                        st.warning(t["confirmar_limpiar_todo"])

    # ---------- EXPORTAR HOJA CONTABLE (PDF, EXCEL, PARA CONTADOR) ----------
    st.markdown("---")
    puede_exportar = tiene_permiso(usuario_actual, "ver_exportar_hoja_pdf") or tiene_permiso(usuario_actual, "ver_hoja_contable")
    if puede_exportar:
        with st.expander(f"📤 {t['exportar_hoja_contable_titulo']}", expanded=True):
            st.caption(t["exportar_hoja_contable_ayuda"])
            df_exp = cargar_db()
            if df_exp.empty:
                st.warning(t["sin_movimientos"])
            else:
                if "fecha" in df_exp.columns:
                    df_exp = df_exp.sort_values("fecha", ascending=True).reset_index(drop=True)
                    df_exp = _recalcular_totales_ledger(df_exp)
                col_pdf, col_xlsx, col_csv = st.columns(3)
                with col_pdf:
                    st.markdown(f"**{t['exportar_opcion_pdf']}**")
                    pdf_hoja = generar_pdf_hoja_contable(df_exp, lang)
                    pdf_hoja.seek(0)
                    st.download_button(
                        label=t["exportar_opcion_pdf"],
                        data=pdf_hoja.getvalue(),
                        file_name=f"Hoja_Contable_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        mime="application/pdf",
                        key="download_hoja_pdf"
                    )
                with col_xlsx:
                    st.markdown(f"**{t['exportar_opcion_excel']}**")
                    try:
                        buf_xlsx = BytesIO()
                        df_exp.to_excel(buf_xlsx, index=False, engine="openpyxl")
                        xlsx_bytes = _formatear_excel_contador(buf_xlsx)
                        st.download_button(
                            label=t["exportar_opcion_excel"],
                            data=xlsx_bytes,
                            file_name=f"Hoja_Contable_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_excel"
                        )
                    except Exception:
                        st.caption("Excel no disponible" if lang == "ES" else "Excel not available")
                with col_csv:
                    st.markdown(f"**{t['exportar_opcion_contador']}**")
                    csv_bytes = "\ufeff" + df_exp.to_csv(index=False, encoding="utf-8", date_format="%Y-%m-%d")
                    st.download_button(
                        label=t["exportar_opcion_contador"],
                        data=csv_bytes.encode("utf-8"),
                        file_name=f"Hoja_Contable_contador_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                        mime="text/csv",
                        key="download_contador"
                    )

    # ---------- INFORME PDF PARA WHATSAPP ----------
    if tiene_permiso(usuario_actual, "ver_informe_pdf"):
        if st.button(f"📄 {t['ver_informe']}", key="btn_pdf"):
            df = cargar_db()
            if df.empty:
                total_ingresos = 0.0
                total_gastos = 0.0
                saldo = 0.0
                ultimo_arqueo = ""
            else:
                total_ingresos = float(df["ingreso"].sum(skipna=True) or 0)
                total_gastos = float(df["gastos"].sum(skipna=True) or 0)
                saldo = total_ingresos - total_gastos
                ing_num = pd.to_numeric(df["ingreso"], errors="coerce").fillna(0)
                ultimo = df.loc[ing_num > 0].tail(1)
                if not ultimo.empty:
                    val = _safe_float_pdf(ultimo["ingreso"].values[0])
                    ultimo_arqueo = f"${val:.2f}"
                else:
                    ultimo_arqueo = "$0.00"
            pdf_bytes = generar_pdf(df, lang, saldo, ultimo_arqueo)
            pdf_bytes.seek(0)
            st.download_button(
                label=t["descargar_pdf_whatsapp"],
                data=pdf_bytes.getvalue(),
                file_name=f"Informe_Tesorería_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                key="download_pdf"
            )

if __name__ == "__main__":
    main()
